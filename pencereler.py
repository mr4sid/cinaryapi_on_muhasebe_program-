# pencereler.py dosyasının içeriği 
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime, date, timedelta
import os
import shutil
import threading
import traceback
import calendar
import multiprocessing
# Üçüncü Parti Kütüphaneler
from PIL import Image, ImageTk
import openpyxl
from veritabani import OnMuhasebe
# Yerel Uygulama Modülleri
from yardimcilar import (sort_treeview_column, setup_numeric_entry, setup_date_entry,
                         validate_numeric_input_generic, format_on_focus_out_numeric_generic,
                         DatePickerDialog)

class SiparisPenceresi(tk.Toplevel):
    def __init__(self, parent, db_manager, app_ref, siparis_tipi, siparis_id_duzenle=None, yenile_callback=None, initial_cari_id=None, initial_urunler=None):
        super().__init__(parent)
        self.app = app_ref
        self.db = db_manager # db_manager'ı burada da tutalım
        self.parent = parent # parent'ı da tutalım
        self.siparis_tipi = siparis_tipi
        self.siparis_id_duzenle = siparis_id_duzenle
        self.yenile_callback = yenile_callback
        self.initial_cari_id = initial_cari_id
        self.initial_urunler = initial_urunler

        title = "Yeni Sipariş"
        if siparis_id_duzenle:
            title = "Sipariş Güncelleme"
        else:
            title = "Yeni Müşteri Siparişi" if siparis_tipi == 'SATIŞ_SIPARIS' else "Yeni Tedarikçi Siparişi"

        self.title(title)
        self.geometry("1400x820") # Boyutu da burada ayarlayın
        self.transient(parent)
        self.grab_set()

        # Yerel içe aktarma
        from arayuz import SiparisOlusturmaSayfasi

        self.siparis_frame = SiparisOlusturmaSayfasi(
            self, # Parent olarak bu Toplevel penceresini veriyoruz.
            self.db,
            self.app,
            self.siparis_tipi,
            siparis_id_duzenle=self.siparis_id_duzenle,
            yenile_callback=self.yenile_callback,
            initial_cari_id=self.initial_cari_id,
            initial_urunler=self.initial_urunler
        )
        self.siparis_frame.pack(expand=True, fill=tk.BOTH)

        # Pencere kapatma olayını yakala ve on_kapat metodunu çağır.
        self.protocol("WM_DELETE_WINDOW", self.on_kapat)

    def on_kapat(self):
        # Sadece yeni bir sipariş oluşturuluyorsa (düzenleme modunda değilse) taslağı kaydet.
        # Düzenleme modunda olan bir sipariş kapatılırsa taslak kaydetmeye gerek yok.
        if self.siparis_id_duzenle is None and self.siparis_frame:
            self.siparis_frame._save_current_form_data_to_temp()
        self.destroy()

class CariHesapEkstresiPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, cari_id, cari_tip, pencere_basligi, parent_list_refresh_func=None):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.cari_id = cari_id
        self.cari_tip = cari_tip
        self.pencere_basligi_str = pencere_basligi
        self.parent_list_refresh_func = parent_list_refresh_func

        self.title(f"Cari Hesap Ekstresi: {self.pencere_basligi_str}")
        self.geometry("1300x850")
        self.transient(parent_app)
        self.grab_set()

        self.app.register_cari_ekstre_window(self)
        self.protocol("WM_DELETE_WINDOW", self.destroy_and_unreg_parent)
        
        main_container = ttk.Frame(self)
        main_container.pack(expand=True, fill=tk.BOTH)
        main_container.rowconfigure(1, weight=1)
        main_container.columnconfigure(0, weight=1)

        # 1. BÖLÜM: Özet ve Bilgi Alanı (Sekmelerin Üstünde)
        self.ozet_ve_bilgi_frame = ttk.LabelFrame(main_container, text="Cari Özet Bilgileri", padding="10")
        self.ozet_ve_bilgi_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        self._create_ozet_bilgi_alani()

        # 2. BÖLÜM: Sekmeli Yapı (Notebook)
        self.notebook = ttk.Notebook(main_container)
        self.notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

        # Sekme 1: Hesap Hareketleri
        self.hesap_hareketleri_tab = ttk.Frame(self.notebook, padding="5")
        self.notebook.add(self.hesap_hareketleri_tab, text="Hesap Hareketleri")
        self._create_hesap_hareketleri_tab()

        # Sekme 2: Siparişler
        self.siparisler_tab = ttk.Frame(self.notebook, padding="5")
        self.notebook.add(self.siparisler_tab, text="Siparişler")
        self._create_siparisler_tab()

        # 3. BÖLÜM: Hızlı İşlem Formu (Sekmelerin Altında)
        self.hizli_islemler_ana_frame = ttk.Frame(main_container)
        self.hizli_islemler_ana_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=5)
        self.hizli_islemler_ana_frame.columnconfigure((0, 1, 2), weight=1)
        self._create_hizli_islem_alanlari()

        self._yukle_ozet_bilgileri()
        self.ekstreyi_yukle()

    def _on_tab_change(self, event):
        selected_tab_text = self.notebook.tab(self.notebook.select(), "text")
        if selected_tab_text == "Siparişler":
            self._siparisleri_yukle()
        elif selected_tab_text == "Hesap Hareketleri":
            self.ekstreyi_yukle()

    def _create_hesap_hareketleri_tab(self):
        parent_frame = self.hesap_hareketleri_tab
        parent_frame.rowconfigure(1, weight=1)
        parent_frame.columnconfigure(0, weight=1)
        
        filter_frame = ttk.Frame(parent_frame, padding="5")
        filter_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=0)
        self._create_filter_alani(filter_frame)

        tree_frame = ttk.Frame(parent_frame, padding="5")
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=0)
        self._create_treeview_alani(tree_frame)

    def _create_siparisler_tab(self):
        """Siparişler sekmesinin içeriğini oluşturur."""
        parent_frame = self.siparisler_tab
        parent_frame.rowconfigure(0, weight=1)
        parent_frame.columnconfigure(0, weight=1)
        
        # Sütunların teknik (kod içi) isimleri
        cols = ("ID", "Sipariş No", "Tarih", "Teslimat Tarihi", "Toplam Tutar", "Durum", "Fatura No", "Not Var", "Oluşturan")
        self.siparisler_tree = ttk.Treeview(parent_frame, columns=cols, show='headings', selectmode="browse")
        
        col_defs = [ 
            ("ID", "ID", 40, tk.CENTER), 
            ("Sipariş No", "Sipariş No", 150, tk.CENTER), 
            ("Tarih", "Sipariş Tarihi", 100, tk.CENTER),
            ("Teslimat Tarihi", "Teslimat Tarihi", 100, tk.CENTER), 
            ("Toplam Tutar", "Toplam Tutar", 120, tk.CENTER), 
            ("Durum", "Durum", 120, tk.CENTER),
            ("Fatura No", "İlişkili Fatura No", 150, tk.CENTER), 
            ("Not Var", "Not Var", 60, tk.CENTER), 
            ("Oluşturan", "Oluşturan Kullanıcı", 100, tk.CENTER) 
        ]
        for col_id, col_text, w, a in col_defs:
            self.siparisler_tree.column(col_id, width=w, anchor=a, stretch=tk.NO)
            self.siparisler_tree.heading(col_id, text=col_text, command=lambda _c=col_id: sort_treeview_column(self.siparisler_tree, _c, False))
            
        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=self.siparisler_tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.siparisler_tree.configure(yscrollcommand=vsb.set)
        self.siparisler_tree.pack(expand=True, fill=tk.BOTH)
        self.siparisler_tree.bind("<Double-1>", self._on_siparis_double_click)
        
        # Renklendirme için etiketler (tags)
        self.siparisler_tree.tag_configure('tamamlandi', foreground='green')
        self.siparisler_tree.tag_configure('iptal_edildi', foreground='gray', font=('Segoe UI', 9, 'overstrike'))

    def _siparisleri_yukle(self):
        for i in self.siparisler_tree.get_children():
            self.siparisler_tree.delete(i)
        
        # veritabanından bu cariye ait siparişleri çek
        siparis_tipi_for_db = 'MUSTERI' if self.cari_tip == 'MUSTERI' else 'TEDARIKCI'
        siparisler = self.db.get_siparisler_by_cari(siparis_tipi_for_db, self.cari_id)
        
        for siparis in siparisler:
            tarih_obj = siparis['tarih']
            teslimat_tarihi_obj = siparis['teslimat_tarihi']

            # Gelen veri zaten bir tarih nesnesi olduğu için doğrudan formatlıyoruz.
            tarih_formatted = tarih_obj.strftime('%d.%m.%Y') if isinstance(tarih_obj, (datetime, date)) else str(tarih_obj or '')
            teslimat_formatted = teslimat_tarihi_obj.strftime('%d.%m.%Y') if isinstance(teslimat_tarihi_obj, (datetime, date)) else '-'
            
            self.siparisler_tree.insert("", "end", iid=siparis['id'], values=(
                siparis['id'], 
                siparis['siparis_no'], 
                tarih_formatted, 
                teslimat_formatted,
                self.db._format_currency(siparis['toplam_tutar']), 
                siparis['durum']
            ))
        self.app.set_status(f"{len(siparisler)} adet sipariş listelendi.")

    def _on_siparis_double_click(self, event):
        selected_item_iid = self.siparisler_tree.focus()
        if not selected_item_iid: return
        
        siparis_id = int(selected_item_iid)
        from pencereler import SiparisDetayPenceresi
        SiparisDetayPenceresi(self.app, self.db, siparis_id)

    def _create_bottom_bar(self):
        self.bottom_frame.columnconfigure(0, weight=1)
        islemler_frame = ttk.Frame(self.bottom_frame)
        islemler_frame.pack(side=tk.LEFT)
        # Butonlar artık burada değil, sağ tık menüsünde. Bu alan boş kalabilir veya başka amaçla kullanılabilir.
        # Şimdilik boş bırakıyoruz.

    def _create_ozet_bilgi_alani(self):
        """Pencerenin üst kısmındaki özet bilgi alanını oluşturur."""
        frame = self.ozet_ve_bilgi_frame
        frame.columnconfigure(1, weight=1)
        frame.columnconfigure(3, weight=1)
        frame.columnconfigure(4, weight=0) # Yeni: Butonlar için sabit sütun

        finans_ozet_cerceve = ttk.Frame(frame)
        finans_ozet_cerceve.grid(row=0, column=0, rowspan=4, sticky="ns", padx=(0, 20))
        label_font_buyuk = ("Segoe UI", 12, "bold"); deger_font_buyuk = ("Segoe UI", 12); label_font_kucuk = ("Segoe UI", 9)
        ttk.Label(finans_ozet_cerceve, text="Net Bakiye:", font=label_font_buyuk).grid(row=0, column=0, sticky="w")
        self.lbl_ozet_net_bakiye = ttk.Label(finans_ozet_cerceve, text="0,00 TL", font=deger_font_buyuk, foreground="blue")
        self.lbl_ozet_net_bakiye.grid(row=0, column=1, sticky="w", padx=5)
        satis_alis_label_text = "Toplam Satış:" if self.cari_tip == 'MUSTERI' else "Toplam Alış:"
        ttk.Label(finans_ozet_cerceve, text=satis_alis_label_text, font=label_font_kucuk).grid(row=1, column=0, sticky="w", pady=(10,0))
        self.lbl_ozet_toplam_satis_alis = ttk.Label(finans_ozet_cerceve, text="0,00 TL", font=label_font_kucuk)
        self.lbl_ozet_toplam_satis_alis.grid(row=1, column=1, sticky="w", padx=5, pady=(10,0))
        tahsilat_odeme_label_text = "Toplam Tahsilat:" if self.cari_tip == 'MUSTERI' else "Toplam Ödeme:"
        ttk.Label(finans_ozet_cerceve, text=tahsilat_odeme_label_text, font=label_font_kucuk).grid(row=2, column=0, sticky="w")
        self.lbl_ozet_toplam_tahsilat_odeme = ttk.Label(finans_ozet_cerceve, text="0,00 TL", font=label_font_kucuk)
        self.lbl_ozet_toplam_tahsilat_odeme.grid(row=2, column=1, sticky="w", padx=5)
        ttk.Label(frame, text="Cari Adı:").grid(row=0, column=2, sticky="w", padx=5)
        self.lbl_cari_detay_ad = ttk.Label(frame, text="-")
        self.lbl_cari_detay_ad.grid(row=0, column=3, sticky="w")
        ttk.Label(frame, text="Telefon:").grid(row=1, column=2, sticky="w", padx=5)
        self.lbl_cari_detay_tel = ttk.Label(frame, text="-")
        self.lbl_cari_detay_tel.grid(row=1, column=3, sticky="w")
        ttk.Label(frame, text="Adres:").grid(row=2, column=2, sticky="w", padx=5)
        self.lbl_cari_detay_adres = ttk.Label(frame, text="-", wraplength=300)
        self.lbl_cari_detay_adres.grid(row=2, column=3, sticky="w")
        ttk.Label(frame, text="Vergi No:").grid(row=3, column=2, sticky="w", padx=5)
        self.lbl_cari_detay_vergi = ttk.Label(frame, text="-")
        self.lbl_cari_detay_vergi.grid(row=3, column=3, sticky="w")

        # Cari Bilgilerini Güncelle Butonunun konumunu güncelliyoruz
        guncelle_btn = ttk.Button(frame, text="Cari Bilgilerini Güncelle", command=self._cari_bilgileri_guncelle)
        guncelle_btn.grid(row=3, column=4, sticky="se", padx=10) # <-- Bu satırı güncelleyin

    def _create_filter_alani(self, filter_frame):
        d = datetime.now()
        baslangic_varsayilan = (d - timedelta(days=365)).strftime('%Y-%m-%d')
        bitis_varsayilan = d.strftime('%Y-%m-%d')
        ttk.Label(filter_frame, text="Başlangıç Tarihi:").pack(side=tk.LEFT, padx=(0, 2))
        self.bas_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.bas_tarih_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.bas_tarih_entry.insert(0, baslangic_varsayilan)
        setup_date_entry(self.app, self.bas_tarih_entry)
        ttk.Button(filter_frame, text="🗓️", command=lambda: DatePickerDialog(self, self.bas_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)
        ttk.Label(filter_frame, text="Bitiş Tarihi:").pack(side=tk.LEFT, padx=(10, 2))
        self.bit_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.bit_tarih_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.bit_tarih_entry.insert(0, bitis_varsayilan)
        setup_date_entry(self.app, self.bit_tarih_entry)
        ttk.Button(filter_frame, text="🗓️", command=lambda: DatePickerDialog(self, self.bit_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)
        ttk.Button(filter_frame, text="Filtrele", command=self.ekstreyi_yukle, style="Accent.TButton").pack(side=tk.LEFT, padx=10)
        
        # Özet alanına taşıdığımız butonlar için yeni bir çerçeve
        export_buttons_frame = ttk.Frame(self.ozet_ve_bilgi_frame) # self.ozet_ve_bilgi_frame'in içine yerleştiriyoruz
        export_buttons_frame.grid(row=0, column=4, rowspan=2, sticky="ne", padx=10) # Konumunu ayarlayın

        ttk.Button(export_buttons_frame, text="PDF'e Aktar", command=self.pdf_aktar).pack(pady=2, fill=tk.X)
        ttk.Button(export_buttons_frame, text="Excel'e Aktar", command=self.excel_aktar).pack(pady=2, fill=tk.X)

    def _create_treeview_alani(self, tree_frame):
        cols = ("ID", "Tarih", "Saat", "İşlem Tipi", "Referans", "Ödeme Türü", "Açıklama/Detay", "Borç", "Alacak", "Bakiye")
        self.ekstre_tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode="browse")
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.ekstre_tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.ekstre_tree.configure(yscrollcommand=vsb.set)
        self.ekstre_tree.pack(expand=True, fill=tk.BOTH)
        
        self.ekstre_tree.bind("<Button-3>", self._show_context_menu)
        self.ekstre_tree.bind("<Double-1>", self.on_double_click_hareket_detay)
        
        col_defs = [
            ("ID", 40, tk.CENTER, tk.NO), ("Tarih", 80, tk.CENTER, tk.NO),
            ("Saat", 60, tk.CENTER, tk.NO), ("İşlem Tipi", 120, tk.CENTER, tk.NO),
            ("Referans", 120, tk.CENTER, tk.NO), ("Ödeme Türü", 100, tk.CENTER, tk.NO),
            ("Açıklama/Detay", 300, tk.W, tk.YES), # Açıklama sola yaslı kalsın
            ("Borç", 100, tk.CENTER, tk.NO),
            ("Alacak", 100, tk.CENTER, tk.NO), ("Bakiye", 120, tk.CENTER, tk.NO)
        ]
        for cn, w, a, s in col_defs:
            self.ekstre_tree.column(cn, width=w, anchor=a, stretch=s)
            self.ekstre_tree.heading(cn, text=cn, command=lambda _c=cn: sort_treeview_column(self.ekstre_tree, _c, False))
        
        self.ekstre_tree.tag_configure('devir', font=('Segoe UI', 9, 'bold'), background='#EFEFEF')
        self.ekstre_tree.tag_configure('acik_hesap', foreground='red')
        self.ekstre_tree.tag_configure('tahsilat_odeme', foreground='green')
        self.ekstre_tree.tag_configure('pesin_islem', foreground='gray', font=('Segoe UI', 9, 'italic'))


    def _create_hizli_islem_alanlari(self):
        # Ödeme/Tahsilat Formu
        odeme_tahsilat_frame = ttk.LabelFrame(self.hizli_islemler_ana_frame, text="Ödeme Ekle" if self.cari_tip == 'TEDARIKCI' else "Tahsilat Ekle", padding="10")
        odeme_tahsilat_frame.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Label(odeme_tahsilat_frame, text="Ödeme Tipi:").pack(anchor=tk.W)
        self.ot_odeme_tipi_combo = ttk.Combobox(odeme_tahsilat_frame, state="readonly", values=["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"])
        self.ot_odeme_tipi_combo.pack(fill=tk.X, pady=2)
        self.ot_odeme_tipi_combo.set("NAKİT")
        ttk.Label(odeme_tahsilat_frame, text="Tutar:").pack(anchor=tk.W)
        self.ot_tutar_entry = ttk.Entry(odeme_tahsilat_frame)
        self.ot_tutar_entry.pack(fill=tk.X, pady=2)
        setup_numeric_entry(self.app, self.ot_tutar_entry)
        ttk.Label(odeme_tahsilat_frame, text="Not:").pack(anchor=tk.W)
        self.ot_not_entry = ttk.Entry(odeme_tahsilat_frame)
        self.ot_not_entry.pack(fill=tk.X, pady=2)
        ot_kaydet_btn_text = "Ödeme Ekle" if self.cari_tip == 'TEDARIKCI' else "Tahsilat Ekle"
        ttk.Button(odeme_tahsilat_frame, text=ot_kaydet_btn_text, command=self._hizli_odeme_tahsilat_kaydet, style="Accent.TButton").pack(fill=tk.X, pady=(5,0))

        # Veresiye Borç Formu
        borc_frame = ttk.LabelFrame(self.hizli_islemler_ana_frame, text="Veresiye Borç Ekle", padding="10")
        borc_frame.grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Label(borc_frame, text="Türü Seçiniz:").pack(anchor=tk.W)
        self.borc_tur_combo = ttk.Combobox(borc_frame, state="readonly", values=["Satış Faturası", "Diğer Borç"])
        self.borc_tur_combo.pack(fill=tk.X, pady=2)
        self.borc_tur_combo.set("Diğer Borç")
        ttk.Label(borc_frame, text="Tutar:").pack(anchor=tk.W)
        self.borc_tutar_entry = ttk.Entry(borc_frame)
        self.borc_tutar_entry.pack(fill=tk.X, pady=2)
        setup_numeric_entry(self.app, self.borc_tutar_entry)
        ttk.Label(borc_frame, text="Not:").pack(anchor=tk.W)
        self.borc_not_entry = ttk.Entry(borc_frame)
        self.borc_not_entry.pack(fill=tk.X, pady=2)
        ttk.Button(borc_frame, text="Veresiye Ekle", command=self._hizli_veresiye_borc_kaydet, style="Accent.TButton").pack(fill=tk.X, pady=(5,0))

        # Alacak Ekleme Formu
        alacak_frame = ttk.LabelFrame(self.hizli_islemler_ana_frame, text="Alacak Ekleme", padding="10")
        alacak_frame.grid(row=0, column=2, sticky="ew", padx=(5, 0))
        ttk.Label(alacak_frame, text="Türü Seçiniz:").pack(anchor=tk.W)
        self.alacak_tur_combo = ttk.Combobox(alacak_frame, state="readonly", values=["İade Faturası", "Diğer Alacak"])
        self.alacak_tur_combo.pack(fill=tk.X, pady=2)
        self.alacak_tur_combo.set("Diğer Alacak")
        ttk.Label(alacak_frame, text="Tutar:").pack(anchor=tk.W)
        self.alacak_tutar_entry = ttk.Entry(alacak_frame)
        self.alacak_tutar_entry.pack(fill=tk.X, pady=2)
        setup_numeric_entry(self.app, self.alacak_tutar_entry)
        ttk.Label(alacak_frame, text="Not:").pack(anchor=tk.W)
        self.alacak_not_entry = ttk.Entry(alacak_frame)
        self.alacak_not_entry.pack(fill=tk.X, pady=2)
        ttk.Button(alacak_frame, text="Alacak Kaydet", command=self._hizli_alacak_kaydet, style="Accent.TButton").pack(fill=tk.X, pady=(5,0))

    def _show_context_menu(self, event):
        """Treeview üzerinde sağ tıklandığında bağlamsal menüyü oluşturur ve gösterir."""
        # Sağ tıklanan öğeyi seçili hale getir
        item_id = self.ekstre_tree.identify_row(event.y)
        if not item_id:
            return
        
        self.ekstre_tree.selection_set(item_id)
        
        tags = self.ekstre_tree.item(item_id, "tags")
        
        # Menüyü oluştur
        context_menu = tk.Menu(self, tearoff=0)
        
        can_delete = False
        can_update = False
        
        if tags and len(tags) >= 3:
            ref_tip = tags[2]
            if ref_tip in ['TAHSILAT', 'ODEME', 'VERESIYE_BORC_MANUEL', 'FATURA']:
                can_delete = True
            if ref_tip == 'FATURA':
                can_update = True
        
        if can_delete:
            context_menu.add_command(label="İşlemi Sil", command=self.secili_islemi_sil)
        
        if can_update:
            context_menu.add_command(label="Faturayı Güncelle", command=self.secili_islemi_guncelle)
            
        # Eğer menüde hiç öğe yoksa gösterme
        if can_delete or can_update:
            context_menu.tk_popup(event.x_root, event.y_root)


    def _yukle_ozet_bilgileri(self):
        ozet_data = self.db.get_cari_ozet_bilgileri(self.cari_id, self.cari_tip)
        cari_data = self.db.musteri_getir_by_id(self.cari_id) if self.cari_tip == 'MUSTERI' else self.db.tedarikci_getir_by_id(self.cari_id)

        net_bakiye = ozet_data.get("net_bakiye", 0.0)
        self.lbl_ozet_net_bakiye.config(text=self.db._format_currency(abs(net_bakiye)))
        if net_bakiye > 0: self.lbl_ozet_net_bakiye.config(foreground="red")
        elif net_bakiye < 0: self.lbl_ozet_net_bakiye.config(foreground="green")
        else: self.lbl_ozet_net_bakiye.config(foreground="black")

        satis_alis_tutar = ozet_data.get("toplam_satis") if self.cari_tip == 'MUSTERI' else ozet_data.get("toplam_alis")
        self.lbl_ozet_toplam_satis_alis.config(text=self.db._format_currency(satis_alis_tutar))

        tahsilat_odeme_tutar = ozet_data.get("toplam_tahsilat") if self.cari_tip == 'MUSTERI' else ozet_data.get("toplam_odeme")
        self.lbl_ozet_toplam_tahsilat_odeme.config(text=self.db._format_currency(tahsilat_odeme_tutar))

        if cari_data:
            self.lbl_cari_detay_ad.config(text=cari_data['ad'])
            self.lbl_cari_detay_tel.config(text=cari_data['telefon'] or "-")
            self.lbl_cari_detay_adres.config(text=cari_data['adres'] or "-")
            self.lbl_cari_detay_vergi.config(text=f"{cari_data['vergi_dairesi'] or '-'} / {cari_data['vergi_no'] or '-'}")


    def _cari_bilgileri_guncelle(self):
        """Cari düzenleme penceresini açar."""
        if self.cari_tip == 'MUSTERI':
            musteri_db = self.db.musteri_getir_by_id(self.cari_id)
            if musteri_db:
                from pencereler import YeniMusteriEklePenceresi, YeniTedarikciEklePenceresi
                YeniMusteriEklePenceresi(self.app, self.db, self._ozet_ve_liste_yenile, musteri_duzenle=musteri_db, app_ref=self.app)
        elif self.cari_tip == 'TEDARIKCI':
            tedarikci_db = self.db.tedarikci_getir_by_id(self.cari_id)
            if tedarikci_db:
                YeniTedarikciEklePenceresi(self.app, self.db, self._ozet_ve_liste_yenile, tedarikci_duzenle=tedarikci_db, app_ref=self.app)

    def _ozet_ve_liste_yenile(self):
        self._yukle_ozet_bilgileri()
        self.ekstreyi_yukle()
        if self.parent_list_refresh_func:
            self.parent_list_refresh_func()

    def _yukle_kasa_banka_hesaplarini_quick(self):
        self.quick_kasa_banka_map = {}
        display_values = [""]
        hesaplar = self.db.kasa_banka_listesi_al()
        if hesaplar:
            for h_id, h_ad, h_no, h_bakiye, h_para_birimi, h_tip, h_acilis_tarihi, h_banka, h_sube_adi, h_varsayilan_odeme_turu in hesaplar:
                bakiye_formatted = self.db._format_currency(h_bakiye)
                display_text = f"{h_ad} ({h_tip}) - Bakiye: {bakiye_formatted}"
                if h_tip == "BANKA" and h_banka:
                    display_text += f" ({h_banka})"
                self.quick_kasa_banka_map[display_text] = h_id
                display_values.append(display_text)
            
            self.quick_kasa_banka_combo['values'] = display_values
            if len(display_values) > 1:
                self.quick_kasa_banka_combo.current(1)
                self.quick_kasa_banka_combo.config(state="readonly")
            else:
                self.quick_kasa_banka_combo.current(0)
                self.quick_kasa_banka_combo.config(state=tk.DISABLED)
        else:
            self.quick_kasa_banka_combo['values'] = ["Hesap Yok"]
            self.quick_kasa_banka_combo.set("Hesap Yok")
            self.quick_kasa_banka_combo.config(state=tk.DISABLED)

    def _hizli_odeme_tahsilat_kaydet(self):
        odeme_tipi = self.ot_odeme_tipi_combo.get()
        tutar_str = self.ot_tutar_entry.get()
        not_str = self.ot_not_entry.get() or f"Manuel {self.cari_tip.capitalize()} {odeme_tipi}"

        if not tutar_str:
            messagebox.showwarning("Eksik Bilgi", "Lütfen tutar giriniz.", parent=self)
            return

        # Basitlik adına, tüm peşin işlemleri MERKEZİ NAKİT kasasına işliyoruz.
        # Daha sonra burası da bir combobox ile seçilebilir hale getirilebilir.
        kasa_id = self.db.default_nakit_kasa_id

        if self.cari_tip == 'MUSTERI':
            success, message = self.db.tahsilat_ekle(self.cari_id, datetime.now().strftime('%Y-%m-%d'), tutar_str, odeme_tipi, not_str, kasa_id)
        else: # TEDARIKCI
            success, message = self.db.odeme_ekle(self.cari_id, datetime.now().strftime('%Y-%m-%d'), tutar_str, odeme_tipi, not_str, kasa_id)
        
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self._ozet_ve_liste_yenile()
            self.ot_tutar_entry.delete(0, tk.END)
            self.ot_not_entry.delete(0, tk.END)
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _hizli_veresiye_borc_kaydet(self):
        tur = self.borc_tur_combo.get()
        tutar_str = self.borc_tutar_entry.get()
        not_str = self.borc_not_entry.get() or f"Manuel {tur}"

        if not tutar_str:
            messagebox.showwarning("Eksik Bilgi", "Lütfen tutar giriniz.", parent=self)
            return

        if tur == "Satış Faturası":
            messagebox.showinfo("Yönlendirme", "Fatura oluşturmak için lütfen ana menüden 'Yeni Satış Faturası' ekranını kullanın.", parent=self)
        else: # Diğer Borç
            success, message = self.db.veresiye_borc_ekle(self.cari_id, self.cari_tip, datetime.now().strftime('%Y-%m-%d'), tutar_str, not_str)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self._ozet_ve_liste_yenile()
                self.borc_tutar_entry.delete(0, tk.END)
                self.borc_not_entry.delete(0, tk.END)
            else:
                messagebox.showerror("Hata", message, parent=self)

    def _hizli_alacak_kaydet(self):
        # Bu özellik henüz tam olarak kodlanmamıştır. Örnek bir uyarı gösterir.
        messagebox.showinfo("Geliştirme Aşamasında", "Alacak ekleme özelliği henüz tamamlanmamıştır.", parent=self)

    def destroy_and_unreg_parent(self):
        self.app.unregister_cari_ekstre_window(self)
        if self.parent_list_refresh_func: self.parent_list_refresh_func()
        self.destroy()

    def excel_aktar(self):
        dosya_yolu = filedialog.asksaveasfilename(
            initialfile=f"Cari_Ekstresi_{self.pencere_basligi_str.replace(' ', '_').replace('(', '_').replace(')', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx")],
            title="Cari Hesap Ekstresini Excel'e Kaydet",
            parent=self
        )
        if dosya_yolu:
            bekleme_penceresi = BeklemePenceresi(self, message="Ekstre Excel'e aktarılıyor, lütfen bekleyiniz...")
            # Ayrı thread'de işlemi başlat
            threading.Thread(target=lambda: self._generate_ekstre_excel_threaded(
                self.cari_tip, self.cari_id, self.bas_tarih_entry.get(), self.bit_tarih_entry.get(),
                dosya_yolu, bekleme_penceresi
            )).start()
        else:
            self.app.set_status("Excel'e aktarma iptal edildi.")

    def pdf_aktar(self):
        dosya_yolu = filedialog.asksaveasfilename(
            initialfile=f"Cari_Ekstresi_{self.pencere_basligi_str.replace(' ', '_').replace('(', '_').replace(')', '')}_{datetime.now().strftime('%Y%m%d')}.pdf",
            defaultextension=".pdf",
            filetypes=[("PDF Dosyaları", "*.pdf")],
            title="Cari Hesap Ekstresini PDF'e Kaydet",
            parent=self
        )
        if dosya_yolu:
            bekleme_penceresi = BeklemePenceresi(self, message="Ekstre PDF'e aktarılıyor, lütfen bekleyiniz...")

            # Sonuçları ana sürece geri bildirmek için bir Kuyruk oluştur
            self.result_queue = multiprocessing.Queue()
            from main import _pdf_olusturma_islemi 

            # PDF oluşturma işlemini ayrı bir süreçte başlat
            self.pdf_process = multiprocessing.Process(target=_pdf_olusturma_islemi, args=(
                self.db.db_name, # Veritabanı dosya yolu
                self.cari_tip,
                self.cari_id,
                self.bas_tarih_entry.get(),
                self.bit_tarih_entry.get(),
                dosya_yolu,
                self.result_queue # Sonuç kuyruğu
            ))
            self.pdf_process.start() # Süreci başlat

            self.after(100, self._check_pdf_process_completion, bekleme_penceresi)
        else:
            self.app.set_status("PDF'e aktarma iptal edildi.")

    def _check_pdf_process_completion(self, bekleme_penceresi):
        # Süreç hala çalışıyor mu kontrol et
        if self.pdf_process.is_alive():
            self.after(100, self._check_pdf_process_completion, bekleme_penceresi) # 100ms sonra tekrar kontrol et
        else:
            # Süreç tamamlandı, sonuçları al
            bekleme_penceresi.kapat()
            success, message = self.result_queue.get() # Kuyruktan sonucu al
            
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.app.set_status(message)
            else:
                messagebox.showerror("Hata", message, parent=self)
                self.app.set_status(f"Ekstre PDF'e aktarılırken hata: {message}")
            
            self.pdf_process.join() # Sürecin tamamen bitmesini bekle (kaynakları serbest bırakmak için)

    def _generate_ekstre_excel_threaded(self, cari_tip, cari_id, bas_t, bit_t, dosya_yolu, bekleme_penceresi):
        success, message = self.db.cari_ekstresi_excel_olustur(cari_tip, cari_id, bas_t, bit_t, dosya_yolu)
        self.app.after(0, bekleme_penceresi.kapat)
        if success:
            self.app.after(0, lambda: messagebox.showinfo("Başarılı", message, parent=self))
            self.app.after(0, lambda: self.app.set_status(message))
        else:
            self.app.after(0, lambda: messagebox.showerror("Hata", message, parent=self))
            self.app.after(0, lambda: self.app.set_status(f"Ekstre Excel'e aktarılırken hata: {message}"))            

    def _generate_ekstre_pdf_threaded(self, cari_tip, cari_id, bas_t, bit_t, dosya_yolu, bekleme_penceresi):
        success = False # Başlangıçta başarısız olarak ayarla
        message = "Bilinmeyen bir hata oluştu." # Başlangıç hata mesajı

        try:
            # Tüm PDF oluşturma mantığı veritabanı sınıfında olmalı
            success, message = self.db.cari_ekstresi_pdf_olustur(cari_tip, cari_id, bas_t, bit_t, dosya_yolu)
        except Exception as e:
            # Thread içinde oluşan hatayı yakala ve mesajı ayarla
            message = f"PDF oluşturma sırasında beklenmeyen bir hata oluştu: {e}\n{traceback.format_exc()}"
            from arayuz import logging
            logging.error(f"Cari Ekstresi PDF thread hatası: {message}") # Loglama ekle

        finally:
            # UI güncellemelerini her zaman ana thread'e geri gönder
            self.app.after(0, bekleme_penceresi.kapat)
            if success:
                self.app.after(0, lambda: messagebox.showinfo("Başarılı", message, parent=self))
                self.app.after(0, lambda: self.app.set_status(message))
            else:
                self.app.after(0, lambda: messagebox.showerror("Hata", message, parent=self))
                self.app.after(0, lambda: self.app.set_status(f"Ekstre PDF'e aktarılırken hata: {message}"))

    def ekstreyi_yukle(self):
        self._yukle_ozet_bilgileri()
        
        for i in self.ekstre_tree.get_children():
            self.ekstre_tree.delete(i)
        
        bas_t = self.bas_tarih_entry.get()
        bit_t = self.bit_tarih_entry.get()

        hareketler_listesi, devreden_bakiye, _, _ = self.db.cari_hesap_ekstresi_al(
            self.cari_tip, self.cari_id, bas_t, bit_t
        )
        
        self.ekstre_tree.insert("", "end", values=("", "", "", "DEVİR", "", "", "", "", self.db._format_currency(devreden_bakiye)), tags=('devir',))

        bakiye = devreden_bakiye
        for hareket in hareketler_listesi:
            # <<< NİHAİ GÖSTERİM MANTIĞI BURADA >>>
            
            # 1. KURAL: Peşin bir faturanın otomatik ödeme/tahsilat kaydını atla.
            if hareket['referans_tip'] in ['FATURA_SATIS_PESIN', 'FATURA_ALIS_PESIN']:
                continue 
            
            tutar = hareket['tutar']
            islem_tipi_db = hareket['islem_tipi']
            ref_tip = hareket['referans_tip']
            odeme_turu = hareket['odeme_turu'] or ''
            
            borc_str, alacak_str = "", ""
            tags_list = []

            # 2. KURAL: Bakiye hesaplama ve renklendirme
            if ref_tip == 'FATURA' and odeme_turu in self.db.pesin_odeme_turleri:
                # Peşin faturalar bakiyeyi etkilemez, rengi gri ve italik olur.
                tags_list.append('pesin_islem')
            elif islem_tipi_db in ['TAHSILAT', 'ODEME']:
                # Manuel tahsilat/ödemeler bakiyeyi etkiler, rengi yeşil olur.
                if self.cari_tip == 'MUSTERI': bakiye -= tutar
                else: bakiye += tutar
                tags_list.append('tahsilat_odeme')
            else: # Bu, AÇIK HESAP fatura veya MANUEL borç demektir.
                # Veresiye işlemler bakiyeyi etkiler, rengi kırmızı olur.
                if self.cari_tip == 'MUSTERI': bakiye += tutar
                else: bakiye -= tutar
                tags_list.append('acik_hesap')

            # 3. Sütunları doldurma
            if self.cari_tip == 'MUSTERI':
                if islem_tipi_db == 'ALACAK': alacak_str = self.db._format_currency(tutar)
                elif islem_tipi_db == 'TAHSILAT': borc_str = self.db._format_currency(tutar)
            elif self.cari_tip == 'TEDARIKCI':
                if islem_tipi_db == 'BORC': alacak_str = self.db._format_currency(tutar)
                elif islem_tipi_db == 'ODEME': borc_str = self.db._format_currency(tutar)
            
            try: formatted_date = datetime.strptime(str(hareket['tarih']), '%Y-%m-%d').strftime('%d.%m.%Y')
            except: formatted_date = hareket['tarih'] or ''
            
            ref_gosterim = hareket['fatura_no'] if ref_tip == 'FATURA' else (ref_tip or '-')
            tags_list.extend([str(hareket['referans_id']), str(ref_tip)])

            self.ekstre_tree.insert("", "end", iid=hareket['id'], values=(
                hareket['id'], formatted_date, hareket['islem_saati'] or '', 
                islem_tipi_db, ref_gosterim, odeme_turu, 
                hareket['aciklama'] or '', borc_str, alacak_str,
                self.db._format_currency(bakiye)
            ), tags=tuple(tags_list))

    def on_tree_select(self, event):
        selected_item_iid = self.ekstre_tree.focus()
        can_delete = False
        can_update_fatura = False

        if selected_item_iid:
            tags = self.ekstre_tree.item(selected_item_iid, "tags")
            if tags and len(tags) > 2:
                ref_tip = tags[2]
                if ref_tip in ['TAHSILAT', 'ODEME', 'VERESIYE_BORC_MANUEL']:
                    can_delete = True
                elif ref_tip == 'FATURA':
                    can_delete = True 
                    can_update_fatura = True 
        
        self.sil_button_bottom.config(state=tk.NORMAL if can_delete else tk.DISABLED)
        self.guncelle_button_bottom.config(state=tk.NORMAL if can_update_fatura else tk.DISABLED)


    def secili_islemi_sil(self):
        selected_item_iid = self.ekstre_tree.focus()
        if not selected_item_iid:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir işlem seçin.", parent=self)
            return

        tags = self.ekstre_tree.item(selected_item_iid, "tags")
        aciklama_text = self.ekstre_tree.item(selected_item_iid, "values")[5]

        if tags and len(tags) >= 3:
            ref_id_str, ref_tip = tags[1], tags[2]
            cari_hareket_id = int(selected_item_iid)

            confirm_msg = f"'{aciklama_text}' açıklamalı işlemi silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz."
            if ref_tip == 'FATURA':
                ref_no = self.ekstre_tree.item(selected_item_iid, "values")[4]
                confirm_msg = f"'{ref_no}' referanslı FATURA kaydını ve ilişkili tüm hareketleri silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz."
            
            if messagebox.askyesno("Silme Onayı", confirm_msg, icon='warning', parent=self):
                success, message = (False, "Bilinmeyen işlem tipi.")
                if ref_tip == 'FATURA':
                    fatura_id_to_delete = int(ref_id_str)
                    success, message = self.db.fatura_sil(fatura_id_to_delete)
                else:
                    success, message = self.db.tahsilat_odeme_sil(cari_hareket_id)

                if success:
                    messagebox.showinfo("Başarılı", message, parent=self)
                    self.ekstreyi_yukle()
                    if self.parent_list_refresh_func: self.parent_list_refresh_func()
                else:
                    messagebox.showerror("Hata", message, parent=self)
        else:
            messagebox.showerror("Hata", "İşlem tipi belirlenemedi, silinemiyor.", parent=self)

    def secili_islemi_guncelle(self):
        selected_item_iid = self.ekstre_tree.focus()
        if not selected_item_iid:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir fatura işlemi seçin.", parent=self)
            return

        tags = self.ekstre_tree.item(selected_item_iid, "tags")
        if tags and len(tags) >= 3 and tags[2] == 'FATURA':
            fatura_id_to_update = int(tags[1])
            from pencereler import FaturaGuncellemePenceresi
            FaturaGuncellemePenceresi(self.app, self.db, fatura_id_to_update, self.ekstreyi_yukle)
        else:
            messagebox.showinfo("Bilgi", "Sadece fatura işlemleri güncellenebilir.", parent=self)


    def on_double_click_hareket_detay(self, event):
        selected_item_iid = self.ekstre_tree.focus()
        if not selected_item_iid: return

        tags = self.ekstre_tree.item(selected_item_iid, "tags")
        item_values = self.ekstre_tree.item(selected_item_iid, "values") # Diğer bilgilere erişmek için

        if tags and len(tags) >= 3:
            # tags'in formatı: ('tag_name', 'referans_id_str', 'referans_tip_str')
            ref_id_str = tags[1]
            ref_tip_str = tags[2] # Örneğin 'FATURA', 'TAHSILAT', 'ODEME', 'VERESIYE_BORC_MANUEL'

            try:
                ref_id = int(ref_id_str)
            except ValueError:
                messagebox.showerror("Hata", "Referans ID okunamadı.", parent=self)
                return

            if ref_tip_str == 'FATURA':
                # Fatura Detay Penceresini Aç
                from pencereler import FaturaDetayPenceresi
                FaturaDetayPenceresi(self.app, self.db, ref_id)
            elif ref_tip_str in ['TAHSILAT', 'ODEME', 'VERESIYE_BORC_MANUEL']:
                # Manuel Tahsilat/Ödeme/Veresiye Borç için detay gösterimi
                # Şu anda bu işlemler için ayrı bir detay penceresi yok.
                # Basit bir bilgi kutusu gösterebilir veya gelecekte detay penceresi geliştirebilirsiniz.
                messagebox.showinfo("İşlem Detayı", 
                                    f"Bu bir {ref_tip_str} işlemidir.\n"
                                    f"Tarih: {item_values[1]}\n"
                                    f"Tutar: {item_values[7] if item_values[7] else item_values[8]}\n" # Borç veya Alacak
                                    f"Açıklama: {item_values[6]}\n"
                                    f"Referans ID: {ref_id}",
                                    parent=self)
            else:
                messagebox.showinfo("Detay", "Bu işlem tipi için detay görüntüleme mevcut değil.", parent=self)
        else:
            messagebox.showinfo("Detay", "Bu satırda detay görüntülenecek bir referans yok.", parent=self)


class FaturaGuncellemePenceresi(tk.Toplevel):
    def __init__(self, parent, db_manager, fatura_id_duzenle, yenile_callback_liste=None):
        super().__init__(parent)
        self.app = parent.app
        self.db = db_manager
        self.yenile_callback_liste = yenile_callback_liste
        self.fatura_id_duzenle = fatura_id_duzenle # Bu ID'yi de saklayalım

        fatura_ana_bilgileri = self.db.fatura_getir_by_id(fatura_id_duzenle)
        if not fatura_ana_bilgileri:
            messagebox.showerror("Hata", "Güncellenecek fatura bilgileri bulunamadı.", parent=self)
            self.destroy()
            return

        fatura_tipi = fatura_ana_bilgileri['tip']

        self.title(f"Fatura Güncelleme: {fatura_ana_bilgileri['fatura_no']}")
        self.geometry("1400x820")
        self.transient(parent)
        self.grab_set()

        # Yerel içe aktarma
        from arayuz import FaturaOlusturmaSayfasi

        self.fatura_olusturma_frame = FaturaOlusturmaSayfasi(
            self,
            self.db,
            self.app,
            fatura_tipi,
            duzenleme_id=fatura_id_duzenle, # Doğru parametre adı
            yenile_callback=self._fatura_guncellendi_callback
        )
        self.fatura_olusturma_frame.pack(expand=True, fill=tk.BOTH)

        # Pencere kapatma olayını yakala ve on_kapat metodunu çağır.
        self.protocol("WM_DELETE_WINDOW", self.on_kapat)

    def on_kapat(self):
        """Pencere kapatıldığında çağrılır."""
        if self.yenile_callback_liste:
            self.yenile_callback_liste() # Fatura listesini yenile
        self.destroy()

    def _fatura_guncellendi_callback(self):
        if self.yenile_callback_liste:
            self.yenile_callback_liste()
        self.destroy()

    def destroy_and_callback(self):
        if self.yenile_callback_liste:
            self.yenile_callback_liste()
        self.destroy()

class IadeFaturasiPenceresi(tk.Toplevel):
    def __init__(self, parent, db_instance: OnMuhasebe, original_invoice_id=None, invoice_type=None, current_user_id=None):
        super().__init__(parent)
        self.parent = parent
        self.db = db_instance # db_manager yerine 'db' kullanıyoruz
        self.original_invoice_id = original_invoice_id
        self.invoice_type = invoice_type if invoice_type else 'SATIŞ_İADE' # Varsayılan
        self.current_user_id = current_user_id

        self.cari_id = None
        self.cari_tip = None
        self.urun_kalemleri = []

        self.title("İade Faturası Oluştur")
        self.geometry("900x700")
        self.grab_set()
        self.focus_set()

        self._create_widgets()
        self._load_initial_data()

    def _create_widgets(self):
        general_frame = ttk.LabelFrame(self, text="İade Faturası Bilgileri", padding=10)
        general_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(general_frame, text="İade Fatura No:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.invoice_no_entry = ttk.Entry(general_frame)
        self.invoice_no_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(general_frame, text="İade Tarihi:").grid(row=0, column=2, padx=5, pady=2, sticky="w")
        self.date_entry = DateEntry(general_frame, width=10, date_picker_dialog=DatePickerDialog)
        self.date_entry.grid(row=0, column=3, padx=5, pady=2, sticky="ew")
        self.date_entry.set_date(datetime.date.today())

        ttk.Label(general_frame, text="Cari Hesap:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        self.cari_adi_var = tk.StringVar(value="Cari Hesap Seçiniz...")
        self.cari_hesap_label = ttk.Label(general_frame, textvariable=self.cari_adi_var, foreground="blue")
        self.cari_hesap_label.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        ttk.Button(general_frame, text="Seç", command=self._select_cari_hesap).grid(row=1, column=2, padx=5, pady=2, sticky="w")

        if self.original_invoice_id:
            ttk.Label(general_frame, text="Orijinal Fatura ID:").grid(row=2, column=0, padx=5, pady=2, sticky="w")
            ttk.Label(general_frame, text=str(self.original_invoice_id)).grid(row=2, column=1, padx=5, pady=2, sticky="w")

        ttk.Label(general_frame, text="Açıklama:").grid(row=3, column=0, padx=5, pady=2, sticky="w")
        self.description_entry = ttk.Entry(general_frame)
        self.description_entry.grid(row=3, column=1, columnspan=3, padx=5, pady=2, sticky="ew")

        general_frame.grid_columnconfigure(1, weight=1)

        items_frame = ttk.LabelFrame(self, text="İade Edilen Ürünler", padding=10)
        items_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(items_frame, columns=("urun_id", "kod", "ad", "miktar", "birim_fiyat", "kdv_oran", "tutar"), show="headings")
        self.tree.heading("kod", text="Ürün Kodu", command=lambda: sort_treeview_column(self.tree, "kod", False))
        self.tree.heading("ad", text="Ürün Adı", command=lambda: sort_treeview_column(self.tree, "ad", False))
        self.tree.heading("miktar", text="Miktar", command=lambda: sort_treeview_column(self.tree, "miktar", True))
        self.tree.heading("birim_fiyat", text="Birim Fiyat", command=lambda: sort_treeview_column(self.tree, "birim_fiyat", True))
        self.tree.heading("kdv_oran", text="KDV Oranı (%)", command=lambda: sort_treeview_column(self.tree, "kdv_oran", True))
        self.tree.heading("tutar", text="Tutar", command=lambda: sort_treeview_column(self.tree, "tutar", True))

        self.tree.column("urun_id", width=0, stretch=tk.NO)
        self.tree.column("kod", width=80, anchor="w")
        self.tree.column("ad", width=200, anchor="w")
        self.tree.column("miktar", width=80, anchor="e")
        self.tree.column("birim_fiyat", width=100, anchor="e")
        self.tree.column("kdv_oran", width=80, anchor="e")
        self.tree.column("tutar", width=100, anchor="e")

        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<Double-1>", self._on_tree_double_click)

        item_buttons_frame = ttk.Frame(items_frame)
        item_buttons_frame.pack(fill=tk.X, pady=5)
        ttk.Button(item_buttons_frame, text="Ürün Ekle", command=self._add_product).pack(side=tk.LEFT, padx=5)
        ttk.Button(item_buttons_frame, text="Seçili Ürünü Sil", command=self._delete_selected_product).pack(side=tk.LEFT, padx=5)

        totals_frame = ttk.LabelFrame(self, text="Toplamlar", padding=10)
        totals_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(totals_frame, text="Ara Toplam (KDV Hariç):").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.subtotal_var = tk.DoubleVar(value=0.0)
        ttk.Label(totals_frame, textvariable=self.subtotal_var, font=("TkDefaultFont", 10, "bold")).grid(row=0, column=1, padx=5, pady=2, sticky="e")

        ttk.Label(totals_frame, text="Toplam KDV:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        self.vat_total_var = tk.DoubleVar(value=0.0)
        ttk.Label(totals_frame, textvariable=self.vat_total_var, font=("TkDefaultFont", 10, "bold")).grid(row=1, column=1, padx=5, pady=2, sticky="e")

        ttk.Label(totals_frame, text="Genel Toplam (KDV Dahil):").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        self.grand_total_var = tk.DoubleVar(value=0.0)
        ttk.Label(totals_frame, textvariable=self.grand_total_var, font=("TkDefaultFont", 12, "bold")).grid(row=2, column=1, padx=5, pady=2, sticky="e")

        totals_frame.grid_columnconfigure(1, weight=1)

        button_frame = ttk.Frame(self)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(button_frame, text="İade Faturasını Kaydet", command=self._save_return_invoice).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.RIGHT, padx=5)

    def _load_initial_data(self):
        """Orijinal faturadan veri yükler veya boş formu hazırlar."""
        self.invoice_no_entry.insert(0, self.db.generate_new_return_invoice_no())

        if self.original_invoice_id:
            try:
                original_invoice = self.db.get_invoice_details(self.original_invoice_id)
                original_items = self.db.get_invoice_items(self.original_invoice_id)

                if original_invoice:
                    self.cari_id = original_invoice['cari_id']
                    self.cari_tip = original_invoice['cari_tip']
                    self.cari_adi_var.set(f"{original_invoice['cari_adi']}")
                    self.description_entry.insert(0, f"Orj. Fatura No: {original_invoice['fatura_no']} İadesi")
                    
                    if original_invoice['tip'] == 'SATIŞ':
                        self.invoice_type = 'SATIŞ_İADE'
                    elif original_invoice['tip'] == 'ALIŞ':
                        self.invoice_type = 'ALIŞ_İADE'
                    else:
                        messagebox.showwarning("Uyarı", "Orijinal fatura tipi iade için uygun değil. Varsayılan satış iadesi olarak ayarlandı.")
                        self.invoice_type = 'SATIŞ_İADE'

                if original_items:
                    for item in original_items:
                        self.urun_kalemleri.append({
                            'urun_id': item['urun_id'],
                            'kod': item['urun_kodu'],
                            'ad': item['urun_adi'],
                            'miktar': item['miktar'],
                            'birim_fiyat': item['birim_fiyat'],
                            'kdv_orani': item['kdv_orani'],
                            'kdv_tutari': item['kdv_tutari'],
                            'kalem_toplam_kdv_haric': item['kalem_toplam_kdv_haric'],
                            'kalem_toplam_kdv_dahil': item['kalem_toplam_kdv_dahil']
                        })
                    self._update_treeview()
                    self._calculate_totals()
            except Exception as e:
                messagebox.showerror("Hata", f"Orijinal fatura bilgileri yüklenirken hata oluştu: {e}")
                self.cari_id = None
                self.cari_tip = None
                self.cari_adi_var.set("Cari Hesap Seçiniz...")
                self.urun_kalemleri = []
                self._update_treeview()
                self._calculate_totals()
        
    def _select_cari_hesap(self):
        """Cari Hesap Seçim penceresini açar ve seçimi işler."""
        select_type_for_dialog = None
        if self.invoice_type == 'SATIŞ_İADE':
            select_type_for_dialog = 'MUSTERI'
        elif self.invoice_type == 'ALIŞ_İADE':
            select_type_for_dialog = 'TEDARIKCI'

        cari_secim_penceresi = CariSecimPenceresi(self, self.db, select_type=select_type_for_dialog)
        self.wait_window(cari_secim_penceresi)
        if hasattr(cari_secim_penceresi, 'selected_cari') and cari_secim_penceresi.selected_cari:
            self.cari_id = cari_secim_penceresi.selected_cari['id']
            # CariSecimPenceresi'nden cari_tip'in de geldiğinden emin olun!
            # Musteriler tablosu için 'MUSTERI', Tedarikciler için 'TEDARIKCI' gibi
            self.cari_tip = cari_secim_penceresi.selected_cari['tip']
            self.cari_adi_var.set(f"{cari_secim_penceresi.selected_cari['ad_unvan']} ({cari_secim_penceresi.selected_cari['kod']})")


    def _add_product(self):
        """Ürün Seçim penceresini açar ve seçilen ürünü kalemlere ekler."""
        urun_secim_penceresi = UrunSecimPenceresi(self, self.db)
        self.wait_window(urun_secim_penceresi)
        if hasattr(urun_secim_penceresi, 'selected_product') and urun_secim_penceresi.selected_product:
            selected_product = urun_secim_penceresi.selected_product
            
            found = False
            for item in self.urun_kalemleri:
                if item['urun_id'] == selected_product['id']:
                    item['miktar'] += 1.0
                    found = True
                    break
            if not found:
                birim_fiyat = selected_product['satis_fiyati_kdv_haric'] if self.invoice_type == 'SATIŞ_İADE' else selected_product['alis_fiyati_kdv_haric']
                kdv_orani = selected_product['kdv_orani']

                kalem_toplam_kdv_haric = 1.0 * birim_fiyat
                kdv_tutari = kalem_toplam_kdv_haric * (kdv_orani / 100)
                kalem_toplam_kdv_dahil = kalem_toplam_kdv_haric + kdv_tutari

                self.urun_kalemleri.append({
                    'urun_id': selected_product['id'],
                    'kod': selected_product['urun_kodu'],
                    'ad': selected_product['urun_adi'],
                    'miktar': 1.0,
                    'birim_fiyat': birim_fiyat,
                    'kdv_orani': kdv_orani,
                    'kdv_tutari': kdv_tutari,
                    'kalem_toplam_kdv_haric': kalem_toplam_kdv_haric,
                    'kalem_toplam_kdv_dahil': kalem_toplam_kdv_dahil
                })
            self._update_treeview()
            self._calculate_totals()

    def _delete_selected_product(self):
        """Seçili ürünü Treeview'dan ve listeden siler."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir ürün seçin.")
            return

        items_to_delete_indices = []
        for item_id in selected_item:
            items_to_delete_indices.append(self.tree.index(item_id))

        items_to_delete_indices.sort(reverse=True)
        for index in items_to_delete_indices:
            self.tree.delete(self.tree.get_children()[index])
            del self.urun_kalemleri[index]

        self._calculate_totals()

    def _on_tree_double_click(self, event):
        """Treeview'da bir satıra çift tıklandığında miktarı düzenleme."""
        item_id = self.tree.focus()
        if not item_id:
            return

        column = self.tree.identify_column(event.x)
        if column == "#4": # 'miktar' sütunu
            row_id = self.tree.identify_row(event.y)
            if not row_id:
                return

            col_index = self.tree.column(column, 'id')
            current_value = float(self.tree.set(item_id, col_index))

            entry = NumericEntry(self.tree, width=8)
            entry.insert(0, str(current_value))
            entry.focus_set()

            x, y, width, height = self.tree.bbox(item_id, column)
            entry.place(x=x, y=y, width=width, height=height)

            def on_edit_complete(event=None):
                try:
                    new_value = float(entry.get())
                    if new_value < 0:
                        raise ValueError("Miktar negatif olamaz.")
                    
                    item_index = self.tree.index(item_id)
                    self.urun_kalemleri[item_index]['miktar'] = new_value
                    self._update_treeview()
                    self._calculate_totals()
                except ValueError as e:
                    messagebox.showerror("Hata", f"Geçersiz miktar girişi: {e}")
                finally:
                    entry.destroy()

            entry.bind("<Return>", on_edit_complete)
            entry.bind("<FocusOut>", on_edit_complete)

    def _update_treeview(self):
        """urun_kalemleri listesindeki verileri Treeview'a yansıtır."""
        for i in self.tree.get_children():
            self.tree.delete(i)
        for item in self.urun_kalemleri:
            self.tree.insert("", "end", values=(
                item['urun_id'],
                item['kod'], item['ad'], item['miktar'],
                f"{item['birim_fiyat']:.2f}", item['kdv_orani'], f"{item['kalem_toplam_kdv_dahil']:.2f}"
            ))

    def _calculate_totals(self):
        """Ürün kalemlerine göre ara toplam, KDV toplamı ve genel toplamı hesaplar."""
        subtotal_kdv_haric = 0.0
        vat_total = 0.0
        for item in self.urun_kalemleri:
            kalem_toplam_kdv_haric = item['miktar'] * item['birim_fiyat']
            kdv_tutari = kalem_toplam_kdv_haric * (item['kdv_orani'] / 100)
            kalem_toplam_kdv_dahil = kalem_toplam_kdv_haric + kdv_tutari
            
            item['kalem_toplam_kdv_haric'] = kalem_toplam_kdv_haric
            item['kdv_tutari'] = kdv_tutari
            item['kalem_toplam_kdv_dahil'] = kalem_toplam_kdv_dahil

            subtotal_kdv_haric += kalem_toplam_kdv_haric
            vat_total += kdv_tutari

        grand_total = subtotal_kdv_haric + vat_total
        self.subtotal_var.set(f"{subtotal_kdv_haric:.2f}")
        self.vat_total_var.set(f"{vat_total:.2f}")
        self.grand_total_var.set(f"{grand_total:.2f}")

    def _save_return_invoice(self):
        """İade faturasını veritabanına kaydeder ve ilgili güncellemeleri yapar."""
        fatura_no = self.invoice_no_entry.get().strip()
        iade_tarihi = self.date_entry.get_date()
        aciklama = self.description_entry.get().strip()

        if not fatura_no or not iade_tarihi or not self.cari_id or not self.urun_kalemleri:
            messagebox.showerror("Hata", "Lütfen tüm gerekli alanları doldurun ve en az bir ürün ekleyin.")
            return
        
        if not self.cari_tip:
             messagebox.showerror("Hata", "Cari tipi belirlenemedi. Lütfen cari hesabı tekrar seçin.")
             return

        try:
            # Transaction başlat (DbManager'ınızda varsa)
            # self.db.start_transaction()

            invoice_data = {
                'fatura_no': fatura_no,
                'iade_tarihi': iade_tarihi.strftime('%Y-%m-%d'),
                'cari_tip': self.cari_tip,
                'cari_id': self.cari_id,
                'orijinal_fatura_id': self.original_invoice_id,
                'iade_tipi': self.invoice_type,
                'toplam_kdv_haric': float(self.subtotal_var.get()),
                'toplam_kdv_dahil': float(self.grand_total_var.get()),
                'aciklama': aciklama
            }

            iade_fatura_id = self.db.add_return_invoice(invoice_data, self.current_user_id)

            for item in self.urun_kalemleri:
                item_data = {
                    'iade_fatura_id': iade_fatura_id,
                    'urun_id': item['urun_id'],
                    'miktar': item['miktar'],
                    'birim_fiyat': item['birim_fiyat'],
                    'kdv_orani': item['kdv_orani'],
                    'kdv_tutari': item['kdv_tutari'],
                    'kalem_toplam_kdv_haric': item['kalem_toplam_kdv_haric'],
                    'kalem_toplam_kdv_dahil': item['kalem_toplam_kdv_dahil']
                }
                self.db.add_return_invoice_item(item_data, self.current_user_id)

                # Stok güncellemesi: Satış iadesi stoğu artırır, Alış iadesi stoğu azaltır
                is_increase_stock = (self.invoice_type == 'SATIŞ_İADE')
                self.db.update_stock(item['urun_id'], item['miktar'], is_increase_stock, self.current_user_id, 'IADE_FATURASI', iade_fatura_id)

            # Cari hesap bakiyesini güncelle
            # İade faturaları cari hesap bakiyesini AZALTIR.
            bakiye_etkisi = -float(self.grand_total_var.get())
            self.db.update_current_account_balance(
                self.cari_id, self.cari_tip, bakiye_etkisi, 'IADE_FATURASI',
                self.current_user_id, iade_fatura_id, 'IADE_FATURASI'
            )

            # self.db.commit_transaction()
            messagebox.showinfo("Başarılı", "İade faturası başarıyla kaydedildi.")
            self.destroy()

            # Ana pencereye veya liste penceresine güncelleme sinyali gönder
            if hasattr(self.parent, 'refresh_data'):
                self.parent.refresh_data()
            elif hasattr(self.parent, 'guncelle_ozet_bilgiler'):
                self.parent.guncelle_ozet_bilgiler()


        except ValueError as ve:
            messagebox.showerror("Hata", f"Veri hatası: {ve}")
            # self.db.rollback_transaction()
        except Exception as e:
            messagebox.showerror("Hata", f"İade faturası kaydedilirken bir hata oluştu: {e}")
            # self.db.rollback_transaction()

class FaturaDetayPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, fatura_id):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.fatura_id = fatura_id
        self.f_no = None
        self.tip = None
        self.main_container = None
        self.ust_frame = None
        self.kalem_tree = None
        self.title(f"Fatura Detayları: Yükleniyor...")
        self.geometry("1300x850")
        self.transient(parent_app)
        self.grab_set()

        self.main_container = ttk.Frame(self, padding="15")
        self.main_container.pack(expand=True, fill=tk.BOTH)

        # Sadece bir kez çağrılacak metotlar
        self._verileri_yukle_ve_arayuzu_doldur() # <-- Bu metodu çağırıyoruz
        self._butonlari_olustur() # <-- Bu metodu çağırıyoruz
        # DEĞİŞİKLİK BİTİŞİ

        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _verileri_yukle_ve_arayuzu_doldur(self):
        """
        YENİ METOT: Faturaya ait tüm verileri veritabanından çeker ve arayüzü sıfırdan oluşturup doldurur.
        Bu metot, hem pencere ilk açıldığında hem de güncelleme sonrası yenileme için kullanılır.
        """
        # Önce mevcut arayüz elemanlarını temizle (yenileme durumları için)
        for widget in self.main_container.winfo_children():
            widget.destroy()

        fatura_ana = self.db.fatura_getir_by_id(self.fatura_id)
        if not fatura_ana:
            messagebox.showerror("Fatura Bulunamadı", "Seçilen fatura bilgileri alınamadı.", parent=self)
            self.destroy()
            return

        # Sınıf özelliklerini güncelle
        self.f_no = fatura_ana['fatura_no']
        self.tip = fatura_ana['tip']
        self.title(f"Fatura Detayları: {self.f_no} ({self.tip})")

        # Verileri yerel değişkenlere ata
        tarih_db = fatura_ana['tarih']
        c_id = fatura_ana['cari_id']
        toplam_kdv_haric_fatura_ana_db = fatura_ana['toplam_kdv_haric']
        toplam_kdv_dahil_fatura_ana_db = fatura_ana['toplam_kdv_dahil']
        odeme_turu_db = fatura_ana['odeme_turu']
        misafir_adi_db = fatura_ana['misafir_adi']
        kasa_banka_id_db = fatura_ana['kasa_banka_id']
        olusturma_tarihi_saat = fatura_ana['olusturma_tarihi_saat']
        olusturan_kullanici_id = fatura_ana['olusturan_kullanici_id']
        son_guncelleme_tarihi_saat = fatura_ana['son_guncelleme_tarihi_saat']
        son_guncelleyen_kullanici_id = fatura_ana['son_guncelleyen_kullanici_id']
        fatura_notlari_db = fatura_ana['fatura_notlari']
        vade_tarihi_db = fatura_ana['vade_tarihi']
        genel_iskonto_tipi_db = fatura_ana['genel_iskonto_tipi']
        genel_iskonto_degeri_db = fatura_ana['genel_iskonto_degeri']

        kullanicilar_map = {k[0]: k[1] for k in self.db.kullanici_listele()}
        olusturan_adi = kullanicilar_map.get(olusturan_kullanici_id, "Bilinmiyor")
        son_guncelleyen_adi = kullanicilar_map.get(son_guncelleyen_kullanici_id, "Bilinmiyor")

        cari_adi_text = "Bilinmiyor"
        if str(c_id) == str(self.db.perakende_musteri_id) and self.tip == 'SATIŞ':
            cari_adi_text = "Perakende Satış Müşterisi"
            if misafir_adi_db: cari_adi_text += f" (Misafir: {misafir_adi_db})"
        else:
            cari_bilgi_db = self.db.musteri_getir_by_id(c_id) if self.tip == 'SATIŞ' else self.db.tedarikci_getir_by_id(c_id)
            if cari_bilgi_db:
                cari_adi_text = f"{cari_bilgi_db['ad']} (Kod: {cari_bilgi_db['musteri_kodu'] if 'musteri_kodu' in cari_bilgi_db.keys() else cari_bilgi_db['tedarikci_kodu']})"
            else:
                cari_adi_text = "Bilinmeyen Cari"

        # --- 1. BÖLÜM: ÜST BİLGİLER FRAME'İ ---
        self.ust_frame = ttk.LabelFrame(self.main_container, text=f"Fatura Genel Bilgileri: {self.f_no}", padding="10")
        self.ust_frame.pack(pady=5, padx=5, fill="x")
        self.ust_frame.columnconfigure(1, weight=1)
        self.ust_frame.columnconfigure(3, weight=1)

        row_idx = 0
        ttk.Label(self.ust_frame, text="Fatura No:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text=self.f_no, font=("Segoe UI", 9)).grid(row=row_idx, column=1, sticky="w", padx=5, pady=2)

        try:
            fatura_tarihi_formatted = datetime.strptime(str(tarih_db), '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            fatura_tarihi_formatted = tarih_db
        ttk.Label(self.ust_frame, text="Tarih:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=2, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text=fatura_tarihi_formatted, font=("Segoe UI", 9)).grid(row=row_idx, column=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        ttk.Label(self.ust_frame, text="Fatura Tipi:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text=self.tip, font=("Segoe UI", 9)).grid(row=row_idx, column=1, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text="Ödeme Türü:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=2, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text=odeme_turu_db if odeme_turu_db else "-", font=("Segoe UI", 9)).grid(row=row_idx, column=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        cari_label_tipi = "Müşteri/Misafir:" if self.tip == 'SATIŞ' else "Tedarikçi:"
        ttk.Label(self.ust_frame, text=cari_label_tipi, font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text=cari_adi_text, font=("Segoe UI", 9)).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        if kasa_banka_id_db:
            kb_bilgi = self.db.kasa_banka_getir_by_id(kasa_banka_id_db)
            if kb_bilgi:
                ttk.Label(self.ust_frame, text="Kasa/Banka:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
                ttk.Label(self.ust_frame, text=kb_bilgi['hesap_adi'], font=("Segoe UI", 9)).grid(row=row_idx, column=1, sticky="w", padx=5, pady=2)
                row_idx += 1

        if odeme_turu_db == "AÇIK HESAP" and vade_tarihi_db:
            ttk.Label(self.ust_frame, text="Vade Tarihi:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
            ttk.Label(self.ust_frame, text=vade_tarihi_db, font=("Segoe UI", 9)).grid(row=row_idx, column=1, sticky="w", padx=5, pady=2)
            row_idx += 1

        genel_iskonto_gosterim_text = "Uygulanmadı"
        if genel_iskonto_tipi_db == 'YUZDE' and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Yüzde %{genel_iskonto_degeri_db:.2f}".replace('.', ',')
        elif genel_iskonto_tipi_db == 'TUTAR' and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Tutar {self.db._format_currency(genel_iskonto_degeri_db)}"
        ttk.Label(self.ust_frame, text="Genel İskonto:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text=genel_iskonto_gosterim_text, font=("Segoe UI", 9)).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        ttk.Label(self.ust_frame, text="Oluşturulma:", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(self.ust_frame, text=f"{olusturma_tarihi_saat or '-'} ({olusturan_adi})", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        if son_guncelleme_tarihi_saat:
            ttk.Label(self.ust_frame, text="Son Güncelleme:", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
            ttk.Label(self.ust_frame, text=f"{son_guncelleme_tarihi_saat} ({son_guncelleyen_adi})", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
            row_idx += 1

        ttk.Label(self.ust_frame, text="Fatura Notları:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="nw", padx=5, pady=5)
        fatura_notlari_display_widget = tk.Text(self.ust_frame, height=3, width=50, font=('Segoe UI', 9), wrap=tk.WORD, state=tk.NORMAL)
        fatura_notlari_display_widget.grid(row=row_idx, column=1, columnspan=3, padx=5, pady=5, sticky="ew")
        fatura_notlari_display_widget.insert("1.0", fatura_notlari_db or "")
        fatura_notlari_display_widget.config(state=tk.DISABLED)

        # --- 2. BÖLÜM: FATURA KALEMLERİ FRAME'İ ---
        kalemler_frame = ttk.LabelFrame(self.main_container, text="Fatura Kalemleri", padding="10")
        kalemler_frame.pack(pady=10, padx=5, expand=True, fill="both")
        cols_kalem = ("Sıra", "Ürün Kodu", "Ürün Adı", "Miktar", "Birim Fiyat", "KDV %", "İskonto 1 (%)", "İskonto 2 (%)", "Uyg. İsk. Tutarı", "Tutar (Dah.)", "Alış Fiyatı (Fatura Anı)")
        self.kalem_tree = ttk.Treeview(kalemler_frame, columns=cols_kalem, show='headings', selectmode="none")
        # --- Sütun Tanımları ---
        col_defs_kalem = [
            ("Sıra", 40, tk.CENTER, tk.NO),
            ("Ürün Kodu", 90, tk.W, tk.NO),
            ("Ürün Adı", 180, tk.W, tk.YES),
            ("Miktar", 60, tk.E, tk.NO),
            ("Birim Fiyat", 90, tk.E, tk.NO),
            ("KDV %", 60, tk.E, tk.NO),
            ("İskonto 1 (%)", 75, tk.E, tk.NO),
            ("İskonto 2 (%)", 75, tk.E, tk.NO),
            ("Uyg. İsk. Tutarı", 100, tk.E, tk.NO),
            ("Tutar (Dah.)", 110, tk.E, tk.NO),
            ("Alış Fiyatı (Fatura Anı)", 120, tk.E, tk.NO)
        ]
        for cn, w, a, s in col_defs_kalem:
            self.kalem_tree.column(cn, width=w, anchor=a, stretch=s)
            self.kalem_tree.heading(cn, text=cn)

        vsb_kalem = ttk.Scrollbar(kalemler_frame, orient="vertical", command=self.kalem_tree.yview)
        hsb_kalem = ttk.Scrollbar(kalemler_frame, orient="horizontal", command=self.kalem_tree.xview)
        self.kalem_tree.configure(yscrollcommand=vsb_kalem.set, xscrollcommand=hsb_kalem.set)
        vsb_kalem.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_kalem.pack(side=tk.BOTTOM, fill=tk.X)
        self.kalem_tree.pack(expand=True, fill=tk.BOTH)

        fatura_kalemleri_db_list = self.db.fatura_detay_al(self.fatura_id)
        self._load_fatura_kalemleri_to_treeview(fatura_kalemleri_db_list) # Kalemleri yükle

        # --- 3. BÖLÜM: ALT TOPLAMLAR FRAME'İ ---
        alt_toplam_iskonto_frame = ttk.Frame(self.main_container, padding="10")
        alt_toplam_iskonto_frame.pack(fill="x", pady=(5,0), padx=5, side=tk.BOTTOM)
        alt_toplam_iskonto_frame.columnconfigure(0, weight=1) # Sol tarafı boş bırakmak için

        toplam_kdv_hesaplanan_detay = toplam_kdv_dahil_fatura_ana_db - toplam_kdv_haric_fatura_ana_db
        toplam_kdv_dahil_kalemler_genel_iskonto_oncesi = sum(k['kalem_toplam_kdv_dahil'] for k in fatura_kalemleri_db_list)
        gercek_uygulanan_genel_iskonto = toplam_kdv_dahil_kalemler_genel_iskonto_oncesi - toplam_kdv_dahil_fatura_ana_db

        ttk.Label(alt_toplam_iskonto_frame, text="Toplam KDV Hariç:", font=('Segoe UI', 9, 'bold')).grid(row=0, column=1, sticky="e", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text=self.db._format_currency(toplam_kdv_haric_fatura_ana_db), font=('Segoe UI', 9, 'bold')).grid(row=0, column=2, sticky="w", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text="Toplam KDV:", font=('Segoe UI', 9, 'bold')).grid(row=1, column=1, sticky="e", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text=self.db._format_currency(toplam_kdv_hesaplanan_detay), font=('Segoe UI', 9, 'bold')).grid(row=1, column=2, sticky="w", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text="Genel Toplam:", font=('Segoe UI', 10, "bold")).grid(row=2, column=1, sticky="e", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text=self.db._format_currency(toplam_kdv_dahil_fatura_ana_db), font=('Segoe UI', 10, "bold")).grid(row=2, column=2, sticky="w", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text="Uygulanan Genel İskonto:", font=('Segoe UI', 9, 'bold')).grid(row=3, column=1, sticky="e", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text=self.db._format_currency(gercek_uygulanan_genel_iskonto if gercek_uygulanan_genel_iskonto > 0 else 0.0), font=('Segoe UI', 9, 'bold')).grid(row=3, column=2, sticky="w", padx=5, pady=2)
            
    def _butonlari_olustur(self):
        """YENİ METOT: Pencerenin altındaki butonları oluşturur. Sadece bir kez çağrılır."""
        button_frame_alt = ttk.Frame(self.main_container, padding="5")
        button_frame_alt.pack(fill="x", side=tk.BOTTOM, padx=5, pady=(0,5))

        ttk.Button(button_frame_alt, text="Güncelle", command=self._open_fatura_guncelleme_penceresi, style="Accent.TButton").pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame_alt, text="Kapat", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(button_frame_alt, text="PDF Yazdır", command=self._handle_pdf_print, style="Accent.TButton").pack(side=tk.RIGHT, padx=5)

    def _handle_pdf_print(self):
        """Fatura detay penceresinden PDF yazdırma işlemini başlatır."""
        dosya_adi_onek = f"{self.tip.capitalize()}Faturasi"
        dosya_yolu = filedialog.asksaveasfilename(
            initialfile=f"{dosya_adi_onek}_{self.f_no.replace('/','_')}.pdf",
            defaultextension=".pdf",
            filetypes=[("PDF Dosyaları","*.pdf")],
            title=f"{self.tip.capitalize()} Faturasını PDF Kaydet",
            parent=self.app
        )
        if dosya_yolu:
            success, message = self.db.fatura_pdf_olustur(self.fatura_id, dosya_yolu)
            if success:
                self.app.set_status(message)
                messagebox.showinfo("Başarılı", message, parent=self.app)
            else:
                self.app.set_status(f"PDF kaydetme başarısız: {message}")
                messagebox.showerror("Hata", message, parent=self.app)
        else:
            self.app.set_status("PDF kaydetme iptal edildi.")

    def _open_fatura_guncelleme_penceresi(self):
        """Faturayı güncellemek için FaturaGuncellemePenceresi'ni açar."""
        from pencereler import FaturaGuncellemePenceresi
        FaturaGuncellemePenceresi(
            self, # parent olarak FaturaDetayPenceresi'nin kendisi veriliyor.
            self.db,
            self.fatura_id, # Güncellenecek faturanın ID'si
            yenile_callback_liste=self._fatura_guncellendi_callback_detay # Güncelleme sonrası bu pencereyi yenileyecek callback
        )

    def _fatura_guncellendi_callback_detay(self):
        """GÜNCELLENDİ: Artık çok daha basit. Sadece yeniden yükleme metodunu çağırıyor."""
        self._verileri_yukle_ve_arayuzu_doldur()
        self.app.set_status(f"Fatura '{self.f_no}' detayları güncellendi.")

        # Ana fatura listesini de yenile (örneğin FaturaListesiSayfasi'nı güncelleyebiliriz)
        if hasattr(self.app, 'fatura_listesi_sayfasi'):
            if hasattr(self.app.fatura_listesi_sayfasi.satis_fatura_frame, 'fatura_listesini_yukle'):
                self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
            if hasattr(self.app.fatura_listesi_sayfasi.alis_fatura_frame, 'fatura_listesini_yukle'):
                self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()

    def _load_fatura_kalemleri_to_treeview(self, kalemler_list):
        for i in self.kalem_tree.get_children():
            self.kalem_tree.delete(i)

        sira_idx = 1
        for kalem_item in kalemler_list:
            # ### HATA DÜZELTMESİ: Verilere artık sütun isimleriyle erişiyoruz ###
            miktar_db = kalem_item['miktar']
            toplam_dahil_db = kalem_item['kalem_toplam_kdv_dahil']
            original_birim_fiyat_kdv_haric_item = kalem_item['birim_fiyat']
            original_kdv_orani_item = kalem_item['kdv_orani']

            # İskontolu Birim Fiyat (KDV Dahil) Hesapla
            iskontolu_birim_fiyat_kdv_dahil = (toplam_dahil_db / miktar_db) if miktar_db != 0 else 0.0

            # Uygulanan Kalem İskonto Tutarı (KDV Dahil) Hesapla
            original_birim_fiyat_kdv_dahil_kalem = original_birim_fiyat_kdv_haric_item * (1 + original_kdv_orani_item / 100)
            uygulanan_kalem_iskonto_tutari = (original_birim_fiyat_kdv_dahil_kalem - iskontolu_birim_fiyat_kdv_dahil) * miktar_db

            self.kalem_tree.insert("", tk.END, values=[
                sira_idx,
                kalem_item['urun_kodu'],
                kalem_item['urun_adi'],
                f"{miktar_db:.2f}".rstrip('0').rstrip('.'),
                self.db._format_currency(iskontolu_birim_fiyat_kdv_dahil),
                f"%{kalem_item['kdv_orani']:.0f}",
                f"{kalem_item['iskonto_yuzde_1']:.2f}".replace('.',','),
                f"{kalem_item['iskonto_yuzde_2']:.2f}".replace('.',','),
                self.db._format_currency(uygulanan_kalem_iskonto_tutari),
                self.db._format_currency(toplam_dahil_db),
                self.db._format_currency(kalem_item['alis_fiyati_fatura_aninda'])
            ])
            sira_idx += 1

    def _load_fatura_kalemleri(self):
        for i in self.kalem_tree.get_children():
            self.kalem_tree.delete(i) # Önce temizle

        fatura_kalemleri_db_list = self.db.fatura_detay_al(self.fatura_id)
        sira_idx = 1
        for kalem_item in fatura_kalemleri_db_list:
            miktar_gosterim = f"{kalem_item[2]:.2f}".rstrip('0').rstrip('.')
            alis_fiyati_fatura_aninda = kalem_item[9]
            iskonto_yuzde_1 = kalem_item[11]
            iskonto_yuzde_2 = kalem_item[12]
            iskontolu_birim_fiyat_kdv_dahil = kalem_item[7] / kalem_item[2] if kalem_item[2] != 0 else 0.0

            original_birim_fiyat_kdv_haric_item = kalem_item[3] 
            original_kdv_orani_item = kalem_item[4] 
            original_birim_fiyat_kdv_dahil_item = original_birim_fiyat_kdv_haric_item * (1 + original_kdv_orani_item / 100)
            
            iskonto_farki_per_birim_detay = original_birim_fiyat_kdv_dahil_item - iskontolu_birim_fiyat_kdv_dahil
            uygulanan_toplam_iskonto_tutari_detay = iskonto_farki_per_birim_detay * kalem_item[2] 
            
            self.kalem_tree.insert("", tk.END, values=[
                sira_idx, 
                kalem_item[0], 
                kalem_item[1], 
                miktar_gosterim, 
                self.db._format_currency(iskontolu_birim_fiyat_kdv_dahil), 
                f"%{kalem_item[4]:.0f}", 
                f"{iskonto_yuzde_1:.2f}".replace('.',','), 
                f"{iskonto_yuzde_2:.2f}".replace('.',','), 
                self.db._format_currency(uygulanan_toplam_iskonto_tutari_detay), 
                self.db._format_currency(kalem_item[7]), 
                self.db._format_currency(alis_fiyati_fatura_aninda)
            ])
            sira_idx += 1

    # Yeni yardımcı metot: Bir Label'ı metinle bulup güncellemek için
    def find_and_update_label_by_text(self, parent_widget, label_text_prefix, new_value_text):
        """
        Bir widget hiyerarşisinde belirli bir etiket metniyle başlayan Label'ı bulur ve değerini günceller.
        Tkinter'ın varsayılan Label objelerini ve ttk.Label objelerini de arar.
        """
        for child in parent_widget.winfo_children():
            if isinstance(child, (ttk.Label, tk.Label)):
                try:
                    current_label_text = child.cget("text")
                    if current_label_text.startswith(label_text_prefix):
                        child.config(text=f"{label_text_prefix} {new_value_text}")
                        return True
                except tk.TclError:
                    pass
            if self.find_and_update_label_by_text(child, label_text_prefix, new_value_text):
                return True
        return False

    # Yeni yardımcı metot: Toplam etiketlerini güncellemek için
    def update_summary_labels_detay(self, toplam_kdv_haric, toplam_kdv_dahil, gercek_uygulanan_genel_iskonto):
        """Fatura Detay penceresindeki alt toplam etiketlerini günceller."""
        toplam_kdv = toplam_kdv_dahil - toplam_kdv_haric

        # Alt kısımdaki toplam etiketlerine (tkh_l, tkdv_l, gt_l) doğrudan erişip güncelleyelim.
        # Bu etiketlerin __init__ içinde self. olarak tanımlanmış olması gerekir.
        self.tkh_l.config(text=f"Toplam KDV Hariç: {self.db._format_currency(toplam_kdv_haric)}")
        self.tkdv_l.config(text=f"Toplam KDV: {self.db._format_currency(toplam_kdv)}")
        self.gt_l.config(text=f"Genel Toplam: {self.db._format_currency(toplam_kdv_dahil)}")
        
        if gercek_uygulanan_genel_iskonto > 0:
            self.lbl_uygulanan_genel_iskonto.config(text=f"Uygulanan Genel İskonto: {self.db._format_currency(gercek_uygulanan_genel_iskonto)}")
        else:
            self.lbl_uygulanan_genel_iskonto.config(text="Uygulanan Genel İskonto: 0,00 TL")

class SiparisDetayPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, siparis_id, yenile_callback=None):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.siparis_id = siparis_id
        self.yenile_callback = yenile_callback

        siparis_ana_info = self.db.get_siparis_by_id(self.siparis_id)
        if not siparis_ana_info:
            messagebox.showerror("Sipariş Bulunamadı", "Seçilen sipariş bilgileri alınamadı.", parent=self)
            self.destroy()
            return
        
        self.siparis_ana = siparis_ana_info 
        self.s_no = self.siparis_ana['siparis_no']
        durum_db = self.siparis_ana['durum']

        # Fatura ana verilerini ayıklama (17 elemanlı tuple'dan)
        _id, s_no_db, tarih_db, c_tip_db, c_id_db, toplam_tutar_db, durum_db, fatura_id_ref_db, \
        olusturma_tarihi_saat, olusturan_kullanici_id, son_guncelleme_tarihi_saat, \
        son_guncelleyen_kullanici_id, siparis_notlari_db, onay_durumu_db, teslimat_tarihi_db, \
        genel_iskonto_tipi_db, genel_iskonto_degeri_db = self.siparis_ana # Artık self.siparis_ana'yı kullanın
        
        self.s_no = s_no_db 

        self.title(f"Sipariş Detayları: {self.s_no} ({durum_db})")
        self.geometry("1000x700")
        self.transient(parent_app) 
        self.grab_set()
        self.resizable(True, True)

        # TÜM KULLANICILARI ÇEK VE BİR SÖZLÜĞE DÖNÜŞTÜR (ID -> Kullanıcı Adı)
        kullanicilar_map = {k[0]: k[1] for k in self.db.kullanici_listele()}

        # Oluşturan ve Güncelleyen kullanıcı adlarını al
        olusturan_adi = kullanicilar_map.get(olusturan_kullanici_id, "Bilinmiyor") 
        son_guncelleyen_adi = kullanicilar_map.get(son_guncelleyen_kullanici_id, "Bilinmiyor") 

        # Cari bilgisini al (Müşteri/Tedarikçi)
        cari_adi_text = "Bilinmiyor"
        if c_tip_db == 'MUSTERI':
            cari_bilgi_db = self.db.musteri_getir_by_id(c_id_db)
            cari_adi_text = f"{cari_bilgi_db[2]} (Kod: {cari_bilgi_db[1]})" if cari_bilgi_db else "Bilinmiyor"
        elif c_tip_db == 'TEDARIKCI':
            cari_bilgi_db = self.db.tedarikci_getir_by_id(c_id_db)
            cari_adi_text = f"{cari_bilgi_db[2]} (Kod: {cari_bilgi_db[1]})" if cari_bilgi_db else "Bilinmiyor"

        # Ana Kapsayıcı Frame
        main_container = ttk.Frame(self, padding="15")
        main_container.pack(expand=True, fill=tk.BOTH)

        # --- Üst Kısım: Sipariş Genel Bilgileri ---
        # DÜZELTME: ust_frame'in text argümanında 's_no' yerine 'self.s_no' kullanın
        ust_frame = ttk.LabelFrame(main_container, text=f"Sipariş Genel Bilgileri: {self.s_no}", padding="10")
        ust_frame.pack(pady=5, padx=5, fill="x")
        
        ust_frame.columnconfigure(1, weight=1) 
        ust_frame.columnconfigure(3, weight=1) 

        row_idx = 0
        ttk.Label(ust_frame, text="Sipariş No:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        # DÜZELTME: text argümanında 's_no' yerine 'self.s_no' kullanın
        ttk.Label(ust_frame, text=self.s_no, font=("Segoe UI", 9)).grid(row=row_idx, column=1, sticky="w", padx=5, pady=2)
        try:
            siparis_tarihi_formatted = datetime.strptime(tarih_db, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            siparis_tarihi_formatted = tarih_db 
        ttk.Label(ust_frame, text="Tarih:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=2, sticky="w", padx=5, pady=2)
        ttk.Label(ust_frame, text=siparis_tarihi_formatted, font=("Segoe UI", 9)).grid(row=row_idx, column=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        ttk.Label(ust_frame, text="Cari Tipi:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(ust_frame, text=c_tip_db, font=("Segoe UI", 9)).grid(row=row_idx, column=1, sticky="w", padx=5, pady=2)
        ttk.Label(ust_frame, text="Durum:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=2, sticky="w", padx=5, pady=2)
        ttk.Label(ust_frame, text=durum_db, font=("Segoe UI", 9)).grid(row=row_idx, column=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        ttk.Label(ust_frame, text="Cari Bilgisi:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(ust_frame, text=cari_adi_text, font=("Segoe UI", 9)).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        ttk.Label(ust_frame, text="Teslimat Tarihi:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        try:
            teslimat_tarihi_formatted = datetime.strptime(teslimat_tarihi_db, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            teslimat_tarihi_formatted = teslimat_tarihi_db
        ttk.Label(ust_frame, text=teslimat_tarihi_formatted if teslimat_tarihi_formatted else "-", font=("Segoe UI", 9)).grid(row=row_idx, column=1, sticky="w", padx=5, pady=2)
        row_idx += 1

        # Genel İskonto Bilgisi
        genel_iskonto_gosterim_text = "Uygulanmadı"
        if genel_iskonto_tipi_db == 'YUZDE' and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Yüzde %{genel_iskonto_degeri_db:.2f}".replace('.', ',')
        elif genel_iskonto_tipi_db == 'TUTAR' and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Tutar {self.db._format_currency(genel_iskonto_degeri_db)}"
        
        ttk.Label(ust_frame, text="Genel İskonto:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(ust_frame, text=genel_iskonto_gosterim_text, font=("Segoe UI", 9)).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
        row_idx += 1

        ttk.Label(ust_frame, text="Oluşturulma:", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(ust_frame, text=f"{olusturma_tarihi_saat if olusturma_tarihi_saat else '-'} ({olusturan_adi})", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
        row_idx += 1
        
        if son_guncelleme_tarihi_saat:
            ttk.Label(ust_frame, text="Son Güncelleme:", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
            ttk.Label(ust_frame, text=f"{son_guncelleme_tarihi_saat} ({son_guncelleyen_adi})", font=("Segoe UI", 8, "italic")).grid(row=row_idx, column=1, columnspan=3, sticky="w", padx=5, pady=2)
            row_idx += 1

        ttk.Label(ust_frame, text="Sipariş Notları:", font=("Segoe UI", 9, "bold")).grid(row=row_idx, column=0, sticky="nw", padx=5, pady=5) 
        siparis_notlari_display = tk.Text(ust_frame, height=3, width=50, font=('Segoe UI', 9), wrap=tk.WORD, state=tk.DISABLED)
        siparis_notlari_display.grid(row=row_idx, column=1, columnspan=3, padx=5, pady=5, sticky="ew")
        siparis_notlari_display.insert("1.0", siparis_notlari_db if siparis_notlari_db else "")
        row_idx += 1


        # --- Orta Kısım: Sipariş Kalemleri ---
        kalemler_frame = ttk.LabelFrame(main_container, text="Sipariş Kalemleri", padding="10")
        kalemler_frame.pack(pady=10, padx=5, expand=True, fill="both")

        # GÜNCEL SÜTUN BAŞLIKLARI VE GENİŞLİKLERİ (Sepet Treeview ile uyumlu)
        cols_kalem = ("Sıra", "Ürün Kodu", "Ürün Adı", "Miktar", "Birim Fiyat", "KDV %", "İskonto 1 (%)", "İskonto 2 (%)", "Uyg. İsk. Tutarı", "Tutar (Dah.)", "Alış Fiyatı (Sipariş Anı)", "Satış Fiyatı (Sipariş Anı)") # Yeni
        self.kalem_tree = ttk.Treeview(kalemler_frame, columns=cols_kalem, show='headings', selectmode="none") 
        
        col_widths_kalem = {
            "Sıra": 40, "Ürün Kodu":90, "Ürün Adı":180, "Miktar":60, "Birim Fiyat":90, "KDV %":60, 
            "İskonto 1 (%)":75, "İskonto 2 (%)":75, "Uyg. İsk. Tutarı":100, "Tutar (Dah.)":110, 
            "Alış Fiyatı (Sipariş Anı)":120, "Satış Fiyatı (Sipariş Anı)":120 # Yeni
        }
        col_anchors_kalem = {
            "Sıra":tk.CENTER, "Miktar":tk.E, "Birim Fiyat":tk.E, "KDV %":tk.E, 
            "İskonto 1 (%)":tk.E, "İskonto 2 (%)":tk.E, "Uyg. İsk. Tutarı":tk.E, "Tutar (Dah.)":tk.E,
            "Alış Fiyatı (Sipariş Anı)":tk.E, "Satış Fiyatı (Sipariş Anı)":tk.E # Yeni
        } 
        
        for col in cols_kalem:
            self.kalem_tree.heading(col, text=col)
            self.kalem_tree.column(col, width=col_widths_kalem.get(col, 80), anchor=col_anchors_kalem.get(col, tk.W), stretch=tk.YES)

        vsb_kalem = ttk.Scrollbar(kalemler_frame, orient="vertical", command=self.kalem_tree.yview)
        hsb_kalem = ttk.Scrollbar(kalemler_frame, orient="horizontal", command=self.kalem_tree.xview) # Horizontal scrollbar
        self.kalem_tree.configure(yscrollcommand=vsb_kalem.set, xscrollcommand=hsb_kalem.set)
        vsb_kalem.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_kalem.pack(side=tk.BOTTOM, fill=tk.X) # Pack horizontal scrollbar
        self.kalem_tree.pack(expand=True, fill=tk.BOTH)
        
        siparis_kalemleri_db_list = self.db.get_siparis_kalemleri(self.siparis_id)
        # k_db yapısı (17 elemanlı):
        # (id, siparis_id, urun_id, miktar, birim_fiyat, kdv_orani, kdv_tutari, kalem_toplam_kdv_haric,
        #  kalem_toplam_kdv_dahil, alis_fiyati_siparis_aninda, satis_fiyati_siparis_aninda, iskonto_yuzde_1,
        #  iskonto_yuzde_2, olusturma_tarihi_saat, olusturan_kullanici_id, son_guncelleme_tarihi_saat, son_guncelleyen_kullanici_id)
        
        sira_idx = 1
        for k_db in siparis_kalemleri_db_list:
            urun_id_db = k_db[2]
            urun_info = self.db.stok_getir_by_id(urun_id_db) # Ürün adını çekmek için
            if not urun_info: continue
            urun_adi_db = urun_info[2]

            miktar_gosterim = f"{k_db[3]:.2f}".rstrip('0').rstrip('.') # Miktar (index 3)
            # Birim fiyat (index 4) aslında iskontosuz kdv hariç fiyat. 
            # Treeview'de iskontolu kdv dahil nihai birim fiyatı göstermeliyiz.
            # Bu, kalem_toplam_kdv_dahil (index 8) / miktar (index 3) olarak hesaplanabilir.
            iskontolu_birim_fiyat_kdv_dahil_display = (k_db[8] / k_db[3]) if k_db[3] != 0 else 0.0

            iskonto_yuzde_1_display = f"{k_db[11]:.2f}".replace('.', ',').rstrip('0').rstrip(',') # İskonto 1 (index 11)
            iskonto_yuzde_2_display = f"{k_db[12]:.2f}".replace('.', ',').rstrip('0').rstrip(',') # İskonto 2 (index 12)
            
            # Uygulanan toplam kalem iskontosu (TL olarak) hesapla
            # Bu, iskontosuz orijinal birim fiyat KDV dahilinden, iskontolu KDV dahil birim fiyatı çıkarılarak bulunur.
            original_birim_fiyat_kdv_dahil_kalem = k_db[4] * (1 + k_db[5] / 100) # orijinal birim_fiyat (index 4) ve kdv_orani (index 5)
            iskonto_farki_per_birim_detay = original_birim_fiyat_kdv_dahil_kalem - iskontolu_birim_fiyat_kdv_dahil_display
            uygulanan_toplam_iskonto_tutari_detay = iskonto_farki_per_birim_detay * k_db[3] 

            alis_fiyati_siparis_aninda_display = self.db._format_currency(k_db[9]) # Alış Fiyatı (index 9)
            satis_fiyati_siparis_aninda_display = self.db._format_currency(k_db[10]) # Satış Fiyatı (index 10)
            
            self.kalem_tree.insert("", tk.END, values=[
                sira_idx, # Sıra No
                urun_info[1], # Ürün Kodu (stok_getir_by_id'den)
                urun_adi_db, # Ürün Adı
                miktar_gosterim, # Miktar
                self.db._format_currency(iskontolu_birim_fiyat_kdv_dahil_display), # Birim Fiyat (KDV Dahil - İSKONTOLU)
                f"%{k_db[5]:.0f}", # KDV Oranı (orijinal)
                iskonto_yuzde_1_display, # İskonto 1 (%)
                iskonto_yuzde_2_display, # İskonto 2 (%)
                self.db._format_currency(uygulanan_toplam_iskonto_tutari_detay), # Uygulanan İskonto Tutarı
                self.db._format_currency(k_db[8]), # Tutar (Dah.) (iskontolu toplam)
                alis_fiyati_siparis_aninda_display, # Alış Fiyatı (Sipariş Anı)
                satis_fiyati_siparis_aninda_display # Satış Fiyatı (Sipariş Anı)
            ])
            sira_idx += 1
            
        # --- Alt Kısım: Sipariş Toplamları ve Butonlar ---
        alt_toplam_iskonto_frame = ttk.Frame(main_container, padding="10")
        alt_toplam_iskonto_frame.pack(fill="x", pady=(5,0), padx=5, side=tk.BOTTOM)

        alt_toplam_iskonto_frame.columnconfigure(0, weight=1) # Sol tarafı boş bırakmak için
        alt_toplam_iskonto_frame.columnconfigure(1, weight=0) 
        alt_toplam_iskonto_frame.columnconfigure(2, weight=0) 

        ttk.Label(alt_toplam_iskonto_frame, text="Genel Toplam (KDV Dahil):", font=('Segoe UI', 10, 'bold')).grid(row=0, column=1, sticky="e", padx=5, pady=2)
        ttk.Label(alt_toplam_iskonto_frame, text=self.db._format_currency(toplam_tutar_db), font=('Segoe UI', 10, 'bold')).grid(row=0, column=2, sticky="w", padx=5, pady=2)
        
        # Eğer siparişte genel iskonto varsa göster
        if genel_iskonto_tipi_db != 'YOK' and genel_iskonto_degeri_db > 0:
            ttk.Label(alt_toplam_iskonto_frame, text="Uygulanan Genel İskonto:", font=('Segoe UI', 9, 'bold')).grid(row=1, column=1, sticky="e", padx=5, pady=2)
            
            genel_iskonto_gosterim_detay = ""
            if genel_iskonto_tipi_db == 'YUZDE':
                genel_iskonto_gosterim_detay = f"%{genel_iskonto_degeri_db:.2f}".replace('.', ',').rstrip('0').rstrip(',')
            elif genel_iskonto_tipi_db == 'TUTAR':
                genel_iskonto_gosterim_detay = self.db._format_currency(genel_iskonto_degeri_db)
            
            ttk.Label(alt_toplam_iskonto_frame, text=genel_iskonto_gosterim_detay, font=('Segoe UI', 9)).grid(row=1, column=2, sticky="w", padx=5, pady=2)


        # Butonlar
        button_frame_alt = ttk.Frame(main_container, padding="5")
        button_frame_alt.pack(fill="x", side=tk.BOTTOM, padx=5, pady=(0,5))
        
        # Faturaya Dönüştür butonu
        self.faturaya_donustur_button_detail = ttk.Button(button_frame_alt, text="Faturaya Dönüştür", command=self._faturaya_donustur, style="Accent.TButton")
        self.faturaya_donustur_button_detail.pack(side=tk.RIGHT, padx=5)
        
        # Siparişi Düzenle butonu
        ttk.Button(button_frame_alt, text="Siparişi Düzenle", command=self._siparisi_duzenle).pack(side=tk.RIGHT, padx=5)
        
        # Kapat butonu
        ttk.Button(button_frame_alt, text="Kapat", command=self.destroy).pack(side=tk.RIGHT)

        # Eğer sipariş zaten bir faturaya dönüştürülmüşse, "Faturaya Dönüştür" butonunu pasif yap
        if fatura_id_ref_db:
            self.faturaya_donustur_button_detail.config(state=tk.DISABLED)
            ttk.Label(button_frame_alt, text=f"Bu sipariş Fatura No: '{fatura_id_ref_db}' ile ilişkilendirilmiştir.", foreground="blue", font=("Segoe UI", 8, "italic")).pack(side=tk.RIGHT, padx=10)


    def _faturaya_donustur(self):
        """Bu siparişi satış veya alış faturasına dönüştürür."""
        
        # DÜZELTME: Ödeme Türü Seçim Diyaloğunu açın
        from arayuz import OdemeTuruSecimDialog # Lokal import

        # Cari tipine göre fatura tipi belirlenmeli
        fatura_tipi_for_dialog = 'SATIŞ' if self.siparis_ana[3] == 'MUSTERI' else 'ALIŞ'
        
        # Callback fonksiyonu olarak _faturaya_donustur_on_dialog_confirm'i gönderiyoruz.
        OdemeTuruSecimDialog(
            self.app, 
            self.db, 
            fatura_tipi_for_dialog, # Diyaloğa fatura tipini gönder
            self.siparis_ana[4], # Diyaloğa cari ID'sini gönder (perakende kontrolü için)
            self._faturaya_donustur_on_dialog_confirm # Callback fonksiyonu
        )

    def _faturaya_donustur_on_dialog_confirm(self, selected_odeme_turu, selected_kasa_banka_id, selected_vade_tarihi):
        """
        OdemeTuruSecimDialog'dan onay geldikten sonra çağrılır.
        Siparişi faturaya dönüştürme işlemini gerçekleştirir.
        """
        if selected_odeme_turu is None: # Eğer dialog iptal edildiyse
            self.app.set_status("Faturaya dönüştürme iptal edildi (ödeme türü seçilmedi).")
            return

        confirm_msg = (f"'{self.s_no}' numaralı siparişi '{selected_odeme_turu}' ödeme türü ile faturaya dönüştürmek istediğinizden emin misiniz?\n"
                       f"Bu işlem sonucunda yeni bir fatura oluşturulacak ve sipariş durumu güncellenecektir.")
        if selected_odeme_turu == "AÇIK HESAP" and selected_vade_tarihi:
            confirm_msg += f"\nVade Tarihi: {selected_vade_tarihi}"
        if selected_kasa_banka_id:
            kb_bilgi = self.db.kasa_banka_getir_by_id(selected_kasa_banka_id)
            if kb_bilgi:
                confirm_msg += f"\nİşlem Kasa/Banka: {kb_bilgi[1]}"

        confirm = messagebox.askyesno("Faturaya Dönüştür Onayı", confirm_msg, parent=self.app)
        if not confirm:
            return

        # self.db.siparis_faturaya_donustur metodunu çağır
        # Burada siparis_faturaya_donustur metodunun ek parametreler alması gerekecek.
        success, message = self.db.siparis_faturaya_donustur(
            self.siparis_id, 
            self.app.current_user[0] if self.app and self.app.current_user else None,
            selected_odeme_turu,       # Yeni
            selected_kasa_banka_id,    # Yeni
            selected_vade_tarihi       # Yeni
        )

        if success:
            messagebox.showinfo("Başarılı", message, parent=self.app)
            self.destroy() # Faturaya dönüştürme başarılıysa sipariş detay penceresini kapat.
            # Sipariş listesini yenile (Ana pencerede)
            if hasattr(self.app, 'siparis_listesi_sayfasi') and hasattr(self.app.siparis_listesi_sayfasi, 'siparis_listesini_yukle'):
                self.app.siparis_listesi_sayfasi.siparis_listesini_yukle()
            # Fatura listesini yenile (Ana pencerede)
            if hasattr(self.app, 'fatura_listesi_sayfasi'):
                if hasattr(self.app.fatura_listesi_sayfasi.satis_fatura_frame, 'fatura_listesini_yukle'):
                    self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
                if hasattr(self.app.fatura_listesi_sayfasi.alis_fatura_frame, 'fatura_listesini_yukle'):
                    self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self.app)

    def _siparisi_duzenle(self):
        """Bu siparişi düzenleme penceresinde açar."""
        # Sipariş oluşturma/düzenleme penceresini açmak için SiparisOlusturmaSayfasi'nı çağır
        from arayuz import SiparisOlusturmaSayfasi # Lokal import
        siparis_tipi_db = 'SATIŞ_SIPARIS' if self.siparis_ana['cari_tip'] == 'MUSTERI' else 'ALIŞ_SIPARIS'
        SiparisPenceresi(
            parent=self.app, 
            db_manager=self.db,
            app_ref=self.app,
            siparis_tipi=siparis_tipi_db,
            siparis_id_duzenle=self.siparis_id,
            yenile_callback=self.yenile_callback # Ana listeden gelen yenileme fonksiyonunu aktarıyoruz
        )
        self.destroy()

class YoneticiAyarlariPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.title("Yönetici Ayarları ve Veri İşlemleri")
        self.geometry("600x450") # Pencereyi biraz büyütelim
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Veri Sıfırlama ve Bakım", font=("Segoe UI", 16, "bold")).pack(pady=15)

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # <<< DÜZELTME: Yeni buton ve açıklaması eklendi >>>
        buttons_info = [
            ("Stok Envanterini Yeniden Hesapla", "Tüm stokları faturalara göre sıfırdan hesaplar. Geçmiş hatalı silme işlemlerini düzeltir.", self.db.stok_envanterini_yeniden_hesapla),
            ("Stok Verilerini Temizle", "Bu işlem tüm ürünleri ve ilişkili kalemleri siler.", self.db.clear_stok_data),
            ("Müşteri Verilerini Temizle", "Bu işlem perakende müşteri hariç tüm müşterileri ve ilişkili hareketlerini siler.", self.db.clear_musteri_data),
            ("Tedarikçi Verilerini Temizle", "Bu işlem tüm tedarikçileri ve ilişkili hareketlerini siler.", self.db.clear_tedarikci_data),
            ("Kasa/Banka Verilerini Temizle", "Bu işlem tüm kasa/banka hesaplarını temizler ve ilişkili referansları kaldırır.", self.db.clear_kasa_banka_data),
            ("Tüm İşlem Verilerini Temizle", "Bu işlem faturalar, gelir/gider, cari hareketler, siparişler ve teklifler gibi tüm operasyonel verileri siler. Stok ve Kasa/Banka bakiyeleri sıfırlanır. Ana kayıtlar korunur.", self.db.clear_all_transaction_data),
            ("Tüm Verileri Temizle (Kullanıcılar Hariç)", "Bu işlem kullanıcılar ve şirket ayarları hariç tüm veritabanını temizler. Program yeniden başlatılacaktır.", self.db.clear_all_data)
        ]

        for i, (text, desc, func) in enumerate(buttons_info):
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(fill=tk.X, pady=5)
            
            # Stok Yeniden Hesaplama butonu için farklı bir stil ve renk
            style_name = "Accent.TButton" if "Yeniden Hesapla" in text else "TButton"
            btn = ttk.Button(btn_frame, text=text, command=lambda f=func, t=text: self._confirm_and_run_utility(f, t), style=style_name)
            btn.pack(side=tk.LEFT, padx=5)
            
            ttk.Label(btn_frame, text=desc, wraplength=350, font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        ttk.Button(self, text="Kapat", command=self.destroy).pack(pady=10)

    def _confirm_and_run_utility(self, utility_function, button_text):
        """Veri işleminden önce onay alır ve işlemi gerçekleştirir."""
        confirm_message = f"'{button_text}' işlemini gerçekleştirmek istediğinizden emin misiniz?\n\nBU İŞLEM GERİ ALINAMAZ!"
        if button_text == "Tüm Verileri Temizle (Kullanıcılar Hariç)":
             confirm_message += "\n\nBu işlemden sonra program yeniden başlatılacaktır."

        if messagebox.askyesno("Onay Gerekli", confirm_message, icon='warning', parent=self):
            success, message = utility_function()

            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.app.set_status(message)

                # İlgili listeleri yenile
                self.app.stok_yonetimi_sayfasi.stok_listesini_yenile()
                self.app.ana_sayfa.guncelle_ozet_bilgiler()

                if button_text == "Tüm Verileri Temizle (Kullanıcılar Hariç)":
                    self.app.cikis_yap_ve_giris_ekranina_don()
            else:
                messagebox.showerror("Hata", message, parent=self)
                self.app.set_status(f"'{button_text}' işlemi sırasında hata oluştu: {message}")
        else:
            self.app.set_status(f"'{button_text}' işlemi iptal edildi.")

    def _confirm_and_clear_data(self, clear_function, button_text):
        """Veri temizleme işleminden önce onay alır ve işlemi gerçekleştirir."""
        confirm_message = f"'{button_text}' işlemini gerçekleştirmek istediğinizden emin misiniz?\n\nBU İŞLEM GERİ ALINAMAZ!"
        if button_text == "Tüm Verileri Temizle (Kullanıcılar Hariç)":
             confirm_message += "\n\nBu işlemden sonra program yeniden başlatılacaktır."

        if messagebox.askyesno("Onay Gerekli", confirm_message, icon='warning', parent=self):
            success, message = clear_function()

            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.app.set_status(message)

                if button_text == "Tüm Verileri Temizle (Kullanıcılar Hariç)":
                    messagebox.showinfo("Bilgi", "Tüm veriler temizlendi. Program yeniden başlatılıyor...", parent=self)
                    self.app.cikis_yap_ve_giris_ekranina_don()
                else:
                    if hasattr(self.app, 'ana_sayfa') and hasattr(self.app.ana_sayfa, 'guncelle_ozet_bilgiler'):
                        self.app.ana_sayfa.guncelle_ozet_bilgiler()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, 'stok_listesini_yenile'):
                        self.app.stok_yonetimi_sayfasi.stok_listesini_yenile()
                    if hasattr(self.app, 'musteri_yonetimi_sayfasi') and hasattr(self.app.musteri_yonetimi_sayfasi, 'musteri_listesini_yenile'):
                        self.app.musteri_yonetimi_sayfasi.musteri_listesini_yenile()
                    if hasattr(self.app, 'tedarikci_yonetimi_sayfasi') and hasattr(self.app.tedarikci_yonetimi_sayfasi, 'tedarikci_listesini_yenile'):
                        self.app.tedarikci_yonetimi_sayfasi.tedarikci_listesini_yenile()
                    if hasattr(self.app, 'kasa_banka_yonetimi_sayfasi') and hasattr(self.app.kasa_banka_yonetimi_sayfasi, 'hesap_listesini_yenile'):
                        self.app.kasa_banka_yonetimi_sayfasi.hesap_listesini_yenile()
                    if hasattr(self.app, 'fatura_listesi_sayfasi') and hasattr(self.app.fatura_listesi_sayfasi, 'satis_fatura_frame') and hasattr(self.app.fatura_listesi_sayfasi.satis_fatura_frame, 'fatura_listesini_yukle'):
                         self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
                    if hasattr(self.app, 'fatura_listesi_sayfasi') and hasattr(self.app.fatura_listesi_sayfasi, 'alis_fatura_frame') and hasattr(self.app.fatura_listesi_sayfasi.alis_fatura_frame, 'fatura_listesini_yukle'):
                         self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()
                    if hasattr(self.app, 'gelir_gider_sayfasi') and hasattr(self.app.gelir_gider_sayfasi, 'gelir_listesi_frame') and hasattr(self.app.gelir_gider_sayfasi.gelir_listesi_frame, 'gg_listesini_yukle'):
                        self.app.gelir_gider_sayfasi.gelir_listesi_frame.gg_listesini_yukle()
                    if hasattr(self.app, 'gelir_gider_sayfasi') and hasattr(self.app.gelir_gider_sayfasi, 'gider_listesi_frame') and hasattr(self.app.gelir_gider_sayfasi.gider_listesi_frame, 'gg_listesini_yukle'):
                        self.app.gelir_gider_sayfasi.gider_listesi_frame.gg_listesini_yukle()
            else:
                messagebox.showerror("Hata", message, parent=self)
                self.app.set_status(f"'{button_text}' işlemi sırasında hata oluştu: {message}")
        else:
            self.app.set_status(f"'{button_text}' işlemi iptal edildi.")

class SirketBilgileriPenceresi(tk.Toplevel):
    def __init__(self, parent, db_manager):
        super().__init__(parent)
        self.db = db_manager
        self.app_parent = parent # Ana App referansı
        self.title("Şirket Bilgileri")
        self.geometry("550x400")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text="Şirket Bilgileri Yönetimi", font=("Segoe UI", 16, "bold")).pack(pady=10)
        
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # Labels ve karşılık gelen veritabanı anahtarlarını doğrudan eşleştiriyoruz
        # Bu, labels listesindeki "Şirket Adı:" ile db_key_map'teki "sirket_adı" karmaşasını ortadan kaldırır.
        # entries sözlüğü artık doğrudan veritabanı anahtarlarını tutacak.
        self.field_definitions = [
            ("Şirket Adı:", "sirket_adi", ttk.Entry),
            ("Adres:", "sirket_adresi", tk.Text, {"height": 3}),
            ("Telefon:", "sirket_telefonu", ttk.Entry),
            ("E-mail:", "sirket_email", ttk.Entry),
            ("Vergi Dairesi:", "sirket_vergi_dairesi", ttk.Entry),
            ("Vergi No:", "sirket_vergi_no", ttk.Entry),
            ("Logo Yolu:", "sirket_logo_yolu", ttk.Entry)
        ]
        self.entries = {}

        for i, (label_text, db_key_name, widget_type, *args) in enumerate(self.field_definitions):
            ttk.Label(main_frame, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky=tk.W)
            
            widget_options = args[0] if args else {}

            if widget_type == tk.Text:
                self.entries[db_key_name] = tk.Text(main_frame, width=40, **widget_options)
            else: # ttk.Entry
                self.entries[db_key_name] = ttk.Entry(main_frame, width=50, **widget_options)
            
            self.entries[db_key_name].grid(row=i, column=1, padx=5, pady=5, sticky=tk.EW)
            
            if db_key_name == "sirket_logo_yolu":
                logo_button = ttk.Button(main_frame, text="Gözat...", command=self.logo_gozat)
                logo_button.grid(row=i, column=2, padx=5, pady=5, sticky=tk.W)

        main_frame.columnconfigure(1, weight=1) # Entry'lerin genişlemesi için

        self.yukle_bilgiler()

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=len(self.field_definitions), column=0, columnspan=3, pady=10, sticky=tk.E)
        
        ttk.Button(button_frame, text="Kaydet", command=self.kaydet_bilgiler, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.LEFT)

    def logo_gozat(self):
        dosya_yolu = filedialog.askopenfilename(
            title="Logo Seçin",
            filetypes=(("PNG Dosyaları", "*.png"), ("JPEG Dosyaları", "*.jpg;*.jpeg"), ("Tüm Dosyalar", "*.*")),
            parent=self
        )
        if dosya_yolu:
            self.entries["sirket_logo_yolu"].delete(0, tk.END)
            self.entries["sirket_logo_yolu"].insert(0, dosya_yolu)

    def yukle_bilgiler(self):
        mevcut_bilgiler = self.db.sirket_bilgilerini_yukle()
        for db_key_name, entry_widget in self.entries.items():
            if isinstance(entry_widget, tk.Text):
                entry_widget.delete("1.0", tk.END)
                entry_widget.insert("1.0", mevcut_bilgiler.get(db_key_name, ""))
            else:
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, mevcut_bilgiler.get(db_key_name, ""))
    
    def kaydet_bilgiler(self):
        yeni_bilgiler = {}
        for db_key_name, entry_widget in self.entries.items():
            if isinstance(entry_widget, tk.Text):
                yeni_bilgiler[db_key_name] = entry_widget.get("1.0", tk.END).strip()
            else:
                yeni_bilgiler[db_key_name] = entry_widget.get().strip()

        print(f"DEBUG: kaydet_bilgiler - yeni_bilgiler sözlüğü: {yeni_bilgiler}")
        success, message = self.db.sirket_bilgilerini_kaydet(yeni_bilgiler)
        if success:
            if hasattr(self.app_parent, 'ana_sayfa') and hasattr(self.app_parent.ana_sayfa, 'guncelle_sirket_adi'):
                self.app_parent.ana_sayfa.guncelle_sirket_adi()
            if hasattr(self.app_parent, 'set_status'):
                 self.app_parent.set_status(message)
            messagebox.showinfo("Başarılı", message, parent=self)
            self.destroy()
        else:
            messagebox.showerror("Hata", message, parent=self)

class StokHareketiPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, urun_id, urun_adi, mevcut_stok, hareket_yönü, yenile_stok_listesi_callback, parent_pencere=None):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.urun_id = urun_id
        self.urun_adi = urun_adi
        self.mevcut_stok = mevcut_stok
        self.hareket_yönü = hareket_yönü
        self.yenile_stok_listesi_callback = yenile_stok_listesi_callback

        self.urun_karti_penceresi_ref = parent_pencere # <-- Bu referans burada saklanıyor
        print(f"DEBUG: StokHareketiPenceresi __init__ - parent_pencere: {parent_pencere}") # <-- YENİ DEBUG
        if parent_pencere:
            print(f"DEBUG: StokHareketiPenceresi __init__ - parent_pencere tipi: {type(parent_pencere)}")

        self.urun_karti_penceresi_ref = None
        if isinstance(self.master, tk.Toplevel) and self.master.winfo_class() == 'Toplevel':
            self.urun_karti_penceresi_ref = self.master
            
        self.title(f"Stok Hareketi: {self.urun_adi}")
        self.geometry("400x350")
        self.resizable(False, False)
        self.transient(parent_app)
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # Başlığı hareket yönüne göre ayarla
        baslik_text = ""
        if self.hareket_yönü == "EKLE":
            baslik_text = "Stok Girişi İşlemi"
            islem_tipleri = ["Giriş (Manuel)", "Sayım Fazlası", "İade Girişi"]
        elif self.hareket_yönü == "EKSILT":
            baslik_text = "Stok Çıkışı İşlemi"
            islem_tipleri = ["Çıkış (Manuel)", "Sayım Eksiği", "Zayiat"]
        else:
            baslik_text = "Stok Hareketi İşlemi" # Varsayılan veya hata durumu
            islem_tipleri = ["Giriş (Manuel)", "Çıkış (Manuel)", "Sayım Fazlası", "Sayım Eksiği", "Zayiat", "İade Girişi"]

        ttk.Label(main_frame, text=f"{baslik_text}\nÜrün: {self.urun_adi}", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10), sticky=tk.W)
        ttk.Label(main_frame, text=f"Mevcut Stok: {self.mevcut_stok:.2f}", font=("Segoe UI", 10)).grid(row=1, column=0, columnspan=2, pady=(0, 15), sticky=tk.W)

        # İşlem Tipi (dinamik olarak ayarlanmış)
        ttk.Label(main_frame, text="İşlem Tipi:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.islem_tipi_combo = ttk.Combobox(main_frame, values=islem_tipleri, state="readonly", width=25)
        self.islem_tipi_combo.grid(row=2, column=1, padx=5, pady=5, sticky=tk.EW)
        self.islem_tipi_combo.set(islem_tipleri[0]) # Varsayılan olarak ilk seçeneği belirle

        # ... (Diğer giriş alanları ve butonlar aynı kalacak) ...
        # Miktar
        ttk.Label(main_frame, text="Miktar:").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        self.miktar_entry = ttk.Entry(main_frame, width=25)
        self.miktar_entry.grid(row=3, column=1, padx=5, pady=5, sticky=tk.EW)
        setup_numeric_entry(self.app, self.miktar_entry, allow_negative=False, decimal_places=2)
        self.miktar_entry.insert(0, "0,00")

        # Tarih
        ttk.Label(main_frame, text="Tarih:").grid(row=4, column=0, padx=5, pady=5, sticky=tk.W)
        self.tarih_entry = ttk.Entry(main_frame, width=20)
        self.tarih_entry.grid(row=4, column=1, padx=5, pady=5, sticky=tk.EW)
        self.tarih_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.tarih_entry)
        ttk.Button(main_frame, text="🗓️", command=lambda: DatePickerDialog(self.app, self.tarih_entry), width=3).grid(row=4, column=2, padx=2, pady=5, sticky=tk.W)

        # Açıklama
        ttk.Label(main_frame, text="Açıklama:").grid(row=5, column=0, padx=5, pady=5, sticky=tk.NW)
        self.aciklama_text = tk.Text(main_frame, height=3, width=25, font=('Segoe UI', 9))
        self.aciklama_text.grid(row=5, column=1, padx=5, pady=5, sticky=tk.EW)

        main_frame.columnconfigure(1, weight=1) # Miktar ve Açıklama Entry'sinin genişlemesi için

        # Butonlar
        button_frame = ttk.Frame(self, padding="10")
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Button(button_frame, text="Kaydet", command=self._kaydet_stok_hareketi, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.RIGHT, padx=5)

    def _kaydet_stok_hareketi(self):
        islem_tipi = self.islem_tipi_combo.get()
        miktar_str = self.miktar_entry.get().strip()
        tarih_str = self.tarih_entry.get().strip()
        aciklama = self.aciklama_text.get("1.0", tk.END).strip()

        if not miktar_str or not tarih_str:
            messagebox.showerror("Eksik Bilgi", "Miktar ve Tarih alanları boş bırakılamaz.", parent=self)
            return

        try:
            miktar = float(miktar_str.replace(',', '.'))
            if miktar <= 0:
                messagebox.showerror("Geçersiz Miktar", "Miktar pozitif bir sayı olmalıdır.", parent=self)
                return
            datetime.strptime(tarih_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Geçersiz Giriş", "Miktar sayısal, Tarih YYYY-AA-GG formatında olmalıdır.", parent=self)
            return

        success, message = self.db.stok_hareketi_ekle(
            self.urun_id,
            islem_tipi,
            miktar,
            tarih_str,
            aciklama
        )

        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            
            self.yenile_stok_listesi_callback() # Ana stok listesini yenile

            print("DEBUG: _kaydet_stok_hareketi - self.urun_karti_penceresi_ref kontrol ediliyor.") # <-- YENİ DEBUG
            if self.urun_karti_penceresi_ref and hasattr(self.urun_karti_penceresi_ref, 'refresh_data_and_ui'):
                print("DEBUG: _kaydet_stok_hareketi - self.urun_karti_penceresi_ref var ve refresh_data_and_ui metodu var. Çağrılıyor.") # <-- YENİ DEBUG
                try:
                    self.urun_karti_penceresi_ref.refresh_data_and_ui() # <-- Bu çağrı doğru olmalı
                    self.urun_karti_penceresi_ref.update_idletasks() # UI güncellemesini zorla
                    self.urun_karti_penceresi_ref.update() # UI güncellemesini daha da zorla
                    if hasattr(self.urun_karti_penceresi_ref, 'entry_stok') and self.urun_karti_penceresi_ref.entry_stok:
                        self.urun_karti_penceresi_ref.entry_stok.focus_set()
                        self.urun_karti_penceresi_ref.entry_stok.selection_range(0, tk.END)
                except Exception as e_update_card:
                    print(f"UYARI: Ürün Kartı penceresi güncellenirken hata oluştu: {e_update_card}")
                    traceback.print_exc() # Detaylı hata çıktısı
            else:
                print("DEBUG: _kaydet_stok_hareketi - self.urun_karti_penceresi_ref yok veya refresh_data_and_ui metodu yok.") # <-- YENİ DEBUG
            
            self.after(50, self.destroy)
    def _load_stok_hareketleri(self, event=None):
        for i in self.stok_hareket_tree.get_children():
            self.stok_hareket_tree.delete(i)

        if not self.urun_id:
            self.stok_hareket_tree.insert("", tk.END, values=("", "", "Ürün Seçili Değil", "", "", "", "", ""))
            return

        islem_tipi_filtre = self.stok_hareket_tip_filter_cb.get()
        bas_tarih_str = self.stok_hareket_bas_tarih_entry.get()
        bit_tarih_str = self.stok_hareket_bit_tarih_entry.get()

        hareketler = self.db.stok_hareketleri_listele(
            self.urun_id,
            islem_tipi=islem_tipi_filtre if islem_tipi_filtre != "TÜMÜ" else None,
            baslangic_tarih=bas_tarih_str if bas_tarih_str else None,
            bitis_tarih=bit_tarih_str if bit_tarih_str else None
        )

        if not hareketler:
            self.stok_hareket_tree.insert("", tk.END, values=("", "", "Hareket Bulunamadı", "", "", "", "", ""))
            return

        for hareket in hareketler:
            # ### HATA DÜZELTMESİ BURADA ###
            # hareket[2] zaten bir tarih nesnesi olduğu için strptime kullanmıyoruz.
            tarih_obj = hareket[2]
            if isinstance(tarih_obj, (datetime, date)):
                tarih_formatted = tarih_obj.strftime('%d.%m.%Y')
            else:
                tarih_formatted = str(tarih_obj) # Beklenmedik bir durum olursa diye
                
            miktar_formatted = f"{hareket[4]:.2f}".rstrip('0').rstrip('.')
            onceki_stok_formatted = f"{hareket[5]:.2f}".rstrip('0').rstrip('.')
            sonraki_stok_formatted = f"{hareket[6]:.2f}".rstrip('0').rstrip('.')
            
            self.stok_hareket_tree.insert("", tk.END, values=(
                hareket[0],
                tarih_formatted,
                hareket[3],
                miktar_formatted,
                onceki_stok_formatted,
                sonraki_stok_formatted,
                hareket[7] if hareket[7] else "-",
                hareket[8] if hareket[8] else "-"
            ))
        self.app.set_status(f"Ürün '{self.urun_adi}' için {len(hareketler)} stok hareketi listelendi.")

class IlgiliFaturalarDetayPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, urun_id, urun_adi):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.urun_id = urun_id
        self.urun_adi = urun_adi
        self.title(f"{self.urun_adi} - İlgili Faturalar")
        self.geometry("1000x600")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text=f"{self.urun_adi} Ürününün Yer Aldığı Faturalar", font=("Segoe UI", 16, "bold")).pack(pady=10, anchor=tk.W, padx=10)

        filter_frame = ttk.Frame(self, padding="5")
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(filter_frame, text="Fatura Tipi:").pack(side=tk.LEFT, padx=(0,2))
        self.fatura_tipi_filter_cb = ttk.Combobox(filter_frame, width=15, values=["TÜMÜ", "ALIŞ", "SATIŞ"], state="readonly")
        self.fatura_tipi_filter_cb.pack(side=tk.LEFT, padx=(0,10))
        self.fatura_tipi_filter_cb.set("TÜMÜ")
        self.fatura_tipi_filter_cb.bind("<<ComboboxSelected>>", self._load_ilgili_faturalar)

        ttk.Button(filter_frame, text="Filtrele", command=self._load_ilgili_faturalar, style="Accent.TButton").pack(side=tk.LEFT)

        cols_fatura = ("ID", "Fatura No", "Tarih", "Tip", "Cari/Misafir", "KDV Hariç Top.", "KDV Dahil Top.")
        self.ilgili_faturalar_tree = ttk.Treeview(self, columns=cols_fatura, show='headings', selectmode="browse")

        col_defs_fatura = [
            ("ID", 40, tk.E, tk.NO),
            ("Fatura No", 120, tk.W, tk.YES),
            ("Tarih", 85, tk.CENTER, tk.NO),
            ("Tip", 70, tk.CENTER, tk.NO),
            ("Cari/Misafir", 200, tk.W, tk.YES),
            ("KDV Hariç Top.", 120, tk.E, tk.NO),
            ("KDV Dahil Top.", 120, tk.E, tk.NO)
        ]
        for cn,w,a,s in col_defs_fatura:
            self.ilgili_faturalar_tree.column(cn, width=w, anchor=a, stretch=s)
            self.ilgili_faturalar_tree.heading(cn, text=cn, command=lambda c=cn: sort_treeview_column(self.ilgili_faturalar_tree, c, False))

        vsb_fatura = ttk.Scrollbar(self, orient="vertical", command=self.ilgili_faturalar_tree.yview)
        hsb_fatura = ttk.Scrollbar(self, orient="horizontal", command=self.ilgili_faturalar_tree.xview)
        self.ilgili_faturalar_tree.configure(yscrollcommand=vsb_fatura.set, xscrollcommand=hsb_fatura.set)
        vsb_fatura.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_fatura.pack(side=tk.BOTTOM, fill=tk.X)
        self.ilgili_faturalar_tree.pack(expand=True, fill=tk.BOTH, padx=10, pady=5)

        self.ilgili_faturalar_tree.bind("<Double-1>", self._on_fatura_double_click)

        self._load_ilgili_faturalar() # İlk yükleme

        ttk.Button(self, text="Kapat", command=self.destroy).pack(pady=10)

    def _load_ilgili_faturalar(self, event=None):
        for i in self.ilgili_faturalar_tree.get_children():
            self.ilgili_faturalar_tree.delete(i)

        if not self.urun_id:
            self.ilgili_faturalar_tree.insert("", tk.END, values=("", "", "", "", "Ürün seçili değil.", "", ""))
            return

        fatura_tipi_filtre = self.fatura_tipi_filter_cb.get()
        
        faturalar = self.db.get_faturalar_by_urun_id(self.urun_id, fatura_tipi=fatura_tipi_filtre)

        if not faturalar:
            self.ilgili_faturalar_tree.insert("", tk.END, values=("", "", "", "", "Bu ürüne ait fatura bulunamadı.", "", ""))
            return

        for fatura_item in faturalar:
            fatura_id = fatura_item[0]
            fatura_no = fatura_item[1]
            tarih_str = fatura_item[2]
            fatura_tip = fatura_item[3]
            cari_adi = fatura_item[4]
            toplam_kdv_haric = fatura_item[5]
            toplam_kdv_dahil = fatura_item[6]

            try:
                formatted_tarih = datetime.strptime(tarih_str, '%Y-%m-%d').strftime('%d.%m.%Y')
            except ValueError:
                formatted_tarih = tarih_str

            self.ilgili_faturalar_tree.insert("", tk.END, iid=fatura_id, values=(
                fatura_id,
                fatura_no,
                formatted_tarih,
                fatura_tip,
                cari_adi,
                self.db._format_currency(toplam_kdv_haric),
                self.db._format_currency(toplam_kdv_dahil)
            ))
        self.app.set_status(f"Ürün '{self.urun_adi}' için {len(faturalar)} fatura listelendi.")


    def _on_fatura_double_click(self, event):
        selected_item_iid = self.ilgili_faturalar_tree.focus()
        if not selected_item_iid:
            return
        
        fatura_id = self.ilgili_faturalar_tree.item(selected_item_iid)['values'][0]
        if fatura_id:
            FaturaDetayPenceresi(self.app, self.db, fatura_id)

class KategoriMarkaYonetimiPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, refresh_callback=None):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.refresh_callback = refresh_callback # Ürün kartı combobox'larını yenilemek için callback
        self.title("Kategori & Marka Yönetimi")
        self.geometry("800x500")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Kategori & Marka Yönetimi", font=("Segoe UI", 16, "bold")).pack(pady=10, anchor=tk.W, padx=10)

        # Ana içerik çerçevesi
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(expand=True, fill=tk.BOTH)
        main_frame.columnconfigure(0, weight=1) # Kategori Frame için
        main_frame.columnconfigure(1, weight=1) # Marka Frame için
        main_frame.rowconfigure(0, weight=1) # Kategori/Marka Frame'ler için

        # Sol taraf: Kategori Yönetimi
        kategori_frame = ttk.LabelFrame(main_frame, text="Kategori Yönetimi", padding="10")
        kategori_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        kategori_frame.columnconfigure(1, weight=1)
        kategori_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(kategori_frame, text="Kategori Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.kategori_entry = ttk.Entry(kategori_frame, width=30)
        self.kategori_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(kategori_frame, text="Ekle", command=self._kategori_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(kategori_frame, text="Güncelle", command=self._kategori_guncelle_ui).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(kategori_frame, text="Sil", command=self._kategori_sil_ui).grid(row=0, column=4, padx=5, pady=5)

        self.kategori_tree = ttk.Treeview(kategori_frame, columns=("ID", "Kategori Adı"), show='headings', selectmode="browse")
        self.kategori_tree.heading("ID", text="ID"); self.kategori_tree.column("ID", width=50, stretch=tk.NO)
        self.kategori_tree.heading("Kategori Adı", text="Kategori Adı"); self.kategori_tree.column("Kategori Adı", width=200, stretch=tk.YES)
        self.kategori_tree.grid(row=1, column=0, columnspan=5, padx=5, pady=10, sticky="nsew")
        self.kategori_tree.bind("<<TreeviewSelect>>", self._on_kategori_select)
        self._kategori_listesini_yukle()


        # Sağ taraf: Marka Yönetimi
        marka_frame = ttk.LabelFrame(main_frame, text="Marka Yönetimi", padding="10")
        marka_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        marka_frame.columnconfigure(1, weight=1)
        marka_frame.grid_rowconfigure(1, weight=1)


        ttk.Label(marka_frame, text="Marka Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.marka_entry = ttk.Entry(marka_frame, width=30)
        self.marka_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(marka_frame, text="Ekle", command=self._marka_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(marka_frame, text="Güncelle", command=self._marka_guncelle_ui).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(marka_frame, text="Sil", command=self._marka_sil_ui).grid(row=0, column=4, padx=5, pady=5)

        self.marka_tree = ttk.Treeview(marka_frame, columns=("ID", "Marka Adı"), show='headings', selectmode="browse")
        self.marka_tree.heading("ID", text="ID"); self.marka_tree.column("ID", width=50, stretch=tk.NO)
        self.marka_tree.heading("Marka Adı", text="Marka Adı"); self.marka_tree.column("Marka Adı", width=200, stretch=tk.YES)
        self.marka_tree.grid(row=1, column=0, columnspan=5, padx=5, pady=10, sticky="nsew")
        self.marka_tree.bind("<<TreeviewSelect>>", self._on_marka_select)
        self._marka_listesini_yukle()

        ttk.Button(self, text="Kapat", command=self.destroy).pack(pady=10)

        # Pencere kapandığında callback'i çağır
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        if self.refresh_callback:
            self.refresh_callback() # Ürün kartı combobox'larını yenile
        self.destroy()

    def _kategori_listesini_yukle(self):
        for i in self.kategori_tree.get_children(): self.kategori_tree.delete(i)
        kategoriler = self.db.kategori_listele()
        for kat_id, kat_ad in kategoriler: self.kategori_tree.insert("", tk.END, values=(kat_id, kat_ad), iid=kat_id)
        # _yukle_kategori_marka_comboboxlari() doğrudan burada çağrılmaz, _on_close ile veya manuel çağrılır.
        # Ürün kartında bağlı combobox'ları yenilemek için App'e bir callback verilecek.

    def _on_kategori_select(self, event):
        selected_item = self.kategori_tree.focus()
        if selected_item:
            values = self.kategori_tree.item(selected_item, 'values')
            self.kategori_entry.delete(0, tk.END)
            self.kategori_entry.insert(0, values[1])
        else:
            self.kategori_entry.delete(0, tk.END)

    def _kategori_ekle_ui(self):
        kategori_adi = self.kategori_entry.get().strip()
        success, message = self.db.kategori_ekle(kategori_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.kategori_entry.delete(0, tk.END)
            self._kategori_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _kategori_guncelle_ui(self):
        selected_item = self.kategori_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir kategori seçin.", parent=self)
            return
        kategori_id = self.kategori_tree.item(selected_item)['values'][0]
        yeni_kategori_adi = self.kategori_entry.get().strip()
        success, message = self.db.kategori_guncelle(kategori_id, yeni_kategori_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.kategori_entry.delete(0, tk.END)
            self._kategori_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _kategori_sil_ui(self):
        selected_item = self.kategori_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir kategori seçin.", parent=self)
            return
        kategori_id = self.kategori_tree.item(selected_item)['values'][0]
        kategori_adi = self.kategori_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{kategori_adi}' kategorisini silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.kategori_sil(kategori_id)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.kategori_entry.delete(0, tk.END)
                self._kategori_listesini_yukle()
            else:
                messagebox.showerror("Hata", message, parent=self)

    def _marka_listesini_yukle(self):
        for i in self.marka_tree.get_children(): self.marka_tree.delete(i)
        markalar = self.db.marka_listele()
        for mar_id, mar_ad in markalar: self.marka_tree.insert("", tk.END, values=(mar_id, mar_ad), iid=mar_id)
        # _yukle_kategori_marka_comboboxlari() doğrudan burada çağrılmaz.

    def _on_marka_select(self, event):
        selected_item = self.marka_tree.focus()
        if selected_item:
            values = self.marka_tree.item(selected_item, 'values')
            self.marka_entry.delete(0, tk.END)
            self.marka_entry.insert(0, values[1])
        else:
            self.marka_entry.delete(0, tk.END)

    def _marka_ekle_ui(self):
        marka_adi = self.marka_entry.get().strip()
        success, message = self.db.marka_ekle(marka_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.marka_entry.delete(0, tk.END)
            self._marka_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _marka_guncelle_ui(self):
        selected_item = self.marka_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir marka seçin.", parent=self)
            return
        marka_id = self.marka_tree.item(selected_item)['values'][0]
        yeni_marka_adi = self.marka_entry.get().strip()
        success, message = self.db.marka_guncelle(marka_id, yeni_marka_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.marka_entry.delete(0, tk.END)
            self._marka_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _marka_sil_ui(self):
        selected_item = self.marka_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir marka seçin.", parent=self)
            return
        marka_id = self.marka_tree.item(selected_item)['values'][0]
        marka_adi = self.marka_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{marka_adi}' markasını silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.marka_sil(marka_id)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.marka_entry.delete(0, tk.END)
                self._marka_listesini_yukle()
            else:
                messagebox.showerror("Hata", message, parent=self)

class UrunNitelikYonetimiPenceresi(tk.Toplevel):
    def __init__(self, parent_notebook, db_manager, app_ref, refresh_callback=None):
        super().__init__(parent_notebook)
        self.db = db_manager
        self.app = app_ref
        self.refresh_callback = refresh_callback

        self.title("Ürün Grubu, Birimi ve Menşe Ülke Yönetimi")
        self.geometry("800x600")
        self.transient(parent_notebook.winfo_toplevel())
        self.grab_set()
        self.resizable(False, False)

        main_frame = self
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=0)

        # --- Ürün Grubu Yönetimi ---
        urun_grubu_frame = ttk.LabelFrame(main_frame, text="Ürün Grubu Yönetimi", padding="10")
        urun_grubu_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        urun_grubu_frame.columnconfigure(1, weight=1)
        urun_grubu_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(urun_grubu_frame, text="Grup Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.urun_grubu_entry = ttk.Entry(urun_grubu_frame, width=30)
        self.urun_grubu_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(urun_grubu_frame, text="Ekle", command=self._urun_grubu_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        # DEĞİŞİKLİK: "Güncelle" butonu kaldırıldı, sil butonu sağa kaydırıldı
        ttk.Button(urun_grubu_frame, text="Sil", command=self._urun_grubu_sil_ui).grid(row=0, column=3, padx=5, pady=5)

        self.urun_grubu_tree = ttk.Treeview(urun_grubu_frame, columns=("ID", "Grup Adı"), show='headings', selectmode="browse")
        self.urun_grubu_tree.heading("ID", text="ID"); self.urun_grubu_tree.column("ID", width=50, stretch=tk.NO)
        self.urun_grubu_tree.heading("Grup Adı", text="Grup Adı"); self.urun_grubu_tree.column("Grup Adı", width=200, stretch=tk.YES)
        # DEĞİŞİKLİK: Columnspan 4 oldu çünkü bir buton kaldırıldı
        self.urun_grubu_tree.grid(row=1, column=0, columnspan=4, padx=5, pady=10, sticky="nsew")
        self.urun_grubu_tree.bind("<<TreeviewSelect>>", self._on_urun_grubu_select)
        self.urun_grubu_tree.bind("<ButtonRelease-3>", self._open_urun_grubu_context_menu) # Sağ tık menüsü
        self._urun_grubu_listesini_yukle()

        # --- Ürün Birimi Yönetimi ---
        urun_birimi_frame = ttk.LabelFrame(main_frame, text="Ürün Birimi Yönetimi", padding="10")
        urun_birimi_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        urun_birimi_frame.columnconfigure(1, weight=1)
        urun_birimi_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(urun_birimi_frame, text="Birim Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.urun_birimi_entry = ttk.Entry(urun_birimi_frame, width=30)
        self.urun_birimi_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(urun_birimi_frame, text="Ekle", command=self._urun_birimi_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        # DEĞİŞİKLİK: "Güncelle" butonu kaldırıldı, sil butonu sağa kaydırıldı
        ttk.Button(urun_birimi_frame, text="Sil", command=self._urun_birimi_sil_ui).grid(row=0, column=3, padx=5, pady=5)

        self.urun_birimi_tree = ttk.Treeview(urun_birimi_frame, columns=("ID", "Birim Adı"), show='headings', selectmode="browse")
        self.urun_birimi_tree.heading("ID", text="ID"); self.urun_birimi_tree.column("ID", width=50, stretch=tk.NO)
        self.urun_birimi_tree.heading("Birim Adı", text="Birim Adı"); self.urun_birimi_tree.column("Birim Adı", width=200, stretch=tk.YES)
        # DEĞİŞİKLİK: Columnspan 4 oldu
        self.urun_birimi_tree.grid(row=1, column=0, columnspan=4, padx=5, pady=10, sticky="nsew")
        self.urun_birimi_tree.bind("<<TreeviewSelect>>", self._on_urun_birimi_select)
        self.urun_birimi_tree.bind("<ButtonRelease-3>", self._open_birim_context_menu) # Sağ tık menüsü
        self._urun_birimi_listesini_yukle()

        # --- Ülke (Menşe) Yönetimi ---
        ulke_frame = ttk.LabelFrame(main_frame, text="Menşe Ülke Yönetimi", padding="10")
        ulke_frame.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
        ulke_frame.columnconfigure(1, weight=1)
        ulke_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(ulke_frame, text="Ülke Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.ulke_entry = ttk.Entry(ulke_frame, width=30)
        self.ulke_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(ulke_frame, text="Ekle", command=self._ulke_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        # DEĞİŞİKLİK: "Güncelle" butonu kaldırıldı, sil butonu sağa kaydırıldı
        ttk.Button(ulke_frame, text="Sil", command=self._ulke_sil_ui).grid(row=0, column=3, padx=5, pady=5)

        self.ulke_tree = ttk.Treeview(ulke_frame, columns=("ID", "Ülke Adı"), show='headings', selectmode="browse")
        self.ulke_tree.heading("ID", text="ID"); self.ulke_tree.column("ID", width=50, stretch=tk.NO)
        self.ulke_tree.heading("Ülke Adı", text="Ülke Adı"); self.ulke_tree.column("Ülke Adı", width=200, stretch=tk.YES)
        # DEĞİŞİKLİK: Columnspan 4 oldu
        self.ulke_tree.grid(row=1, column=0, columnspan=4, padx=5, pady=10, sticky="nsew")
        self.ulke_tree.bind("<<TreeviewSelect>>", self._on_ulke_select)
        self.ulke_tree.bind("<ButtonRelease-3>", self._open_ulke_context_menu) # Sağ tık menüsü
        self._ulke_listesini_yukle()

        ttk.Button(self, text="Kapat", command=self.destroy).grid(row=2, column=0, columnspan=2, pady=10, sticky="se")

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        if self.refresh_callback:
            self.refresh_callback() # Ürün kartı combobox'larını yenile
        self.destroy()

    # Ürün Grubu Yönetimi Metotları
    def _urun_grubu_listesini_yukle(self):
        for i in self.urun_grubu_tree.get_children():
            self.urun_grubu_tree.delete(i)
        urun_gruplari = self.db.urun_grubu_listele()
        for grup_id, grup_ad in urun_gruplari:
            self.urun_grubu_tree.insert("", tk.END, values=(grup_id, grup_ad), iid=grup_id)
        if hasattr(self.app, '_yukle_urun_grubu_birimi_ulke_comboboxlari'):
            self.app._yukle_urun_grubu_birimi_ulke_comboboxlari()

    def _on_urun_grubu_select(self, event):
        selected_item = self.urun_grubu_tree.focus()
        if selected_item:
            values = self.urun_grubu_tree.item(selected_item, 'values')
            self.urun_grubu_entry.delete(0, tk.END)
            self.urun_grubu_entry.insert(0, values[1])
        else:
            self.urun_grubu_entry.delete(0, tk.END)

    def _urun_grubu_ekle_ui(self):
        grup_adi = self.urun_grubu_entry.get().strip()
        if not grup_adi:
            messagebox.showwarning("Uyarı", "Ürün grubu adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_grubu_ekle(grup_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.urun_grubu_entry.delete(0, tk.END)
            self._urun_grubu_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _urun_grubu_guncelle_ui(self):
        selected_item = self.urun_grubu_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir ürün grubu seçin.", parent=self)
            return
        grup_id = self.urun_grubu_tree.item(selected_item)['values'][0]
        yeni_grup_adi = self.urun_grubu_entry.get().strip()
        if not yeni_grup_adi:
            messagebox.showwarning("Uyarı", "Ürün grubu adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_grubu_guncelle(grup_id, yeni_grup_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.urun_grubu_entry.delete(0, tk.END)
            self._urun_grubu_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _urun_grubu_sil_ui(self):
        selected_item = self.urun_grubu_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir ürün grubu seçin.", parent=self)
            return
        grup_id = self.urun_grubu_tree.item(selected_item)['values'][0]
        grup_adi = self.urun_grubu_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{grup_adi}' ürün grubunu silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.urun_grubu_sil(grup_id)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.urun_grubu_entry.delete(0, tk.END)
                self._urun_grubu_listesini_yukle()
            else:
                messagebox.showerror("Hata", message, parent=self)

    # Ürün Birimi Yönetimi Metotları
    def _urun_birimi_listesini_yukle(self):
        for i in self.urun_birimi_tree.get_children():
            self.urun_birimi_tree.delete(i)
        urun_birimleri = self.db.urun_birimi_listele()
        for birim_id, birim_ad in urun_birimleri:
            self.urun_birimi_tree.insert("", tk.END, values=(birim_id, birim_ad), iid=birim_id)
        if hasattr(self.app, '_yukle_urun_grubu_birimi_ulke_comboboxlari'):
            self.app._yukle_urun_grubu_birimi_ulke_comboboxlari()

    def _on_urun_birimi_select(self, event):
        selected_item = self.urun_birimi_tree.focus()
        if selected_item:
            values = self.urun_birimi_tree.item(selected_item, 'values')
            self.urun_birimi_entry.delete(0, tk.END)
            self.urun_birimi_entry.insert(0, values[1])
        else:
            self.urun_birimi_entry.delete(0, tk.END)

    def _urun_birimi_ekle_ui(self):
        birim_adi = self.urun_birimi_entry.get().strip()
        if not birim_adi:
            messagebox.showwarning("Uyarı", "Ürün birimi adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_birimi_ekle(birim_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.urun_birimi_entry.delete(0, tk.END)
            self._urun_birimi_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _urun_birimi_guncelle_ui(self):
        selected_item = self.urun_birimi_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir ürün birimi seçin.", parent=self)
            return
        birim_id = self.urun_birimi_tree.item(selected_item)['values'][0]
        yeni_birim_adi = self.urun_birimi_entry.get().strip()
        if not yeni_birim_adi:
            messagebox.showwarning("Uyarı", "Ürün birimi adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_birimi_guncelle(birim_id, yeni_birim_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.urun_birimi_entry.delete(0, tk.END)
            self._urun_birimi_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _urun_birimi_sil_ui(self):
        selected_item = self.urun_birimi_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir ürün birimi seçin.", parent=self)
            return
        birim_id = self.urun_birimi_tree.item(selected_item)['values'][0]
        birim_adi = self.urun_birimi_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{birim_adi}' ürün birimini silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.urun_birimi_sil(birim_id)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.urun_birimi_entry.delete(0, tk.END)
                self._urun_birimi_listesini_yukle()
            else:
                messagebox.showerror("Hata", message, parent=self)

    def _open_urun_grubu_context_menu(self, event):
        item_id = self.urun_grubu_tree.identify_row(event.y)
        if not item_id: return

        self.urun_grubu_tree.selection_set(item_id)
        grup_id = int(item_id) # iid zaten ID'dir

        context_menu = tk.Menu(self, tearoff=0)
        context_menu.add_command(label="Güncelle", command=lambda: self._urun_grubu_duzenle_popup(grup_id))
        context_menu.add_command(label="Sil", command=self._urun_grubu_sil_ui)

        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def _urun_grubu_duzenle_popup(self, grup_id):
        from pencereler import GrupDuzenlePenceresi # Yeni pop-up sınıfı
        # Grup bilgilerini veritabanından çek
        self.db.c.execute("SELECT id, grup_adi FROM urun_gruplari WHERE id=?", (grup_id,))
        grup_info = self.db.c.fetchone()

        if grup_info:
            GrupDuzenlePenceresi(self, self.db, grup_info, self._urun_grubu_listesini_yukle)
        else:
            messagebox.showerror("Hata", "Ürün grubu bilgisi bulunamadı.", parent=self)
    # DEĞİŞİKLİK BİTİŞİ

    # DEĞİŞİKLİK BAŞLIYOR: Ürün Birimi için sağ tık menüsü metotları (Sizin sağ tık kodunuz)
    def _open_birim_context_menu(self, event):
        item_id = self.urun_birimi_tree.identify_row(event.y)
        if not item_id: return

        self.urun_birimi_tree.selection_set(item_id)
        birim_id = int(item_id)

        context_menu = tk.Menu(self, tearoff=0)
        context_menu.add_command(label="Güncelle", command=lambda: self._urun_birimi_duzenle_popup(birim_id))
        context_menu.add_command(label="Sil", command=self._urun_birimi_sil_ui)

        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def _urun_birimi_duzenle_popup(self, birim_id):
        # Birim bilgilerini veritabanından çek (sadece birim_id ve birim_adi'nı döndüren bir metoda ihtiyacımız var)
        # Bu metot veritabani.py içinde olmalı: urun_birimi_getir_by_id
        self.db.c.execute("SELECT id, birim_adi FROM urun_birimleri WHERE id=?", (birim_id,))
        birim_info = self.db.c.fetchone()

        if birim_info:
            from pencereler import BirimDuzenlePenceresi # Daha önce tanımladığımız sınıf
            BirimDuzenlePenceresi(self, self.db, birim_info, self._urun_birimi_listesini_yukle)
        else:
            messagebox.showerror("Hata", "Ürün birimi bilgisi bulunamadı.", parent=self)
    # DEĞİŞİKLİK BİTİŞİ

    # DEĞİŞİKLİK BAŞLIYOR: Menşe Ülke için sağ tık menüsü metotları
    def _open_ulke_context_menu(self, event):
        item_id = self.ulke_tree.identify_row(event.y)
        if not item_id: return

        self.ulke_tree.selection_set(item_id)
        ulke_id = int(item_id)

        context_menu = tk.Menu(self, tearoff=0)
        context_menu.add_command(label="Güncelle", command=lambda: self._ulke_duzenle_popup(ulke_id))
        context_menu.add_command(label="Sil", command=self._ulke_sil_ui)

        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def _ulke_duzenle_popup(self, ulke_id):
        from pencereler import UlkeDuzenlePenceresi # Yeni pop-up sınıfı
        # Ülke bilgilerini veritabanından çek
        self.db.c.execute("SELECT id, ulke_adi FROM urun_ulkeleri WHERE id=?", (ulke_id,))
        ulke_info = self.db.c.fetchone()

        if ulke_info:
            UlkeDuzenlePenceresi(self, self.db, ulke_info, self._ulke_listesini_yukle)
        else:
            messagebox.showerror("Hata", "Ülke bilgisi bulunamadı.", parent=self)

    # Ülke (Menşe) Yönetimi Metotları
    def _ulke_listesini_yukle(self):
        for i in self.ulke_tree.get_children():
            self.ulke_tree.delete(i)
        ulkeler = self.db.ulke_listele()
        for ulke_id, ulke_ad in ulkeler:
            self.ulke_tree.insert("", tk.END, values=(ulke_id, ulke_ad), iid=ulke_id)
        if hasattr(self.app, '_yukle_urun_grubu_birimi_ulke_comboboxlari'):
            self.app._yukle_urun_grubu_birimi_ulke_comboboxlari()

    def _on_ulke_select(self, event):
        selected_item = self.ulke_tree.focus()
        if selected_item:
            values = self.ulke_tree.item(selected_item, 'values')
            self.ulke_entry.delete(0, tk.END)
            self.ulke_entry.insert(0, values[1])
        else:
            self.ulke_entry.delete(0, tk.END)

    def _ulke_ekle_ui(self):
        ulke_adi = self.ulke_entry.get().strip()
        if not ulke_adi:
            messagebox.showwarning("Uyarı", "Ülke adı boş olamaz.", parent=self)
            return
        success, message = self.db.ulke_ekle(ulke_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.ulke_entry.delete(0, tk.END)
            self._ulke_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _ulke_guncelle_ui(self):
        selected_item = self.ulke_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir ülke seçin.", parent=self)
            return
        ulke_id = self.ulke_tree.item(selected_item)['values'][0]
        yeni_ulke_adi = self.ulke_entry.get().strip()
        if not yeni_ulke_adi:
            messagebox.showwarning("Uyarı", "Ülke adı boş olamaz.", parent=self)
            return
        success, message = self.db.ulke_guncelle(ulke_id, yeni_ulke_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.ulke_entry.delete(0, tk.END)
            self._ulke_listesini_yukle()
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _ulke_sil_ui(self):
        selected_item = self.ulke_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir ülke seçin.", parent=self)
            return
        ulke_id = self.ulke_tree.item(selected_item)['values'][0]
        ulke_adi = self.ulke_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{ulke_adi}' ülkesini silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.ulke_sil(ulke_id)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.ulke_entry.delete(0, tk.END)
                self._ulke_listesini_yukle()
            else:
                messagebox.showerror("Hata", message, parent=self)

    # Ortak ComboBox Yükleme Metotları
    def _yukle_kategori_marka_comboboxlari(self):
        # Kategoriler
        kategoriler = self.db.kategori_listele()
        self.kategoriler_map = {"Seçim Yok": None}
        kategori_display_values = ["Seçim Yok"]
        for k_id, k_ad in kategoriler:
            self.kategoriler_map[k_ad] = k_id
            kategori_display_values.append(k_ad)
        self.combo_kategori['values'] = kategori_display_values
        if self.urun_duzenle and self.urun_detaylari[22]: # kategori_id'nin indeksi 22
            kategori_adi = self.db.kategori_getir_by_id(self.urun_detaylari[22])
            if kategori_adi: self.combo_kategori.set(kategori_adi[1])
            else: self.combo_kategori.set("Seçim Yok")
        else:
            self.combo_kategori.set("Seçim Yok")

        # Markalar
        markalar = self.db.marka_listele()
        self.markalar_map = {"Seçim Yok": None}
        marka_display_values = ["Seçim Yok"]
        for m_id, m_ad in markalar:
            self.markalar_map[m_ad] = m_id
            marka_display_values.append(m_ad)
        self.combo_marka['values'] = marka_display_values
        if self.urun_duzenle and self.urun_detaylari[23]: # marka_id'nin indeksi 23
            marka_adi = self.db.marka_getir_by_id(self.urun_detaylari[23])
            if marka_adi: self.combo_marka.set(marka_adi[1])
            else: self.combo_marka.set("Seçim Yok")
        else:
            self.combo_marka.set("Seçim Yok")

    def _yukle_urun_grubu_birimi_ulke_comboboxlari(self):
        # Ürün Grupları
        urun_gruplari = self.db.urun_grubu_listele()
        self.urun_gruplari_map = {"Seçim Yok": None}
        urun_grubu_display_values = ["Seçim Yok"]
        for g_id, g_ad in urun_gruplari:
            self.urun_gruplari_map[g_ad] = g_id
            urun_grubu_display_values.append(g_ad)

        self.combo_urun_grubu['values'] = urun_grubu_display_values
        if self.urun_duzenle and self.urun_duzenle[24] is not None: # urun_grubu_id'nin indeksi 24
            grup_adi_tuple = self.db.urun_grubu_getir_by_id(self.urun_duzenle[24])
            if grup_adi_tuple and grup_adi_tuple[1] in urun_grubu_display_values: # Grup adı listede varsa
                self.combo_urun_grubu.set(grup_adi_tuple[1])
            else:
                self.combo_urun_grubu.set("Seçim Yok")
        else:
            self.combo_urun_grubu.set("Seçim Yok")

        # Ürün Birimleri
        urun_birimleri = self.db.urun_birimi_listele()
        self.urun_birimleri_map = {"Seçim Yok": None} # <-- DÜZELTME: urun_birimileri_map yerine urun_birimleri_map
        urun_birimi_display_values = ["Seçim Yok"]
        for b_id, b_ad in urun_birimleri:
            self.urun_birimleri_map[b_ad] = b_id
            urun_birimi_display_values.append(b_ad)

        self.combo_urun_birimi['values'] = urun_birimi_display_values
        if self.urun_duzenle and self.urun_duzenle[25] is not None: # urun_birimi_id'nin indeksi 25
            birim_adi_tuple = self.db.urun_birimi_getir_by_id(self.urun_duzenle[25])
            if birim_adi_tuple and birim_adi_tuple[1] in urun_birimi_display_values: # Birim adı listede varsa
                self.combo_urun_birimi.set(birim_adi_tuple[1])
            else:
                self.combo_urun_birimi.set("Seçim Yok")
        else:
            self.combo_urun_birimi.set("Seçim Yok")

        # Ülkeler (Menşe)
            ulkeler = self.db.ulke_listele()
        self.ulkeler_map = {"Seçim Yok": None}
        ulke_display_values = ["Seçim Yok"]
        for u_id, u_ad in ulkeler:
            self.ulkeler_map[u_ad] = u_id
            ulke_display_values.append(u_ad)

        self.combo_mense['values'] = ulke_display_values
        if self.urun_duzenle and self.urun_duzenle[26] is not None: # ulke_id'nin indeksi 26
            ulke_adi_tuple = self.db.ulke_getir_by_id(self.urun_duzenle[26])
            if ulke_adi_tuple and ulke_adi_tuple[1] in ulke_display_values: # Ülke adı listede varsa
                self.combo_mense.set(ulke_adi_tuple[1])
            else:
                self.combo_mense.set("Seçim Yok")
        else:
            self.combo_mense.set("Seçim Yok")

class UrunKartiPenceresi(tk.Toplevel):
    def __init__(self, parent, db_manager, yenile_callback, urun_duzenle=None, app_ref=None, on_update_reopen_callback=None):
        super().__init__(parent)
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.urun_duzenle = urun_duzenle
        self.app = app_ref
        self.title("Ürün Kartı" if urun_duzenle is None else "Ürün Düzenle")
        self.geometry("950x750")
        self.transient(parent)
        self.grab_set()
        self.resizable(True, True)

        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=0)
        self.grid_columnconfigure(0, weight=1)

        self.sv_kod = tk.StringVar(self)
        self.sv_ad = tk.StringVar(self)
        self.sv_kdv = tk.StringVar(self)
        self.sv_alis_haric = tk.StringVar(self)
        self.sv_alis_dahil = tk.StringVar(self)
        self.sv_satis_haric = tk.StringVar(self)
        self.sv_satis_dahil = tk.StringVar(self)
        self.sv_stok = tk.StringVar(self)
        self.sv_min_stok = tk.StringVar(self)

        self.entry_kod = None
        self.entry_ad = None
        self.entry_urun_detayi = None
        self.entry_kdv = None
        self.entry_alis_haric = None
        self.entry_alis_dahil = None
        self.entry_satis_haric = None
        self.entry_satis_dahil = None
        self.label_kar_orani = None
        self.fiyat_degisiklik_tarihi_label = None
        self.urun_resmi_path = ""
        self.original_image = None
        self.tk_image = None
        self._last_resized_size = (0, 0)
        self.urun_resmi_label = None
        self.image_container_frame = None

        self.entry_stok = None
        self.entry_min_stok = None
        self.combo_kategori = None
        self.combo_marka = None
        self.combo_urun_grubu = None
        self.combo_urun_birimi = None
        self.combo_mense = None

        self.kategoriler_map = {"Seçim Yok": None}
        self.markalar_map = {"Seçim Yok": None}
        self.urun_gruplari_map = {"Seçim Yok": None}
        self.urun_birimleri_map = {"Seçim Yok": None}
        self.ulkeler_map = {"Seçim Yok": None}

        ttk.Label(self, text=self.title(), font=("Segoe UI", 16, "bold")).grid(row=0, column=0, pady=5, sticky="ew")

        self.main_notebook = ttk.Notebook(self)
        self.main_notebook.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

        self.genel_bilgiler_sekmesi_frame = ttk.Frame(self.main_notebook, padding="5")
        self.main_notebook.add(self.genel_bilgiler_sekmesi_frame, text="Genel Bilgiler")

        self.urun_gorsel_ve_operasyon_frame = ttk.Frame(self.genel_bilgiler_sekmesi_frame)
        self.urun_gorsel_ve_operasyon_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

        self._setup_genel_bilgiler_tab(self.genel_bilgiler_sekmesi_frame)

        self.urun_id = self.urun_duzenle[0] if self.urun_duzenle else None
        self.urun_adi_initial = self.urun_duzenle[2] if self.urun_duzenle else "Yeni Ürün"

        # DÜZELTME BAŞLANGICI: StokHareketleriSekmesi'ne 'parent_pencere=self' gönderiyoruz.
        # arayuz.py dosyasından doğru sınıfları import ettiğinizden emin olun.
        from arayuz import StokHareketleriSekmesi, IlgiliFaturalarSekmesi, KategoriMarkaYonetimiSekmesi
        self.stok_hareketleri_sekmesi_frame = StokHareketleriSekmesi(
            self.main_notebook, # parent_notebook
            self.db,
            self.app,
            self.urun_id,
            self.urun_adi_initial,
            parent_pencere=self # <-- Burası kritik düzeltme! UrunKartiPenceresi'nin kendisini gönderiyoruz.
        )
        self.main_notebook.add(self.stok_hareketleri_sekmesi_frame, text="Stok Hareketleri")
        # DÜZELTME BİTİŞİ

        self.ilgili_faturalar_sekmesi_frame = IlgiliFaturalarSekmesi(self.main_notebook, self.db, self.app, self.urun_id, self.urun_adi_initial)
        self.main_notebook.add(self.ilgili_faturalar_sekmesi_frame, text="İlgili Faturalar")

        self.kategori_marka_yonetimi_sekmesi_frame = KategoriMarkaYonetimiSekmesi(self.main_notebook, self.db, self.app)
        self.main_notebook.add(self.kategori_marka_yonetimi_sekmesi_frame, text="Kategori & Marka Yönetimi")

        self.main_notebook.bind("<<NotebookTabChanged>>", self._on_notebook_tab_change)

        bottom_main_buttons_frame = ttk.Frame(self, padding="5")
        bottom_main_buttons_frame.grid(row=2, column=0, sticky="ew", pady=(0, 5), padx=5)

        self.btn_kaydet = ttk.Button(bottom_main_buttons_frame, text="Kaydet", command=self.kaydet, style="Accent.TButton")
        self.btn_kaydet.pack(side=tk.LEFT, padx=2)

        self.btn_sil = ttk.Button(bottom_main_buttons_frame, text="Sil", command=self._urun_sil_butonu)
        self.btn_sil.pack(side=tk.LEFT, padx=2)

        ttk.Button(bottom_main_buttons_frame, text="Kapat", command=self.destroy).pack(side=tk.RIGHT, padx=2)

        if self.urun_duzenle:
            self.urun_detaylari = self.urun_duzenle
            self._load_genel_bilgiler()
            self.btn_sil.config(state=tk.NORMAL)
        else:
            self.urun_detaylari = None
            self.sv_kod.set(self.db.get_next_stok_kodu())
            self.sv_ad.set("")
            self.entry_urun_detayi.delete("1.0", tk.END)
            self.sv_kdv.set("20")
            self.sv_alis_haric.set("0,00")
            self.sv_alis_dahil.set("0,00")
            self.sv_satis_haric.set("0,00")
            self.sv_satis_dahil.set("0,00")
            self.sv_stok.set("0,00")
            self.sv_min_stok.set("0,00")
            self._yukle_kategori_marka_comboboxlari()
            self._yukle_urun_grubu_birimi_ulke_comboboxlari()
            self.urun_resmi_path = ""
            if self.urun_resmi_label:
                self.urun_resmi_label.config(text="Resim Yok", image='')
            self.original_image = None
            self.tk_image = None
            self._last_resized_size = (0,0)
            self.btn_sil.config(state=tk.DISABLED)

        self.after(100, self.entry_kod.focus_set)

    def refresh_data_and_ui(self):
        """
        Ürüne ait en güncel verileri veritabanından çeker ve tüm arayüzü yeniler.
        Bu metot, alt pencerelerden (Stok Hareketi gibi) gelen sinyaller üzerine çağrılır.
        """
        print("DEBUG: UrunKartiPenceresi.refresh_data_and_ui çağrıldı.")
        if not self.urun_id: # ürün ID'si yoksa işlem yapma
            return

        # Veritabanından en güncel ürün verisini çek
        latest_product_data = self.db.stok_getir_by_id(self.urun_id)

        if latest_product_data:
            self.urun_duzenle = latest_product_data # Pencerenin ana veri kaynağını güncelle
            self._load_genel_bilgiler() # Arayüzü bu yeni veriyle doldur

            # UI'ın kendini hemen yenilemesini sağlamak için
            self.update_idletasks()
            # Alternatif olarak: self.update() de kullanılabilir.
            print("DEBUG: Ürün kartı arayüzü en güncel verilerle yenilendi.")
        else:
            print("UYARI: Ürün kartı yenilenirken ürün veritabanından bulunamadı.")
            messagebox.showwarning("Veri Kayıp", "Ürün verileri bulunamadığı için kart yenilenemedi.", parent=self)

    def _setup_urun_nitelik_yonetim_tab(self, parent_frame):
        parent_frame.columnconfigure(0, weight=1)
        parent_frame.columnconfigure(1, weight=1)
        parent_frame.rowconfigure(0, weight=1)

        urun_grubu_frame = ttk.LabelFrame(parent_frame, text="Ürün Grubu Yönetimi", padding="10")
        urun_grubu_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        urun_grubu_frame.columnconfigure(1, weight=1)
        urun_grubu_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(urun_grubu_frame, text="Grup Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.urun_grubu_entry = ttk.Entry(urun_grubu_frame, width=30)
        self.urun_grubu_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(urun_grubu_frame, text="Ekle", command=self._urun_grubu_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(urun_grubu_frame, text="Güncelle", command=self._urun_grubu_guncelle_ui).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(urun_grubu_frame, text="Sil", command=self._urun_grubu_sil_ui).grid(row=0, column=4, padx=5, pady=5)

        self.urun_grubu_tree = ttk.Treeview(urun_grubu_frame, columns=("ID", "Grup Adı"), show='headings', selectmode="browse")
        self.urun_grubu_tree.heading("ID", text="ID"); self.urun_grubu_tree.column("ID", width=50, stretch=tk.NO)
        self.urun_grubu_tree.heading("Grup Adı", text="Grup Adı"); self.urun_grubu_tree.column("Grup Adı", width=200, stretch=tk.YES)
        self.urun_grubu_tree.grid(row=1, column=0, columnspan=5, padx=5, pady=10, sticky="nsew")
        self.urun_grubu_tree.bind("<<TreeviewSelect>>", self._on_urun_grubu_select)
        self._urun_grubu_listesini_yukle()

        urun_birimi_frame = ttk.LabelFrame(parent_frame, text="Ürün Birimi Yönetimi", padding="10")
        urun_birimi_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        urun_birimi_frame.columnconfigure(1, weight=1)
        urun_birimi_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(urun_birimi_frame, text="Birim Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.urun_birimi_entry = ttk.Entry(urun_birimi_frame, width=30)
        self.urun_birimi_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(urun_birimi_frame, text="Ekle", command=self._urun_birimi_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(urun_birimi_frame, text="Güncelle", command=self._urun_birimi_guncelle_ui).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(urun_birimi_frame, text="Sil", command=self._urun_birimi_sil_ui).grid(row=0, column=4, padx=5, pady=5)

        self.urun_birimi_tree = ttk.Treeview(urun_birimi_frame, columns=("ID", "Birim Adı"), show='headings', selectmode="browse")
        self.urun_birimi_tree.heading("ID", text="ID"); self.urun_birimi_tree.column("ID", width=50, stretch=tk.NO)
        self.urun_birimi_tree.heading("Birim Adı", text="Birim Adı"); self.urun_birimi_tree.column("Birim Adı", width=200, stretch=tk.YES)
        self.urun_birimi_tree.grid(row=1, column=0, columnspan=5, padx=5, pady=10, sticky="nsew")
        self.urun_birimi_tree.bind("<<TreeviewSelect>>", self._on_urun_birimi_select)
        self._urun_birimi_listesini_yukle()

        ulke_frame = ttk.LabelFrame(parent_frame, text="Menşe Ülke Yönetimi", padding="10")
        ulke_frame.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
        ulke_frame.columnconfigure(1, weight=1)
        ulke_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(ulke_frame, text="Ülke Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.ulke_entry = ttk.Entry(ulke_frame, width=30)
        self.ulke_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(ulke_frame, text="Ekle", command=self._ulke_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(ulke_frame, text="Güncelle", command=self._ulke_guncelle_ui).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(ulke_frame, text="Sil", command=self._ulke_sil_ui).grid(row=0, column=4, padx=5, pady=5)

        self.ulke_tree = ttk.Treeview(ulke_frame, columns=("ID", "Ülke Adı"), show='headings', selectmode="browse")
        self.ulke_tree.heading("ID", text="ID"); self.ulke_tree.column("ID", width=50, stretch=tk.NO)
        self.ulke_tree.heading("Ülke Adı", text="Ülke Adı"); self.ulke_tree.column("Ülke Adı", width=200, stretch=tk.YES)
        self.ulke_tree.grid(row=1, column=0, columnspan=5, padx=5, pady=10, sticky="nsew")
        self.ulke_tree.bind("<<TreeviewSelect>>", self._on_ulke_select)
        self._ulke_listesini_yukle()

    def _urun_grubu_listesini_yukle(self):
        for i in self.urun_grubu_tree.get_children(): self.urun_grubu_tree.delete(i)
        urun_gruplari = self.db.urun_grubu_listele()
        for grup in urun_gruplari: self.urun_grubu_tree.insert("", tk.END, values=grup, iid=grup[0])
        self._yukle_urun_grubu_birimi_ulke_comboboxlari() # Bağlantılı combobox'ı da yenile

    def _setup_tabs(self):
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(expand=True, fill=tk.BOTH, padx=10, pady=(0, 10))

        # Genel Bilgiler Sekmesi
        self._setup_genel_bilgiler_tab()

        # Stok Hareketleri Sekmesi
        # StokHareketleriSekmesi'ne, kendi sahibi olan pencereyi (self) parametre olarak veriyoruz.
        from arayuz import StokHareketleriSekmesi # Yerel içe aktarma
        self.stok_hareketleri_frame = StokHareketleriSekmesi(
            self.notebook, 
            self.db, 
            self.app, 
            self.urun_id, 
            self.urun_duzenle['urun_adi'] if self.urun_duzenle else "Yeni Ürün",
            parent_pencere=self 
        )
        self.notebook.add(self.stok_hareketleri_frame, text="Stok Hareketleri")

        self.notebook.bind("<<NotebookTabChanged>>", self._on_notebook_tab_change)

    def _on_urun_grubu_select(self, event):
        selected_item = self.urun_grubu_tree.focus()
        if selected_item:
            values = self.urun_grubu_tree.item(selected_item, 'values')
            self.urun_grubu_entry.delete(0, tk.END)
            self.urun_grubu_entry.insert(0, values[1])
        else:
            self.urun_grubu_entry.delete(0, tk.END)

    def _urun_grubu_ekle_ui(self):
        grup_adi = self.urun_grubu_entry.get().strip()
        if not grup_adi:
            messagebox.showwarning("Uyarı", "Ürün grubu adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_grubu_ekle(grup_adi)
        if success:
            messagebox.showinfo("Başarılı", f"'{grup_adi}' ürün grubu başarıyla eklendi.", parent=self)
            self.urun_grubu_entry.delete(0, tk.END)
            self._urun_grubu_listesini_yukle()
        else:
            messagebox.showerror("Hata", f"Ürün grubu eklenirken hata: {message}", parent=self)

    def _urun_grubu_guncelle_ui(self):
        selected_item = self.urun_grubu_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir ürün grubu seçin.", parent=self)
            return
        grup_id = self.urun_grubu_tree.item(selected_item)['values'][0]
        yeni_grup_adi = self.urun_grubu_entry.get().strip()
        if not yeni_grup_adi:
            messagebox.showwarning("Uyarı", "Ürün grubu adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_grubu_guncelle(grup_id, yeni_grup_adi)
        if success:
            messagebox.showinfo("Başarılı", f"'{yeni_grup_adi}' ürün grubu başarıyla güncellendi.", parent=self)
            self.urun_grubu_entry.delete(0, tk.END)
            self._urun_grubu_listesini_yukle()
        else:
            messagebox.showerror("Hata", f"Ürün grubu güncellenirken hata: {message}", parent=self)

    def _urun_grubu_sil_ui(self):
        selected_item = self.urun_grubu_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir ürün grubu seçin.", parent=self)
            return
        grup_id = self.urun_grubu_tree.item(selected_item)['values'][0]
        grup_adi = self.urun_grubu_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{grup_adi}' ürün grubunu silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.urun_grubu_sil(grup_id)
            if success:
                messagebox.showinfo("Başarılı", f"'{grup_adi}' ürün grubu başarıyla silindi.", parent=self)
                self.urun_grubu_entry.delete(0, tk.END)
                self._urun_grubu_listesini_yukle()
            else:
                messagebox.showerror("Hata", f"Ürün grubu silinirken hata: {message}\nBu gruba bağlı ürünler olabilir.", parent=self)

    def _on_urun_birimi_select(self, event):
        selected_item = self.urun_birimi_tree.focus()
        if selected_item:
            values = self.urun_birimi_tree.item(selected_item, 'values')
            self.urun_birimi_entry.delete(0, tk.END)
            self.urun_birimi_entry.insert(0, values[1])
        else:
            self.urun_birimi_entry.delete(0, tk.END)

    def _urun_birimi_ekle_ui(self):
        birim_adi = self.urun_birimi_entry.get().strip()
        if not birim_adi:
            messagebox.showwarning("Uyarı", "Ürün birimi adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_birimi_ekle(birim_adi)
        if success:
            messagebox.showinfo("Başarılı", f"'{birim_adi}' ürün birimi başarıyla eklendi.", parent=self)
            self.urun_birimi_entry.delete(0, tk.END)
            self._urun_birimi_listesini_yukle()
        else:
            messagebox.showerror("Hata", f"Ürün birimi eklenirken hata: {message}", parent=self)            

    def _urun_birimi_guncelle_ui(self):
        selected_item = self.urun_birimi_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir ürün birimi seçin.", parent=self)
            return
        birim_id = self.urun_birimi_tree.item(selected_item)['values'][0]
        yeni_birim_adi = self.urun_birimi_entry.get().strip()
        if not yeni_birim_adi:
            messagebox.showwarning("Uyarı", "Ürün birimi adı boş olamaz.", parent=self)
            return
        success, message = self.db.urun_birimi_guncelle(birim_id, yeni_birim_adi)
        if success:
            messagebox.showinfo("Başarılı", f"'{yeni_birim_adi}' ürün birimi başarıyla güncellendi.", parent=self)
            self.urun_birimi_entry.delete(0, tk.END)
            self._urun_birimi_listesini_yukle()
        else:
            messagebox.showerror("Hata", f"Ürün birimi güncellenirken hata: {message}", parent=self)

    def _urun_birimi_sil_ui(self):
        selected_item = self.urun_birimi_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir ürün birimi seçin.", parent=self)
            return
        birim_id = self.urun_birimi_tree.item(selected_item)['values'][0]
        birim_adi = self.urun_birimi_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{birim_adi}' ürün birimini silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.urun_birimi_sil(birim_id)
            if success:
                messagebox.showinfo("Başarılı", f"'{birim_adi}' ürün birimi başarıyla silindi.", parent=self)
                self.urun_birimi_entry.delete(0, tk.END)
                self._urun_birimi_listesini_yukle()
            else:
                messagebox.showerror("Hata", f"Ürün birimi silinirken hata: {message}\nBu birime bağlı ürünler olabilir.", parent=self)

    # Ülke (Menşe) Yönetimi Metotları (UrunKartiPenceresi içinde)
    def _ulke_listesini_yukle(self):
        for i in self.ulke_tree.get_children(): self.ulke_tree.delete(i)
        ulkeler = self.db.ulke_listele()
        for ulke in ulkeler: self.ulke_tree.insert("", tk.END, values=ulke, iid=ulke[0])
        self._yukle_urun_grubu_birimi_ulke_comboboxlari() # Bağlantılı combobox'ı da yenile

    def _on_ulke_select(self, event):
        selected_item = self.ulke_tree.focus()
        if selected_item:
            values = self.ulke_tree.item(selected_item, 'values')
            self.ulke_entry.delete(0, tk.END)
            self.ulke_entry.insert(0, values[1])
        else:
            self.ulke_entry.delete(0, tk.END)

    def _ulke_ekle_ui(self):
        ulke_adi = self.ulke_entry.get().strip()
        if not ulke_adi:
            messagebox.showwarning("Uyarı", "Ülke adı boş olamaz.", parent=self)
            return
        success, message = self.db.ulke_ekle(ulke_adi)
        if success:
            messagebox.showinfo("Başarılı", f"'{ulke_adi}' ülkesi başarıyla eklendi.", parent=self)
            self.ulke_entry.delete(0, tk.END)
            self._ulke_listesini_yukle()
        else:
            messagebox.showerror("Hata", f"Ülke eklenirken hata: {message}", parent=self)

    def _ulke_guncelle_ui(self):
        selected_item = self.ulke_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir ülke seçin.", parent=self)
            return
        ulke_id = self.ulke_tree.item(selected_item)['values'][0]
        yeni_ulke_adi = self.ulke_entry.get().strip()
        if not yeni_ulke_adi:
            messagebox.showwarning("Uyarı", "Ülke adı boş olamaz.", parent=self)
            return
        success, message = self.db.ulke_guncelle(ulke_id, yeni_ulke_adi)
        if success:
            messagebox.showinfo("Başarılı", f"'{yeni_ulke_adi}' ülkesi başarıyla güncellendi.", parent=self)
            self.ulke_entry.delete(0, tk.END)
            self._ulke_listesini_yukle()
        else:
            messagebox.showerror("Hata", f"Ülke güncellenirken hata: {message}", parent=self)

    def _ulke_sil_ui(self):
        selected_item = self.ulke_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir ülke seçin.", parent=self)
            return
        ulke_id = self.ulke_tree.item(selected_item)['values'][0]
        ulke_adi = self.ulke_tree.item(selected_item)['values'][1]
        if messagebox.askyesno("Onay", f"'{ulke_adi}' ülkesini silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.ulke_sil(ulke_id)
            if success:
                messagebox.showinfo("Başarılı", f"'{ulke_adi}' ülkesi başarıyla silindi.", parent=self)
                self.ulke_entry.delete(0, tk.END)
                self._ulke_listesini_yukle()
            else:
                messagebox.showerror("Hata", f"Ülke silinirken hata: {message}\nBu ülkeye bağlı ürünler olabilir.", parent=self)


    def _urun_birimi_listesini_yukle(self):
        for i in self.urun_birimi_tree.get_children(): self.urun_birimi_tree.delete(i)
        urun_birimleri = self.db.urun_birimi_listele()
        for birim in urun_birimleri: self.urun_birimi_tree.insert("", tk.END, values=birim, iid=birim[0])
        self._yukle_urun_grubu_birimi_ulke_comboboxlari() # Bağlantılı combobox'ı da yenile

    def _urun_sil_butonu(self):
        """Ürün Kartından doğrudan ürün silme işlemini çağırır."""
        if self.urun_id:
            urun_adi = self.entry_ad.get()
            if messagebox.askyesno("Ürün Silme Onayı", f"'{urun_adi}' adlı ürünü silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz.", parent=self.app):
                success, message = self.db.stok_sil(self.urun_id) 
                if success:
                    messagebox.showinfo("Başarılı", message, parent=self.app)
                    self.yenile_callback()
                    self.destroy()
                    self.app.set_status(f"'{urun_adi}' ürünü silindi.")
                else:
                    messagebox.showerror("Hata", message, parent=self.app)
        else:
            messagebox.showwarning("Uyarı", "Bu işlem sadece mevcut bir ürünü düzenlerken kullanılabilir.", parent=self)

    def _yukle_urun_grubu_birimi_ulke_comboboxlari(self):
        # Verileri DB'den al
        urun_gruplari_map = self.db.get_urun_gruplari_for_combobox()
        urun_birimleri_map = self.db.get_urun_birimleri_for_combobox()
        ulkeler_map = self.db.get_ulkeler_for_combobox()

        # Combobox'ları doldurma
        self.urun_gruplari_map = {"Seçim Yok": None, **urun_gruplari_map}
        self.combo_urun_grubu['values'] = ["Seçim Yok"] + sorted(urun_gruplari_map.keys())

        self.urun_birimleri_map = {"Seçim Yok": None, **urun_birimleri_map}
        self.combo_urun_birimi['values'] = ["Seçim Yok"] + sorted(urun_birimleri_map.keys())

        self.ulkeler_map = {"Seçim Yok": None, **ulkeler_map}
        self.combo_mense['values'] = ["Seçim Yok"] + sorted(ulkeler_map.keys())

        # Seçili değerleri ayarla (eğer ürün düzenleniyorsa)
        if self.urun_duzenle:
            urun_grubu_adi = self.urun_duzenle[19] # Ürün Grubu Adı
            urun_birimi_adi = self.urun_duzenle[20] # Ürün Birimi Adı
            ulke_adi = self.urun_duzenle[21] # Ülke Adı
            self.combo_urun_grubu.set(urun_grubu_adi if urun_grubu_adi in self.urun_gruplari_map else "Seçim Yok")
            self.combo_urun_birimi.set(urun_birimi_adi if urun_birimi_adi in self.urun_birimleri_map else "Seçim Yok")
            self.combo_mense.set(ulke_adi if ulke_adi in self.ulkeler_map else "Seçim Yok")
        else:
            self.combo_urun_grubu.set("Seçim Yok")
            self.combo_urun_birimi.set("Seçim Yok")
            self.combo_mense.set("Seçim Yok")

    def _load_stok_hareketleri(self, event=None):
        """Stok hareketleri Treeview'ini ürün ID'sine göre doldurur."""
        for i in self.stok_hareket_tree.get_children():
            self.stok_hareket_tree.delete(i)

        if not self.urun_id:
            self.stok_hareket_tree.insert("", tk.END, values=("", "", "Ürün Seçili Değil", "", "", "", "", ""))
            return

        islem_tipi_filtre = self.stok_hareket_tip_filter_cb.get()
        bas_tarih_str = self.stok_hareket_bas_tarih_entry.get()
        bit_tarih_str = self.stok_hareket_bit_tarih_entry.get()

        # Veritabanından stok hareketlerini çek
        # db.stok_hareketleri_listele metodu bu filtreleri almalı
        hareketler = self.db.stok_hareketleri_listele(
            self.urun_id,
            islem_tipi=islem_tipi_filtre if islem_tipi_filtre != "TÜMÜ" else None,
            baslangic_tarih=bas_tarih_str if bas_tarih_str else None,
            bitis_tarih=bit_tarih_str if bit_tarih_str else None
        )

        if not hareketler:
            self.stok_hareket_tree.insert("", tk.END, values=("", "", "Hareket Bulunamadı", "", "", "", "", ""))
            return

        for hareket in hareketler:
            # hareket: (id, urun_id, tarih, islem_tipi, miktar, onceki_stok, sonraki_stok, aciklama, kaynak)
            tarih_formatted = datetime.strptime(hareket[2], '%Y-%m-%d').strftime('%d.%m.%Y')
            miktar_formatted = f"{hareket[4]:.2f}".rstrip('0').rstrip('.')
            onceki_stok_formatted = f"{hareket[5]:.2f}".rstrip('0').rstrip('.')
            sonraki_stok_formatted = f"{hareket[6]:.2f}".rstrip('0').rstrip('.')
            
            self.stok_hareket_tree.insert("", tk.END, values=(
                hareket[0], # ID
                tarih_formatted, # Tarih
                hareket[3], # İşlem Tipi
                miktar_formatted, # Miktar
                onceki_stok_formatted, # Önceki Stok
                sonraki_stok_formatted, # Sonraki Stok
                hareket[7] if hareket[7] else "-", # Açıklama
                hareket[8] if hareket[8] else "-" # Kaynak
            ))
        self.app.set_status(f"Ürün '{self.urun_adi_initial}' için {len(hareketler)} stok hareketi listelendi.")


    def _stok_ekle_penceresi_ac(self):
        """Stok ekleme penceresini 'EKLE' yönüyle açar."""
        if not self.urun_id:
            messagebox.showwarning("Uyarı", "Lütfen işlem yapmak için bir ürün seçin.", parent=self)
            return

        urun_guncel_bilgi = self.db.stok_getir_by_id(self.urun_id)
        if urun_guncel_bilgi:
            mevcut_stok = urun_guncel_bilgi[3]

            stok_hareketi_popup = StokHareketiPenceresi(
                self.app, # parent_app
                self.db,
                self.urun_id,
                self.urun_detaylari[2], # urun_adi
                mevcut_stok, # mevcut_stok
                "EKLE", # hareket_yönü
                self._stok_hareketi_tamamlandi_callback, # yenile_stok_listesi_callback
                parent_pencere=self # <-- BU PARAMETRENİN DOĞRU GEÇİLDİĞİNDEN EMİN OLUN
            )
            # YENİ EKLENDİ: Pop-up kapanınca tetiklenecek ek callback
            stok_hareketi_popup.protocol("WM_DELETE_WINDOW", lambda: self._stok_hareketi_popup_kapandi(stok_hareketi_popup))
            stok_hareketi_popup.after(100, stok_hareketi_popup.grab_set)

            self.app.set_status("Stok giriş penceresi açıldı.")
        else:
            messagebox.showerror("Hata", "Ürün bilgileri alınamadı.", parent=self)

    def _stok_eksilt_penceresi_ac(self):
        """Stok eksiltme penceresini 'EKSILT' yönüyle açar."""
        if not self.urun_id:
            messagebox.showwarning("Uyarı", "Lütfen işlem yapmak için bir ürün seçin.", parent=self)
            return

        urun_guncel_bilgi = self.db.stok_getir_by_id(self.urun_id)
        if urun_guncel_bilgi:
            mevcut_stok = urun_guncel_bilgi[3]

            stok_hareketi_popup = StokHareketiPenceresi(
                self.app, # parent_app
                self.db,
                self.urun_id,
                self.urun_detaylari[2], # urun_adi
                mevcut_stok, # mevcut_stok
                "EKSILT", # hareket_yönü
                self._stok_hareketi_tamamlandi_callback, # yenile_stok_listesi_callback
                parent_pencere=self # <-- BU PARAMETRENİN DOĞRU GEÇİLDİĞİNDEN EMİN OLUN
            )
            # YENİ EKLENDİ: Pop-up kapanınca tetiklenecek ek callback
            stok_hareketi_popup.protocol("WM_DELETE_WINDOW", lambda: self._stok_hareketi_popup_kapandi(stok_hareketi_popup))
            stok_hareketi_popup.after(100, stok_hareketi_popup.grab_set)

            self.app.set_status("Stok çıkış penceresi açıldı.")
        else:
            messagebox.showerror("Hata", "Ürün bilgileri alınamadı.", parent=self)

    def _stok_hareketi_popup_kapandi(self, popup_instance):
        """
        Stok Hareketi pop-up penceresi (StokHareketiPenceresi) kapatıldığında tetiklenir.
        Ürün kartının stok miktarını anlık olarak günceller.
        """
        print(f"DEBUG: _stok_hareketi_popup_kapandi çağrıldı. Popup kapandı.")

        if popup_instance.winfo_exists():
            popup_instance.destroy()

        self._load_genel_bilgiler()
        self.yenile_callback() # Ana stok listesini de güncelle

        self.update_idletasks()
        self.update()

        if self.entry_stok:
            self.entry_stok.focus_set()
            self.entry_stok.selection_range(0, tk.END)

        print(f"DEBUG: Ürün kartı anlık olarak güncellendi. Güncel Stok: {self.sv_stok.get()}")

    def _guncel_stogu_ui_a_yansit(self, guncel_stok_miktari):
        """
        Ürün kartındaki stok miktarını UI'da anlık olarak günceller.
        """
        # Stok miktarını StringVar'a formatlı şekilde set et
        self.sv_stok.set(f"{guncel_stok_miktari:.2f}".rstrip('0').rstrip('.'))
        # UI'ın kendini yenilemesini tetiklemek için update_idletasks() veya update() kullanabiliriz.
        # Genellikle bu set işlemi yeterli olur, ancak bazen görsel gecikmeleri önlemek için faydalıdır.
        self.update_idletasks() 

    def _stok_hareketi_tamamlandi_callback(self):
        """
        Stok hareketi tamamlandığında (kaydetme başarılı olduğunda) tetiklenir.
        Bu metod artık hem ana stok listesini hem de açık olan ürün kartını günceller.
        """
        print(f"DEBUG: _stok_hareketi_tamamlandi_callback çağrıldı.")

        if self.urun_id:
            guncel_urun_verisi = self.db.stok_getir_by_id(self.urun_id)
            if guncel_urun_verisi:
                self.urun_duzenle = guncel_urun_verisi
                self._load_genel_bilgiler()
                self.update_idletasks()

        self.yenile_callback()

        guncel_urun_stok = self.db.stok_getir_by_id(self.urun_id)
        guncel_stok_miktari_display = f"{guncel_urun_stok[3]:.2f}".rstrip('0').rstrip('.') if guncel_urun_stok else "Bilinmiyor"

        self.app.set_status(f"Stok hareketi başarıyla kaydedildi. Ürün: {self.urun_adi_initial}. Güncel Stok: {guncel_stok_miktari_display}")

    def _load_urun_grubu_birimi_ulke_fields(self):
        if self.urun_detaylari:
            # `urun_detaylari` tuple'ının indeksleri (db.stok_getir_by_id sorgusundan)
            # ug.grup_adi (19), ub.birim_adi (20), ul.ulke_adi (21)

            urun_grubu_adi = self.urun_detaylari[19] if len(self.urun_detaylari) > 19 and self.urun_detaylari[19] is not None else "Seçim Yok"
            urun_birimi_adi = self.urun_detaylari[20] if len(self.urun_detaylari) > 20 and self.urun_detaylari[20] is not None else "Seçim Yok"
            ulke_adi = self.urun_detaylari[21] if len(self.urun_detaylari) > 21 and self.urun_detaylari[21] is not None else "Seçim Yok"

            # self.combo_urun_grubu'na değerleri atama
            if urun_grubu_adi != "Seçim Yok" and urun_grubu_adi in self.urun_gruplari_map:
                self.combo_urun_grubu.set(urun_grubu_adi)
            else:
                self.combo_urun_grubu.set("Seçim Yok")

            # self.combo_urun_birimi'ye değerleri atama
            if urun_birimi_adi != "Seçim Yok" and urun_birimi_adi in self.urun_birimleri_map:
                self.combo_urun_birimi.set(urun_birimi_adi)
            else:
                self.combo_urun_birimi.set("Seçim Yok")

            # self.combo_mense'ye değerleri atama
            if ulke_adi != "Seçim Yok" and ulke_adi in self.ulkeler_map:
                self.combo_mense.set(ulke_adi)
            else:
                self.combo_mense.set("Seçim Yok")

    def _resim_sec(self):
        file_path = filedialog.askopenfilename(
            title="Ürün Resmi Seç",
            filetypes=[("Resim Dosyaları", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"), ("Tüm Dosyalar", "*.*")],
            parent=self
        )
        if file_path:
            try:
                resim_klasoru = os.path.join(self.db.data_dir, "urun_resimleri")
                os.makedirs(resim_klasoru, exist_ok=True)

                file_name = os.path.basename(file_path)
                destination_path = os.path.join(resim_klasoru, file_name)

                shutil.copy2(file_path, destination_path)

                self.urun_resmi_path = destination_path
                self._load_urun_resmi() # Resmi yükle ve göster
                self.app.set_status(f"Resim '{file_name}' başarıyla yüklendi ve kaydedildi.")
            except Exception as e:
                messagebox.showerror("Resim Yükleme Hatası", f"Resim kopyalanırken bir hata oluştu: {e}", parent=self)
                print(f"Resim kopyalanırken hata: {e}")

    def _resim_sil(self):
        if messagebox.askyesno("Resmi Sil", "Ürün resmini silmek istediğinizden emin misiniz?", parent=self):
            self.urun_resmi_path = ""
            self.urun_resmi_label.config(image='', text="Resim Yok")
            self.original_image = None
            self.tk_image = None
            messagebox.showinfo("Resim Silindi", "Ürün resmi başarıyla silindi.", parent=self)

    def _load_urun_resmi(self):
        """
        Ürün resmi yolunu kontrol eder ve resmi ayrı bir thread'de yükleme ve boyutlandırma işlemini başlatır.
        Bu sayede UI'ın donması engellenir.
        """
        self.original_image = None
        self.tk_image = None
        self._last_resized_size = (0, 0)
        self.urun_resmi_label.config(image='', text="Resim Yükleniyor...")

        if self.urun_resmi_path and os.path.exists(self.urun_resmi_path):
            threading.Thread(target=self._perform_image_loading_and_resizing).start()
        else:
            self.urun_resmi_label.config(image='', text="Resim Yok")
            self.original_image = None
            self.tk_image = None
            self._last_resized_size = (0, 0)

    def _perform_image_loading_and_resizing(self):
        """
        Resmi yükler ve boyutlandırır (PIL Image objesi olarak). Bu metot ayrı bir thread'de çalışır.
        Tamamlandığında, UI'a hazır PIL Image referansını ve boyut bilgilerini gönderir.
        """
        try:
            original_img = Image.open(self.urun_resmi_path)
            self.after_idle(lambda: self._update_image_on_ui_thread(original_img))
        except Exception as e:
            self.after_idle(lambda: self.urun_resmi_label.config(image='', text=f"Resim Hatası: {e}"))
            self.after_idle(lambda: setattr(self, 'original_image', None))
            self.after_idle(lambda: setattr(self, 'tk_image', None))
            self.after_idle(lambda: setattr(self, '_last_resized_size', (0, 0)))
            print(f"Arka plan resim yükleme hatası: {e}\n{traceback.format_exc()}")


    def _update_image_on_ui_thread(self, original_img_from_thread):
        """
        Arka plan thread'inden gelen orijinal PIL Image objesini UI'da saklar ve
        boyutlandırma işlemini tetikler. Bu metot sadece ana UI thread'inde çağrılmalıdır.
        """
        try:
            self.original_image = original_img_from_thread
            self._resize_image() # Resim yüklendikten sonra boyutlandırmayı tetikle
        except Exception as e:
            print(f"UI thread resim güncelleme hatası: {e}\n{traceback.format_exc()}")
            self.urun_resmi_label.config(image='', text="Resim Gösterme Hatası")
            self.tk_image = None
            self.original_image = None
            self._last_resized_size = (0, 0)


    def _resize_image(self, event=None):
        """
        Label'ı içeren konteyner boyutu değiştiğinde resmi uygun şekilde yeniden boyutlandırır.
        """
        if not self.original_image:
            self.urun_resmi_label.config(image='', text="Resim Yok")
            self.tk_image = None
            self._last_resized_size = (0, 0)
            return

        container_width = self.image_container_frame.winfo_width()
        container_height = self.image_container_frame.winfo_height()

        if container_width <= 1 or container_height <= 1:
            return

        if self._last_resized_size == (container_width, container_height) and self.tk_image:
            return

        img_width, img_height = self.original_image.size

        ratio_w = container_width / img_width
        ratio_h = container_height / img_height
        ratio = min(ratio_w, ratio_h)

        new_width = int(img_width * ratio)
        new_height = int(img_height * ratio)

        if new_width <= 0: new_width = 1
        if new_height <= 0: new_height = 1

        if self._last_resized_size != (new_width, new_height):
            self._last_resized_size = (new_width, new_height)

            try:
                resized_image = self.original_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
                self.tk_image = ImageTk.PhotoImage(resized_image)
                self.urun_resmi_label.config(image=self.tk_image, text="")
            except Exception as e_resize:
                print(f"Resim yeniden boyutlandırılırken hata: {e_resize}\n{traceback.format_exc()}")
                self.urun_resmi_label.config(image='', text="Resim Boyutlandırma Hatası")
                self.tk_image = None
                self._last_resized_size = (0, 0)
            else:
                # Boyut değişmediyse ve zaten bir resim gösteriliyorsa, ek bir işlem yapma.
                pass

    def _setup_price_change_date_label(self, parent_frame):
        self.fiyat_degisiklik_tarihi_label = ttk.Label(parent_frame, text="Fiyat Değişiklik Tarihi: Yükleniyor...", font=("Segoe UI", 9, "italic"))
        self.fiyat_degisiklik_tarihi_label.grid(row=10, column=2, columnspan=2, padx=5, pady=(5, 0), sticky=tk.SE)

    def _on_tab_change(self, event):
        selected_tab_id = self.notebook.select()
        selected_tab_text = self.notebook.tab(selected_tab_id, "text")

        if selected_tab_text == "Stok Hareketleri":
            if self.urun_id: # Sadece ürün ID'si varsa yükle
                self._load_stok_hareketleri()
        elif selected_tab_text == "İlgili Faturalar":
            if self.urun_id: # Sadece ürün ID'si varsa yükle
                self._load_ilgili_faturalar()
        elif selected_tab_text == "Kategori & Marka Yönetimi": 
            # Bu sekmeye geçildiğinde combobox'lar zaten _yukle_kategori_marka_comboboxlari
            # ve _yukle_urun_grubu_birimi_ulke_comboboxlari tarafından doldurulmuş olmalı.
            # Treeview'ları yenilemek isteyebiliriz:
            self._kategori_listesini_yukle()
            self._marka_listesini_yukle()


    def _setup_genel_bilgiler_tab(self, parent_frame):
        # parent_frame (genel_bilgiler_sekmesi_frame) içindeki grid yapısı
        parent_frame.columnconfigure(0, weight=3) # Sol taraf daha çok genişlesin
        parent_frame.columnconfigure(1, weight=1) # Sağ taraf daha az genişlesin
        parent_frame.rowconfigure(0, weight=1) # Ana satır (dikeyde genişleyebilir)

        # SOL TARAFTAKİ BİLGİLERİ İÇERECEK ANA CONTAINER FRAME
        left_info_container_frame = ttk.Frame(parent_frame)
        left_info_container_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        left_info_container_frame.columnconfigure(0, weight=1)

        # --- 1. TEMEL ÜRÜN BİLGİLERİ GRUBU ---
        basic_info_frame = ttk.LabelFrame(left_info_container_frame, text="Temel Ürün Bilgileri", padding="10")
        basic_info_frame.pack(fill=tk.X, padx=2, pady=2, ipady=5)
        basic_info_frame.columnconfigure(1, weight=1)
        basic_info_frame.columnconfigure(3, weight=1)

        row_in_basic = 0
        ttk.Label(basic_info_frame, text="Ürün Kodu:").grid(row=row_in_basic, column=0, padx=5, pady=2, sticky=tk.W)
        self.entry_kod = ttk.Entry(basic_info_frame, textvariable=self.sv_kod)
        self.entry_kod.grid(row=row_in_basic, column=1, padx=5, pady=2, sticky=tk.EW)

        ttk.Label(basic_info_frame, text="Ürün Adı:").grid(row=row_in_basic, column=2, padx=5, pady=2, sticky=tk.W)
        self.entry_ad = ttk.Entry(basic_info_frame, textvariable=self.sv_ad)
        self.entry_ad.grid(row=row_in_basic, column=3, padx=5, pady=2, sticky=tk.EW)
        row_in_basic += 1

        ttk.Label(basic_info_frame, text="Ürün Detayı:").grid(row=row_in_basic, column=0, padx=5, pady=2, sticky=tk.NW)
        self.entry_urun_detayi = tk.Text(basic_info_frame, height=3, wrap=tk.WORD, font=('Segoe UI', 9))
        self.entry_urun_detayi.grid(row=row_in_basic, column=1, columnspan=3, padx=5, pady=2, sticky=tk.EW)
        # Ürün detayı için dikey scrollbar ekle
        urun_detayi_vsb = ttk.Scrollbar(basic_info_frame, orient="vertical", command=self.entry_urun_detayi.yview)
        urun_detayi_vsb.grid(row=row_in_basic, column=4, sticky="ns")
        self.entry_urun_detayi.config(yscrollcommand=urun_detayi_vsb.set)
        basic_info_frame.columnconfigure(4, weight=0)
        basic_info_frame.rowconfigure(row_in_basic, weight=1)

        # --- 2. STOK DURUMU GRUBU ---
        stock_info_frame = ttk.LabelFrame(left_info_container_frame, text="Stok Durumu", padding="10")
        stock_info_frame.pack(fill=tk.X, padx=2, pady=5, ipady=5)
        stock_info_frame.columnconfigure(1, weight=1)
        stock_info_frame.columnconfigure(3, weight=1)

        row_in_stock = 0
        ttk.Label(stock_info_frame, text="Mevcut Stok:").grid(row=row_in_stock, column=0, padx=5, pady=2, sticky=tk.W)
        self.entry_stok = ttk.Entry(stock_info_frame, textvariable=self.sv_stok)
        self.entry_stok.grid(row=row_in_stock, column=1, padx=5, pady=2, sticky=tk.EW)
        setup_numeric_entry(self.app, self.entry_stok, decimal_places=2)
        self.entry_stok.bind("<FocusOut>", lambda e: self._format_stok_entry(sv_variable=self.sv_stok, decimal_places=2, focus_out=True))


        ttk.Label(stock_info_frame, text="Min. Stok Seviyesi:").grid(row=row_in_stock, column=2, padx=5, pady=2, sticky=tk.W)
        self.entry_min_stok = ttk.Entry(stock_info_frame, textvariable=self.sv_min_stok)
        self.entry_min_stok.grid(row=row_in_stock, column=3, padx=5, pady=2, sticky=tk.EW)
        setup_numeric_entry(self.app, self.entry_min_stok, decimal_places=2)
        self.entry_min_stok.bind("<FocusOut>", lambda e: self._format_stok_entry(sv_variable=self.sv_min_stok, decimal_places=2, focus_out=True))

        # --- 3. FİYATLANDIRMA BİLGİLERİ GRUBU ---
        price_info_frame = ttk.LabelFrame(left_info_container_frame, text="Fiyatlandırma Bilgileri", padding="10")
        price_info_frame.pack(fill=tk.X, padx=2, pady=5, ipady=5)
        price_info_frame.columnconfigure(1, weight=1)
        price_info_frame.columnconfigure(3, weight=1)

        row_in_price = 0
        ttk.Label(price_info_frame, text="KDV Oranı (%):").grid(row=row_in_price, column=0, padx=5, pady=2, sticky=tk.W)
        self.entry_kdv = ttk.Entry(price_info_frame, textvariable=self.sv_kdv)
        self.entry_kdv.grid(row=row_in_price, column=1, padx=5, pady=2, sticky=tk.EW)
        setup_numeric_entry(self.app, self.entry_kdv, decimal_places=0, max_value=100)
        self.entry_kdv.bind("<KeyRelease>", self.otomatik_fiyat_doldur)
        self.entry_kdv.bind("<FocusOut>", lambda e: self.otomatik_fiyat_doldur(e, source_type='kdv_focout', price_type='all', focus_out=True))
        row_in_price += 1

        ttk.Label(price_info_frame, text="Alış Fiyatı (KDV Hariç):").grid(row=row_in_price, column=0, padx=5, pady=2, sticky=tk.W)
        self.entry_alis_haric = ttk.Entry(price_info_frame, textvariable=self.sv_alis_haric)
        self.entry_alis_haric.grid(row=row_in_price, column=1, padx=5, pady=2, sticky=tk.EW)
        setup_numeric_entry(self.app, self.entry_alis_haric, decimal_places=2)
        self.entry_alis_haric.bind("<KeyRelease>", lambda e: self.otomatik_fiyat_doldur(e, source_type='haric', price_type='alis'))
        self.entry_alis_haric.bind("<FocusOut>", lambda e: self.otomatik_fiyat_doldur(e, source_type='haric', price_type='alis', focus_out=True))

        ttk.Label(price_info_frame, text="Alış Fiyatı (KDV Dahil):").grid(row=row_in_price, column=2, padx=5, pady=2, sticky=tk.W)
        self.entry_alis_dahil = ttk.Entry(price_info_frame, textvariable=self.sv_alis_dahil)
        self.entry_alis_dahil.grid(row=row_in_price, column=3, padx=5, pady=2, sticky=tk.EW)
        setup_numeric_entry(self.app, self.entry_alis_dahil, decimal_places=2)
        self.entry_alis_dahil.bind("<KeyRelease>", lambda e: self.otomatik_fiyat_doldur(e, source_type='dahil', price_type='alis'))
        self.entry_alis_dahil.bind("<FocusOut>", lambda e: self.otomatik_fiyat_doldur(e, source_type='dahil', price_type='alis', focus_out=True))
        row_in_price += 1

        ttk.Label(price_info_frame, text="Satış Fiyatı (KDV Hariç):").grid(row=row_in_price, column=0, padx=5, pady=2, sticky=tk.W)
        self.entry_satis_haric = ttk.Entry(price_info_frame, textvariable=self.sv_satis_haric)
        self.entry_satis_haric.grid(row=row_in_price, column=1, padx=5, pady=2, sticky=tk.EW)
        setup_numeric_entry(self.app, self.entry_satis_haric, decimal_places=2)
        self.entry_satis_haric.bind("<KeyRelease>", lambda e: self.otomatik_fiyat_doldur(e, source_type='haric', price_type='satis'))
        self.entry_satis_haric.bind("<FocusOut>", lambda e: self.otomatik_fiyat_doldur(e, source_type='haric', price_type='satis', focus_out=True))
        row_in_price += 1

        ttk.Label(price_info_frame, text="Satış Fiyatı (KDV Dahil):").grid(row=row_in_price, column=0, padx=5, pady=2, sticky=tk.W)
        self.entry_satis_dahil = ttk.Entry(price_info_frame, textvariable=self.sv_satis_dahil)
        self.entry_satis_dahil.grid(row=row_in_price, column=1, padx=5, pady=2, sticky=tk.EW)
        setup_numeric_entry(self.app, self.entry_satis_dahil, decimal_places=2)
        self.entry_satis_dahil.bind("<KeyRelease>", lambda e: self.otomatik_fiyat_doldur(e, source_type='dahil', price_type='satis'))
        self.entry_satis_dahil.bind("<FocusOut>", lambda e: self.otomatik_fiyat_doldur(e, source_type='dahil', price_type='satis', focus_out=True))
        row_in_price += 1

        self.label_kar_orani = ttk.Label(price_info_frame, text="0.00 %", font=("Segoe UI", 9, "bold"))
        self.label_kar_orani.grid(row=row_in_price, column=0, columnspan=2, padx=5, pady=2, sticky=tk.W)

        self.fiyat_degisiklik_tarihi_label = ttk.Label(price_info_frame, text="Fiyat Değişiklik Tarihi: Yükleniyor...", font=("Segoe UI", 9, "italic"))
        self.fiyat_degisiklik_tarihi_label.grid(row=row_in_price, column=2, columnspan=2, padx=5, pady=2, sticky=tk.SE)

        # --- 4. EK NİTELİKLER GRUBU ---
        attributes_info_frame = ttk.LabelFrame(left_info_container_frame, text="Ek Nitelikler", padding="10")
        attributes_info_frame.pack(fill=tk.X, padx=2, pady=5, ipady=5)
        attributes_info_frame.columnconfigure(1, weight=1)
        attributes_info_frame.columnconfigure(3, weight=1)

        row_in_attr = 0
        ttk.Label(attributes_info_frame, text="Kategori:").grid(row=row_in_attr, column=0, padx=5, pady=2, sticky=tk.W)
        self.combo_kategori = ttk.Combobox(attributes_info_frame, state="readonly")
        self.combo_kategori.grid(row=row_in_attr, column=1, padx=5, pady=2, sticky=tk.EW)

        ttk.Label(attributes_info_frame, text="Marka:").grid(row=row_in_attr, column=2, padx=5, pady=2, sticky=tk.W)
        self.combo_marka = ttk.Combobox(attributes_info_frame, state="readonly")
        self.combo_marka.grid(row=row_in_attr, column=3, padx=5, pady=2, sticky=tk.EW)
        row_in_attr += 1

        ttk.Label(attributes_info_frame, text="Ürün Grubu:").grid(row=row_in_attr, column=0, padx=5, pady=2, sticky=tk.W)
        self.combo_urun_grubu = ttk.Combobox(attributes_info_frame, state="readonly")
        self.combo_urun_grubu.grid(row=row_in_attr, column=1, padx=5, pady=2, sticky=tk.EW)

        ttk.Label(attributes_info_frame, text="Ürün Birimi:").grid(row=row_in_attr, column=2, padx=5, pady=2, sticky=tk.W)
        self.combo_urun_birimi = ttk.Combobox(attributes_info_frame, state="readonly")
        self.combo_urun_birimi.grid(row=row_in_attr, column=3, padx=5, pady=2, sticky=tk.EW)
        row_in_attr += 1

        ttk.Label(attributes_info_frame, text="Menşe:").grid(row=row_in_attr, column=0, padx=5, pady=2, sticky=tk.W)
        self.combo_mense = ttk.Combobox(attributes_info_frame, state="readonly")
        self.combo_mense.grid(row=row_in_attr, column=1, padx=5, pady=2, sticky=tk.EW)

        # Bu butona tıklandığında UrunNitelikYonetimiPenceresi açılacak.
        from pencereler import UrunNitelikYonetimiPenceresi # Bu import'un doğru olduğundan emin olun
        ttk.Button(attributes_info_frame, text="Nitelik Yönetimi", command=lambda: UrunNitelikYonetimiPenceresi(self.app, self.db, self._yukle_urun_grubu_birimi_ulke_comboboxlari)).grid(row=row_in_attr, column=2, columnspan=2, padx=5, pady=(10,5), sticky=tk.EW)
        row_in_attr += 1 # Buton ekledikten sonra satır indeksini artırın

        # SAĞ TARAFTAKİ "ÜRÜN GÖRSELİ" VE "OPERASYONLAR" ÇERÇEVELERİNİ YERLEŞTİRME
        self.urun_gorsel_ve_operasyon_frame.columnconfigure(0, weight=1)
        self.urun_gorsel_ve_operasyon_frame.rowconfigure(0, weight=1)
        self.urun_gorsel_ve_operasyon_frame.rowconfigure(1, weight=1)


        # Resim Çerçevesi (şimdi grid ile)
        self.urun_gorsel_frame = ttk.LabelFrame(self.urun_gorsel_ve_operasyon_frame, text="Ürün Görseli", padding="5")
        self.urun_gorsel_frame.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        self.urun_gorsel_frame.columnconfigure(0, weight=1)
        self.urun_gorsel_frame.rowconfigure(0, weight=1)

        # image_container_frame de artık grid ile yönetilmeli
        self.image_container_frame = ttk.Frame(self.urun_gorsel_frame, relief="solid", borderwidth=1)
        self.image_container_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.image_container_frame.grid_propagate(False)

        # urun_resmi_label, image_container_frame içinde pack olarak kalabilir, çünkü image_container_frame artık kendi başına bir yöneticidir.
        self.urun_resmi_label = ttk.Label(self.image_container_frame, text="Resim Yok", anchor=tk.CENTER)
        self.urun_resmi_label.pack(expand=True, fill=tk.BOTH)

        self.image_container_frame.bind("<Configure>", self._resize_image)

        # Resim Seç/Sil butonları
        button_frame_gorsel = ttk.Frame(self.urun_gorsel_frame)
        button_frame_gorsel.grid(row=1, column=0, sticky="ew", padx=2, pady=2)
        button_frame_gorsel.columnconfigure(0, weight=1)
        button_frame_gorsel.columnconfigure(1, weight=1)

        ttk.Button(button_frame_gorsel, text="Resim Seç", command=self._resim_sec, style="Accent.TButton").grid(row=0, column=0, padx=1, pady=1, sticky="ew")
        ttk.Button(button_frame_gorsel, text="Resmi Sil", command=self._resim_sil).grid(row=0, column=1, padx=1, pady=1, sticky="ew")

        # Operasyon butonları (Stok Ekle, Stok Eksilt vb.) (şimdi grid ile)
        self.operation_buttons_frame = ttk.LabelFrame(self.urun_gorsel_ve_operasyon_frame, text="Operasyonlar", padding="5")
        self.operation_buttons_frame.grid(row=1, column=0, sticky="nsew", padx=2, pady=(5,0))
        self.operation_buttons_frame.columnconfigure(0, weight=1)


        # Butonları şimdi grid() ile konumlandırıyoruz (önceki pack yerine)
        button_row_idx = 0
        ttk.Button(self.operation_buttons_frame, text="Stok Ekle", command=self._stok_ekle_penceresi_ac, style="Accent.TButton").grid(row=button_row_idx, column=0, sticky="ew", padx=1, pady=1)
        button_row_idx += 1
        ttk.Button(self.operation_buttons_frame, text="Stok Eksilt", command=self._stok_eksilt_penceresi_ac).grid(row=button_row_idx, column=0, sticky="ew", padx=1, pady=1)
        button_row_idx += 1
        ttk.Button(self.operation_buttons_frame, text="Ürüne ait iadeler (Geliştirilecek)", state=tk.DISABLED).grid(row=button_row_idx, column=0, sticky="ew", padx=1, pady=1)
        button_row_idx += 1
        ttk.Button(self.operation_buttons_frame, text="Ürün üret/tüket (Geliştirilecek)", state=tk.DISABLED).grid(row=button_row_idx, column=0, sticky="ew", padx=1, pady=1)
        button_row_idx += 1

        # Fiyat değişiklik tarihi etiketini de grid() ile konumlandırıyoruz
        self.fiyat_degisiklik_tarihi_label.grid(row=button_row_idx, column=0, sticky="w", padx=5, pady=(5,0))

    def _yukle_kategori_marka_comboboxlari(self):
        # Kategori ve marka verilerini DB'den al
        kategoriler_map = self.db.get_kategoriler_for_combobox()
        markalar_map = self.db.get_markalar_for_combobox()

        # Combobox'ları doldurma
        self.kategoriler_map = {"Seçim Yok": None, **kategoriler_map}
        self.combo_kategori['values'] = ["Seçim Yok"] + sorted(kategoriler_map.keys())

        self.markalar_map = {"Seçim Yok": None, **markalar_map}
        self.combo_marka['values'] = ["Seçim Yok"] + sorted(markalar_map.keys())

        # Seçili değerleri ayarla (eğer ürün düzenleniyorsa)
        if self.urun_duzenle:
            kategori_adi = self.urun_duzenle[14] # Kategori Adı
            marka_adi = self.urun_duzenle[15] # Marka Adı
            self.combo_kategori.set(kategori_adi if kategori_adi in self.kategoriler_map else "Seçim Yok")
            self.combo_marka.set(marka_adi if marka_adi in self.markalar_map else "Seçim Yok")
        else:
            self.combo_kategori.set("Seçim Yok")
            self.combo_marka.set("Seçim Yok")

    def _load_kategori_marka_fields(self):
        if self.urun_duzenle:
            # `urun_duzenle` tuple'ının indeksleri (db.stok_getir_by_id sorgusundan)
            # uk.kategori_adi (14), um.marka_adi (15)

            kategori_adi = self.urun_duzenle[14] if len(self.urun_duzenle) > 14 and self.urun_duzenle[14] is not None else "Seçim Yok"
            marka_adi = self.urun_duzenle[15] if len(self.urun_duzenle) > 15 and self.urun_duzenle[15] is not None else "Seçim Yok"

            # self.combo_kategori'ye değerleri atama
            # Sadece eğer kategori_adi "Seçim Yok" değilse ve haritada varsa set et
            if kategori_adi != "Seçim Yok" and kategori_adi in self.kategoriler_map: 
                self.combo_kategori.set(kategori_adi)
            else:
                self.combo_kategori.set("Seçim Yok") # Yoksa varsayılan

            # self.combo_marka'ya değerleri atama
            # Sadece eğer marka_adi "Seçim Yok" değilse ve haritada varsa set et
            if marka_adi != "Seçim Yok" and marka_adi in self.markalar_map: 
                self.combo_marka.set(marka_adi)
            else:
                self.combo_marka.set("Seçim Yok") # Yoksa varsayılan


    def _setup_kategori_marka_tab(self, parent_frame):
        # Parent frame'in grid yapılandırması (bu sekmenin içindeki düzen)
        parent_frame.columnconfigure(0, weight=1) # Kategori Frame için
        parent_frame.columnconfigure(1, weight=1) # Marka Frame için
        parent_frame.rowconfigure(0, weight=1) # Kategori/Marka Frame'ler için

        # Sol taraf: Kategori Yönetimi
        kategori_frame = ttk.LabelFrame(parent_frame, text="Kategori Yönetimi", padding="10")
        kategori_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew") # Grid kullanıldı
        kategori_frame.columnconfigure(1, weight=1) # Entry'nin genişlemesi için
        kategori_frame.grid_rowconfigure(1, weight=1) # Treeview'in genişlemesi için


        ttk.Label(kategori_frame, text="Kategori Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.kategori_entry = ttk.Entry(kategori_frame, width=30)
        self.kategori_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(kategori_frame, text="Ekle", command=self._kategori_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(kategori_frame, text="Güncelle", command=self._kategori_guncelle_ui).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(kategori_frame, text="Sil", command=self._kategori_sil_ui).grid(row=0, column=4, padx=5, pady=5)

        self.kategori_tree = ttk.Treeview(kategori_frame, columns=("ID", "Kategori Adı"), show='headings', selectmode="browse")
        self.kategori_tree.heading("ID", text="ID"); self.kategori_tree.column("ID", width=50, stretch=tk.NO)
        self.kategori_tree.heading("Kategori Adı", text="Kategori Adı"); self.kategori_tree.column("Kategori Adı", width=200, stretch=tk.YES)
        self.kategori_tree.grid(row=1, column=0, columnspan=5, padx=5, pady=10, sticky="nsew")
        
        self.kategori_tree.bind("<<TreeviewSelect>>", self._on_kategori_select)


        # Sağ taraf: Marka Yönetimi
        marka_frame = ttk.LabelFrame(parent_frame, text="Marka Yönetimi", padding="10")
        marka_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew") # Grid kullanıldı
        marka_frame.columnconfigure(1, weight=1) # Entry'nin genişlemesi için
        marka_frame.grid_rowconfigure(1, weight=1) # Treeview'in genişlemesi için


        ttk.Label(marka_frame, text="Marka Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.marka_entry = ttk.Entry(marka_frame, width=30)
        self.marka_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(marka_frame, text="Ekle", command=self._marka_ekle_ui).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(marka_frame, text="Güncelle", command=self._marka_guncelle_ui).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(marka_frame, text="Sil", command=self._marka_sil_ui).grid(row=0, column=4, padx=5, pady=5)

        self.marka_tree = ttk.Treeview(marka_frame, columns=("ID", "Marka Adı"), show='headings', selectmode="browse")
        self.marka_tree.heading("ID", text="ID"); self.marka_tree.column("ID", width=50, stretch=tk.NO)
        self.marka_tree.heading("Marka Adı", text="Marka Adı"); self.marka_tree.column("Marka Adı", width=200, stretch=tk.YES)
        self.marka_tree.grid(row=1, column=0, columnspan=5, padx=5, pady=10, sticky="nsew")
        
        self.marka_tree.bind("<<TreeviewSelect>>", self._on_marka_select)


    def _setup_stok_hareketleri_tab(self, parent_frame):
        ttk.Label(parent_frame, text="Ürün Stok Hareketleri", font=("Segoe UI", 12, "bold")).pack(pady=5, anchor=tk.W)

        # Filtreleme seçenekleri
        filter_frame = ttk.Frame(parent_frame, padding="5")
        filter_frame.pack(fill=tk.X, padx=0, pady=5)

        ttk.Label(filter_frame, text="İşlem Tipi:").pack(side=tk.LEFT, padx=(0,2))
        self.stok_hareket_tip_filter_cb = ttk.Combobox(filter_frame, width=18, values=["TÜMÜ", "Giriş (Manuel)", "Çıkış (Manuel)", "Sayım Fazlası", "Sayım Eksiği", "Zayiat", "İade Girişi", "Fatura Alış", "Fatura Satış"], state="readonly")
        self.stok_hareket_tip_filter_cb.pack(side=tk.LEFT, padx=(0,10))
        self.stok_hareket_tip_filter_cb.set("TÜMÜ")
        self.stok_hareket_tip_filter_cb.bind("<<ComboboxSelected>>", self._load_stok_hareketleri)

        ttk.Label(filter_frame, text="Başlangıç Tarihi:").pack(side=tk.LEFT, padx=(0,2))
        self.stok_hareket_bas_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.stok_hareket_bas_tarih_entry.pack(side=tk.LEFT, padx=(0,5))
        self.stok_hareket_bas_tarih_entry.insert(0, (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.stok_hareket_bas_tarih_entry)
        ttk.Button(filter_frame, text="🗓️", command=lambda: DatePickerDialog(self.app, self.stok_hareket_bas_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)

        ttk.Label(filter_frame, text="Bitiş Tarihi:").pack(side=tk.LEFT, padx=(0,2))
        self.stok_hareket_bit_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.stok_hareket_bit_tarih_entry.pack(side=tk.LEFT, padx=(0,10))
        self.stok_hareket_bit_tarih_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.stok_hareket_bit_tarih_entry)
        ttk.Button(filter_frame, text="🗓️", command=lambda: DatePickerDialog(self.app, self.stok_hareket_bit_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)

        ttk.Button(filter_frame, text="Yenile", command=self._load_stok_hareketleri, style="Accent.TButton").pack(side=tk.LEFT)


        # Stok Hareketleri Treeview
        cols_stok_hareket = ("ID", "Tarih", "İşlem Tipi", "Miktar", "Önceki Stok", "Sonraki Stok", "Açıklama", "Kaynak")
        self.stok_hareket_tree = ttk.Treeview(parent_frame, columns=cols_stok_hareket, show='headings', selectmode="browse")

        col_defs_stok_hareket = [
            ("ID", 40, tk.E, tk.NO),
            ("Tarih", 80, tk.CENTER, tk.NO),
            ("İşlem Tipi", 100, tk.W, tk.NO),
            ("Miktar", 70, tk.E, tk.NO),
            ("Önceki Stok", 80, tk.E, tk.NO),
            ("Sonraki Stok", 80, tk.E, tk.NO),
            ("Açıklama", 250, tk.W, tk.YES),
            ("Kaynak", 80, tk.W, tk.NO)
        ]
        for cn, w, a, s in col_defs_stok_hareket:
            self.stok_hareket_tree.column(cn, width=w, anchor=a, stretch=s)
            self.stok_hareket_tree.heading(cn, text=cn, command=lambda c=cn: sort_treeview_column(self.stok_hareket_tree, c, False))
        
        vsb_stok_hareket = ttk.Scrollbar(parent_frame, orient="vertical", command=self.stok_hareket_tree.yview)
        hsb_stok_hareket = ttk.Scrollbar(parent_frame, orient="horizontal", command=self.stok_hareket_tree.xview)
        self.stok_hareket_tree.configure(yscrollcommand=vsb_stok_hareket.set, xscrollcommand=hsb_stok_hareket.set)
        vsb_stok_hareket.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_stok_hareket.pack(side=tk.BOTTOM, fill=tk.X)
        self.stok_hareket_tree.pack(expand=True, fill=tk.BOTH)

    def _kategori_listesini_yukle(self):
        for i in self.kategori_tree.get_children(): self.kategori_tree.delete(i)
        kategoriler = self.db.kategori_listele()
        for kat in kategoriler: self.kategori_tree.insert("", tk.END, values=kat, iid=kat[0])
        self._yukle_kategori_marka_comboboxlari()

    def _on_kategori_select(self, event):
        selected_item = self.kategori_tree.focus()
        if selected_item:
            values = self.kategori_tree.item(selected_item, 'values')
            self.kategori_entry.delete(0, tk.END)
            self.kategori_entry.insert(0, values[1])
        else:
            self.kategori_entry.delete(0, tk.END)

    def _kategori_ekle_ui(self):
        kategori_adi = self.kategori_entry.get().strip()
        success, message = self.db.kategori_ekle(kategori_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self._kategori_listesini_yukle()
            if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            # UrunKartiPenceresi'nde aktif olan combobox'ları da güncelleme callback'i
            if self.refresh_callback:
                self.refresh_callback()

    def _kategori_guncelle_ui(self):
        selected_item = self.kategori_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir kategori seçin.", parent=self)
            return
        kategori_id = self.kategori_tree.item(selected_item)['values'][0]
        yeni_kategori_adi = self.kategori_entry.get().strip()

        success, message = self.db.kategori_guncelle(kategori_id, yeni_kategori_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.kategori_entry.delete(0, tk.END)
            self._kategori_listesini_yukle()
            self.app.set_status(f"Kategori '{yeni_kategori_adi}' güncellendi.") 
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _kategori_sil_ui(self):
        selected_item = self.kategori_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir kategori seçin.", parent=self)
            return
        kategori_id = self.kategori_tree.item(selected_item)['values'][0]
        kategori_adi = self.kategori_tree.item(selected_item)['values'][1] # Silinecek kategorinin adını al

        if messagebox.askyesno("Onay", f"'{kategori_adi}' kategorisini silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.kategori_sil(kategori_id)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.kategori_entry.delete(0, tk.END)
                self._kategori_listesini_yukle()
                self.app.set_status(f"Kategori '{kategori_adi}' silindi.") 
            else:
                messagebox.showerror("Hata", message, parent=self)

    def _marka_listesini_yukle(self):
        for i in self.marka_tree.get_children(): self.marka_tree.delete(i)
        markalar = self.db.marka_listele()
        for mar in markalar: self.marka_tree.insert("", tk.END, values=mar, iid=mar[0])
        self._yukle_kategori_marka_comboboxlari()

    def _on_marka_select(self, event):
        selected_item = self.marka_tree.focus()
        if selected_item:
            values = self.marka_tree.item(selected_item, 'values')
            self.marka_entry.delete(0, tk.END)
            self.marka_entry.insert(0, values[1])
        else:
            self.marka_entry.delete(0, tk.END)

    def _marka_ekle_ui(self):
        marka_adi = self.marka_entry.get().strip()
        success, message = self.db.marka_ekle(marka_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.marka_entry.delete(0, tk.END)
            self._marka_listesini_yukle()
            self.app.set_status(f"Marka '{marka_adi}' eklendi.")
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _marka_guncelle_ui(self):
        selected_item = self.marka_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek için bir marka seçin.", parent=self)
            return
        marka_id = self.marka_tree.item(selected_item)['values'][0]
        yeni_marka_adi = self.marka_entry.get().strip()

        success, message = self.db.marka_guncelle(marka_id, yeni_marka_adi)
        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.marka_entry.delete(0, tk.END)
            self._marka_listesini_yukle()
            self.app.set_status(f"Marka '{yeni_marka_adi}' güncellendi.")
        else:
            messagebox.showerror("Hata", message, parent=self)

    def _marka_sil_ui(self):
        selected_item = self.marka_tree.focus()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir marka seçin.", parent=self)
            return
        marka_id = self.marka_tree.item(selected_item)['values'][0]
        marka_adi = self.marka_tree.item(selected_item)['values'][1] # Silinecek markanın adını al

        if messagebox.askyesno("Onay", f"'{marka_adi}' markasını silmek istediğinizden emin misiniz?", parent=self):
            success, message = self.db.marka_sil(marka_id)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self.marka_entry.delete(0, tk.END)
                self._marka_listesini_yukle()
                self.app.set_status(f"Marka '{marka_adi}' silindi.") 
            else:
                messagebox.showerror("Hata", message, parent=self)

    def _load_genel_bilgiler(self):
        if self.urun_duzenle: # Sadece düzenleme modunda veri yükle

            print(f"{datetime.now()}: DEBUG: _load_genel_bilgiler çağrıldı.") 
            print(f"{datetime.now()}: DEBUG: Yüklenen ürün detayları: {self.urun_duzenle}") 

            # Ürün Kodu
            urun_kodu_val = self.urun_duzenle[1] if self.urun_duzenle[1] is not None else ""
            self.sv_kod.set(urun_kodu_val)
            print(f"{datetime.now()}: DEBUG: Ürün Kodu yüklendi: '{self.sv_kod.get()}'")

            # Ürün Adı
            urun_adi_val = self.urun_duzenle[2] if self.urun_duzenle[2] is not None else ""
            self.sv_ad.set(urun_adi_val)
            print(f"{datetime.now()}: DEBUG: Ürün Adı yüklendi: '{self.sv_ad.get()}'")

            # Ürün Detayı (tk.Text widget'ı)
            urun_detayi_db = self.urun_duzenle[16] if len(self.urun_duzenle) > 16 and self.urun_duzenle[16] is not None else ""
            self.entry_urun_detayi.delete("1.0", tk.END)
            self.entry_urun_detayi.insert("1.0", urun_detayi_db)
            print(f"{datetime.now()}: DEBUG: Ürün Detayı yüklendi.")

            # KDV Oranı
            kdv_val = self.urun_duzenle[6] if self.urun_duzenle[6] is not None else 0.0
            self.sv_kdv.set(f"{kdv_val:.0f}".replace('.',','))
            print(f"{datetime.now()}: DEBUG: KDV Oranı yüklendi: {self.sv_kdv.get()}")

            # Alış Fiyatı (KDV Hariç)
            alis_haric_val = self.urun_duzenle[4] if self.urun_duzenle[4] is not None else 0.0
            self.sv_alis_haric.set(f"{alis_haric_val:.2f}".replace('.',','))
            print(f"{datetime.now()}: DEBUG: Alış Fiyatı (Hariç) yüklendi: {self.sv_alis_haric.get()}")

                # Alış Fiyatı (KDV Dahil)
            alis_dahil_val = self.urun_duzenle[8] if self.urun_duzenle[8] is not None else 0.0
            self.sv_alis_dahil.set(f"{alis_dahil_val:.2f}".replace('.',','))
            print(f"{datetime.now()}: DEBUG: Alış Fiyatı (Dahil) yüklendi: {self.sv_alis_dahil.get()}")

            # Satış Fiyatı (KDV Hariç)
            satis_haric_val = self.urun_duzenle[5] if self.urun_duzenle[5] is not None else 0.0
            self.sv_satis_haric.set(f"{satis_haric_val:.2f}".replace('.',','))
            print(f"{datetime.now()}: DEBUG: Satış Fiyatı (Hariç) yüklendi: {self.sv_satis_haric.get()}")

            # Satış Fiyatı (KDV Dahil)
            satis_dahil_val = self.urun_duzenle[9] if self.urun_duzenle[9] is not None else 0.0
            self.sv_satis_dahil.set(f"{satis_dahil_val:.2f}".replace('.',','))
            print(f"{datetime.now()}: DEBUG: Satış Fiyatı (Dahil) yüklendi: {self.sv_satis_dahil.get()}")

            self._calculate_kar_orani()
            print(f"{datetime.now()}: DEBUG: Kar oranı hesaplandı.")

            # Fiyat Değişiklik Tarihi (Label)
            fiyat_deg_tarihi = self.urun_duzenle[18] if len(self.urun_duzenle) > 18 and self.urun_duzenle[18] is not None else "-"
            # Label'a direkt metin atadığımız için burada kontrol gerekli değil, _setup_genel_bilgiler_tab'da tanımlanıyor.
            self.fiyat_degisiklik_tarihi_label.config(text=f"Fiyat Değişiklik Tarihi: {fiyat_deg_tarihi}")
            print(f"{datetime.now()}: DEBUG: Fiyat Değişiklik Tarihi yüklendi: {fiyat_deg_tarihi}")

            # Ürün Resmi Yolu
            self.urun_resmi_path = self.urun_duzenle[17] if len(self.urun_duzenle) > 17 and self.urun_duzenle[17] is not None else ""
            self._load_urun_resmi() # Resim yükleme metodunu çağır
            print(f"{datetime.now()}: DEBUG: Ürün resmi yolu yüklendi: {self.urun_resmi_path}")

            stok_val = self.urun_duzenle[3] if self.urun_duzenle[3] is not None else 0.0
            self.sv_stok.set(f"{stok_val:.2f}".rstrip('0').rstrip('.')) 
            print(f"{datetime.now()}: DEBUG: Stok Miktarı yüklendi: {self.sv_stok.get()}")

            # Min. Stok Seviyesi
            min_stok_val = self.urun_duzenle[7] if self.urun_duzenle[7] is not None else 0.0
            self.sv_min_stok.set(f"{min_stok_val:.2f}".rstrip('0').rstrip('.'))
            print(f"{datetime.now()}: DEBUG: Min. Stok Seviyesi yüklendi: {self.sv_min_stok.get()}")


            # Nitelik Combobox'larının değer listelerini yükle ve sonra seçili değerleri ata.
            self._yukle_kategori_marka_comboboxlari()
            self._yukle_urun_grubu_birimi_ulke_comboboxlari()
            self._load_kategori_marka_fields() # Kategori ve Marka combobox'ları set et
            self._load_urun_grubu_birimi_ulke_fields() # Ürün Grubu, Birimi, Menşe combobox'ları set et

            print(f"{datetime.now()}: DEBUG: Nitelik combobox alanları yüklendi ve atandı.")

            print(f"{datetime.now()}: DEBUG: _load_genel_bilgiler tamamlandı.")
        else:
            print(f"{datetime.now()}: UYARI: _load_genel_bilgiler - self.urun_duzenle boş (Yeni Ürün). Varsayılan değerler __init__ içinde set edildi.")
            # Yeni ürün durumu için entry_urun_detayi'yi temizle
            if self.entry_urun_detayi:
                self.entry_urun_detayi.delete("1.0", tk.END)
            # Yeni ürün durumu için resim etiketini sıfırla
            if self.urun_resmi_label:
                self.urun_resmi_label.config(text="Resim Yok", image='')
    
    def _calculate_and_set_price(self, price_type, source_type, kdv_orani, input_value_str, target_sv): # target_entry yerine target_sv
        """Yardımcı fonksiyon: Fiyatı hesaplar ve ilgili StringVar'a yazar."""
        try:
            if not input_value_str.strip():
                target_sv.set(f"0{','.join(['0'] * 2)}" if 2 > 0 else "0") # Varsayılan 2 ondalık
                return

            value = float(input_value_str.replace(',', '.'))

            if source_type == 'haric':
                calculated_target = value * (1 + kdv_orani / 100)
            elif source_type == 'dahil':
                if (1 + kdv_orani / 100) == 0: 
                    calculated_target = 0.0
                else:
                    calculated_target = value / (1 + kdv_orani / 100)
    
            target_sv.set(f"{calculated_target:.2f}".replace('.',','))
        except ValueError:
            target_sv.set(f"0{','.join(['0'] * 2)}" if 2 > 0 else "0")
        except Exception as e:
            print(f"Otomatik fiyat doldurma hatası: {e}")
            
    def otomatik_fiyat_doldur(self, event=None, source_type=None, price_type=None, focus_out=False):
        try:
            kdv_orani_str = self.sv_kdv.get().strip().replace(',', '.')
            kdv_orani = float(kdv_orani_str) if kdv_orani_str else 0.0

            if not (0 <= kdv_orani <= 100):
                if focus_out:
                    messagebox.showwarning("Geçersiz KDV", "KDV Oranı 0 ile 100 arasında olmalıdır.", parent=self)
                self.sv_kdv.set("0")
                kdv_orani = 0.0

            if event is not None and event.keysym != "Tab" and not focus_out: 
                self.fiyat_degisiklik_tarihi_label.config(text=f"Fiyat Değişiklik Tarihi: {datetime.now().strftime('%d/%m/%Y')}")

            if source_type == 'kdv_focout' or price_type == 'all':
                self._calculate_and_set_price('alis', 'haric', kdv_orani, self.sv_alis_haric.get().strip(), self.sv_alis_dahil)
                self._calculate_and_set_price('alis', 'dahil', kdv_orani, self.sv_alis_dahil.get().strip(), self.sv_alis_haric)
                self._calculate_and_set_price('satis', 'haric', kdv_orani, self.sv_satis_haric.get().strip(), self.sv_satis_dahil)
                self._calculate_and_set_price('satis', 'dahil', kdv_orani, self.sv_satis_dahil.get().strip(), self.sv_satis_haric)
            elif price_type == 'alis':
                if source_type == 'haric':
                    self._calculate_and_set_price('alis', 'haric', kdv_orani, self.sv_alis_haric.get().strip(), self.sv_alis_dahil)
                elif source_type == 'dahil':
                    self._calculate_and_set_price('alis', 'dahil', kdv_orani, self.sv_alis_dahil.get().strip(), self.sv_alis_haric)
            elif price_type == 'satis':
                if source_type == 'haric':
                    self._calculate_and_set_price('satis', 'haric', kdv_orani, self.sv_satis_haric.get().strip(), self.sv_satis_dahil)
                elif source_type == 'dahil':
                    self._calculate_and_set_price('satis', 'dahil', kdv_orani, self.sv_satis_dahil.get().strip(), self.sv_satis_haric)
    
            self._calculate_kar_orani()

        except ValueError:
            if focus_out:
                self.sv_alis_haric.set("0,00")
                self.sv_alis_dahil.set("0,00")
                self.sv_satis_haric.set("0,00")
                self.sv_satis_dahil.set("0,00")
                self.label_kar_orani.config(text="0.00 %")
            pass
        except Exception as e:
            print(f"Otomatik fiyat doldurma hatası: {e}")
            
    def _format_stok_entry(self, event=None, sv_variable=None, decimal_places=2, focus_out=False):
        """
        Stok ve minimum stok giriş alanlarındaki değeri formatlar.
        FocusOut olayına özel olarak tasarlanmıştır.
        """
        if sv_variable is None:
            return

        current_value_str = sv_variable.get().strip()

        if not current_value_str or current_value_str == '-' or current_value_str == ',':
            sv_variable.set(f"0,{str('0' * decimal_places)}" if decimal_places > 0 else "0")
            return

        try:
            # Virgülü noktaya çevirerek float'a dönüştür
            value_float = float(current_value_str.replace(',', '.'))
            # İstenen ondalık basamak sayısına göre formatla
            formatted_value_str = f"{{:.{decimal_places}f}}".format(value_float)
            # Noktayı tekrar virgüle çevir
            final_display_value = formatted_value_str.replace('.', ',').rstrip('0').rstrip(',')
            if final_display_value == "": # Eğer sadece . veya , kalırsa sıfıra çek
                 final_display_value = "0" if decimal_places == 0 else "0,00"
            if final_display_value == "-":
                 final_display_value = "0" if decimal_places == 0 else "0,00"


            sv_variable.set(final_display_value)
        except ValueError:
            # Geçersiz bir değer girildiyse sıfırla
            sv_variable.set(f"0,{str('0' * decimal_places)}" if decimal_places > 0 else "0")
        except Exception as e:
            print(f"Hata: _format_stok_entry - {e}")
            sv_variable.set(f"0,{str('0' * decimal_places)}" if decimal_places > 0 else "0")


    def _on_notebook_tab_change(self, event):
        selected_tab_id = self.main_notebook.select()
        selected_tab_widget = self.main_notebook.nametowidget(selected_tab_id)
        selected_tab_text = self.main_notebook.tab(selected_tab_id, "text")
        
        if selected_tab_text == "Stok Hareketleri":
            if self.urun_id:
                selected_tab_widget.urun_id = self.urun_id
                selected_tab_widget.urun_adi = self.urun_adi_initial
                selected_tab_widget._load_stok_hareketleri()
        elif selected_tab_text == "İlgili Faturalar":
            if self.urun_id:
                selected_tab_widget.urun_id = self.urun_id
                selected_tab_widget.urun_adi = self.urun_adi_initial
                selected_tab_widget._load_ilgili_faturalar()
        elif selected_tab_text == "Kategori & Marka Yönetimi": 
            if hasattr(selected_tab_widget, '_kategori_listesini_yukle'):
                selected_tab_widget._kategori_listesini_yukle()
            if hasattr(selected_tab_widget, '_marka_listesini_yukle'):
                selected_tab_widget._marka_listesini_yukle()
        elif selected_tab_text == "Ürün Nitelik Yönetimi":
            if hasattr(selected_tab_widget, '_urun_grubu_listesini_yukle'):
                selected_tab_widget._urun_grubu_listesini_yukle()
                selected_tab_widget._urun_birimi_listesini_yukle()
                selected_tab_widget._ulke_listesini_yukle()

    def kaydet(self):
        kod = self.sv_kod.get().strip()
        ad = self.sv_ad.get().strip()
        urun_detayi = self.entry_urun_detayi.get("1.0", tk.END).strip()
        stok_str = self.sv_stok.get().strip()
        kdv_str = self.sv_kdv.get().strip()
        min_stok_str = self.sv_min_stok.get().strip()
    
        alis_haric_str = self.sv_alis_haric.get().strip()
        alis_dahil_str = self.sv_alis_dahil.get().strip()
        satis_haric_str = self.sv_satis_haric.get().strip()
        satis_dahil_str = self.sv_satis_dahil.get().strip()
    
        # DÜZELTME: Combobox'lardan değerleri alırken "Seçim Yok" kontrolü
        urun_grubu_id = self.urun_gruplari_map.get(self.combo_urun_grubu.get(), None)
        if self.combo_urun_grubu.get() == "Seçim Yok": urun_grubu_id = None

        urun_birimi_id = self.urun_birimleri_map.get(self.combo_urun_birimi.get(), None)
        if self.combo_urun_birimi.get() == "Seçim Yok": urun_birimi_id = None

        ulke_id = self.ulkeler_map.get(self.combo_mense.get(), None)
        if self.combo_mense.get() == "Seçim Yok": ulke_id = None

        fiyat_degisiklik_tarihi_str = self.fiyat_degisiklik_tarihi_label.cget("text").strip()
        if fiyat_degisiklik_tarihi_str == "Yükleniyor..." or fiyat_degisiklik_tarihi_str == "Hata" or fiyat_degisiklik_tarihi_str == "-" or not fiyat_degisiklik_tarihi_str:
            fiyat_degisiklik_tarihi_str = datetime.now().strftime('%Y-%m-%d')
        else:
            try:
                # Fiyat değişiklik tarihi YYYY-MM-DD formatında kaydedilmeli
                fiyat_degisiklik_tarihi_str = datetime.strptime(fiyat_degisiklik_tarihi_str, '%d.%m.%Y').strftime('%Y-%m-%d')
            except ValueError:
                # Eğer format farklıysa veya tanımsızsa bugünün tarihini kullan
                fiyat_degisiklik_tarihi_str = datetime.now().strftime('%Y-%m-%d')
    
        if not (kod and ad):
            messagebox.showerror("Eksik Bilgi", "Ürün Kodu ve Adı boş bırakılamaz.", parent=self)
            return
    
        try:
            stok = float(stok_str.replace(',', '.')) if stok_str else 0.0
            kdv = float(kdv_str.replace(',', '.')) if kdv_str else 0.0
            min_stok = float(min_stok_str.replace(',', '.')) if min_stok_str else 0.0
    
            alis_haric = float(alis_haric_str.replace(',', '.')) if alis_haric_str else 0.0
            alis_dahil = float(alis_dahil_str.replace(',', '.')) if alis_dahil_str else 0.0
            satis_haric = float(satis_haric_str.replace(',', '.')) if satis_haric_str else 0.0
            satis_dahil = float(satis_dahil_str.replace(',', '.')) if satis_dahil_str else 0.0
    
            if not (0 <= kdv <= 100):
                 messagebox.showerror("Geçersiz Değer", "KDV Oranı 0 ile 100 arasında olmalıdır.", parent=self)
                 return
    
        except ValueError:
            messagebox.showerror("Giriş Hatası","Sayısal alanlar doğru formatta olmalıdır.", parent=self)
            return False

        selected_kategori_name = self.combo_kategori.get()
        kategori_id_to_save = self.kategoriler_map.get(selected_kategori_name, None)
        if selected_kategori_name == "Seçim Yok" or kategori_id_to_save is None:
            kategori_id_to_save = None
    
        selected_marka_name = self.combo_marka.get()
        marka_id_to_save = self.markalar_map.get(selected_marka_name, None)
        if selected_marka_name == "Seçim Yok" or marka_id_to_save is None:
            marka_id_to_save = None
    
        urun_detayi_to_save = urun_detayi if urun_detayi else None
        urun_resmi_yolu_to_save = self.urun_resmi_path if self.urun_resmi_path else None
    
        if self.urun_id: 
            success, message = self.db.stok_guncelle(self.urun_id, kod, ad, stok, alis_haric, satis_haric, kdv, min_stok,
                                     alis_dahil, satis_dahil, kategori_id_to_save, marka_id_to_save,
                                     urun_detayi_to_save, urun_resmi_yolu_to_save, fiyat_degisiklik_tarihi_str,
                                     urun_grubu_id, urun_birimi_id, ulke_id) 
            if success:
                messagebox.showinfo("Başarılı", message, parent=self) # db'den gelen mesajı kullan
                if self.app: self.app.set_status(message) # Durum çubuğunu güncelle
                self.yenile_callback()
                try: self.grab_release()
                except tk.TclError: pass
                self.destroy()
            else:
                messagebox.showerror("Hata", message, parent=self) # db'den gelen hata mesajını göster
        else: 
            result_tuple = self.db.stok_ekle(kod, ad, stok, alis_haric, satis_haric, kdv, min_stok,
                                        alis_dahil, satis_dahil, kategori_id_to_save, marka_id_to_save,
                                        urun_detayi_to_save, urun_resmi_yolu_to_save, fiyat_degisiklik_tarihi_str,
                                        urun_grubu_id, urun_birimi_id, ulke_id)
            success, message_or_id = result_tuple
            if success:
                yeni_id = message_or_id
                messagebox.showinfo("Başarılı", f"'{ad}' ürünü eklendi.", parent=self)
                if self.app: self.app.set_status(f"Yeni ürün '{ad}' eklendi (ID: {yeni_id}).")
                self.yenile_callback()
                try: self.grab_release()
                except tk.TclError: pass
                self.destroy()
            else:
                messagebox.showerror("Hata", message_or_id, parent=self)

    def _calculate_kar_orani(self):
        try:
            alis_fiyati_dahil_str = self.sv_alis_dahil.get().strip().replace(',', '.')
            satis_fiyati_dahil_str = self.sv_satis_dahil.get().strip().replace(',', '.')

            alis_fiyati = float(alis_fiyati_dahil_str) if alis_fiyati_dahil_str else 0.0
            satis_fiyati = float(satis_fiyati_dahil_str) if satis_fiyati_dahil_str else 0.0
    
            if alis_fiyati > 0:
                kar_orani = ((satis_fiyati - alis_fiyati) / alis_fiyati) * 100
                self.label_kar_orani.config(text=f"{kar_orani:,.2f} %")
            else:
                self.label_kar_orani.config(text="0.00 %")
        except ValueError:
            self.label_kar_orani.config(text="Hesaplanamıyor")
        except Exception as e:
            print(f"Kar oranı hesaplanırken hata: {e}")
            self.label_kar_orani.config(text="Hata")

    def _setup_ilgili_faturalar_tab(self, parent_frame):
        ttk.Label(parent_frame, text="Ürünün Yer Aldığı Faturalar", font=("Segoe UI", 12, "bold")).pack(pady=5, anchor=tk.W)

        filter_frame = ttk.Frame(parent_frame, padding="5")
        filter_frame.pack(fill=tk.X, padx=0, pady=5)

        ttk.Label(filter_frame, text="Fatura Tipi:").pack(side=tk.LEFT, padx=(0,2))
        self.fatura_tipi_filter_cb = ttk.Combobox(filter_frame, width=15, values=["TÜMÜ", "ALIŞ", "SATIŞ"], state="readonly")
        self.fatura_tipi_filter_cb.pack(side=tk.LEFT, padx=(0,10))
        self.fatura_tipi_filter_cb.set("TÜMÜ")
        self.fatura_tipi_filter_cb.bind("<<ComboboxSelected>>", self._load_ilgili_faturalar)

        ttk.Button(filter_frame, text="Filtrele", command=self._load_ilgili_faturalar, style="Accent.TButton").pack(side=tk.LEFT)


        cols_fatura = ("ID", "Fatura No", "Tarih", "Tip", "Cari/Misafir", "KDV Hariç Top.", "KDV Dahil Top.")
        self.ilgili_faturalar_tree = ttk.Treeview(parent_frame, columns=cols_fatura, show='headings', selectmode="browse")

        col_defs_fatura = [
            ("ID", 40, tk.E, tk.NO),
            ("Fatura No", 120, tk.W, tk.YES),
            ("Tarih", 85, tk.CENTER, tk.NO),
            ("Tip", 70, tk.CENTER, tk.NO),
            ("Cari/Misafir", 200, tk.W, tk.YES),
            ("KDV Hariç Top.", 120, tk.E, tk.NO),
            ("KDV Dahil Top.", 120, tk.E, tk.NO)
        ]
        for col_name, width, anchor, stretch_opt in col_defs_fatura:
            self.ilgili_faturalar_tree.column(col_name, width=width, anchor=anchor, stretch=stretch_opt)
            self.ilgili_faturalar_tree.heading(col_name, text=col_name, command=lambda c=col_name: sort_treeview_column(self.ilgili_faturalar_tree, c, False))

        vsb_fatura = ttk.Scrollbar(parent_frame, orient="vertical", command=self.ilgili_faturalar_tree.yview)
        hsb_fatura = ttk.Scrollbar(parent_frame, orient="horizontal", command=self.ilgili_faturalar_tree.xview)
        self.ilgili_faturalar_tree.configure(yscrollcommand=vsb_fatura.set, xscrollcommand=hsb_fatura.set)
        vsb_fatura.pack(side=tk.RIGHT, fill=tk.Y)
        hsb_fatura.pack(side=tk.BOTTOM, fill=tk.X)
        self.ilgili_faturalar_tree.pack(expand=True, fill=tk.BOTH)

        self.ilgili_faturalar_tree.bind("<Double-1>", self._on_fatura_double_click)

    def _load_ilgili_faturalar(self, event=None):
        for i in self.ilgili_faturalar_tree.get_children():
            self.ilgili_faturalar_tree.delete(i)

        if not self.urun_id:
            self.ilgili_faturalar_tree.insert("", tk.END, values=("", "", "", "", "Ürün seçili değil.", "", ""))
            return

        fatura_tipi_filtre = self.fatura_tipi_filter_cb.get()
        
        faturalar = self.db.get_faturalar_by_urun_id(self.urun_id, fatura_tipi=fatura_tipi_filtre)

        if not faturalar:
            self.ilgili_faturalar_tree.insert("", tk.END, values=("", "", "", "", "Bu ürüne ait fatura bulunamadı.", "", ""))
            return

        for fatura_item in faturalar:
            fatura_id = fatura_item[0]
            fatura_no = fatura_item[1]
            tarih_str = fatura_item[2]
            fatura_tip = fatura_item[3]
            cari_adi = fatura_item[4]
            toplam_kdv_haric = fatura_item[5]
            toplam_kdv_dahil = fatura_item[6]

            try:
                formatted_tarih = datetime.strptime(tarih_str, '%Y-%m-%d').strftime('%d.%m.%Y')
            except ValueError:
                formatted_tarih = tarih_str

            self.ilgili_faturalar_tree.insert("", tk.END, iid=fatura_id, values=(
                fatura_id,
                fatura_no,
                formatted_tarih,
                fatura_tip,
                cari_adi,
                self.db._format_currency(toplam_kdv_haric),
                self.db._format_currency(toplam_kdv_dahil)
            ))

    def _on_fatura_double_click(self, event):
        selected_item_iid = self.ilgili_faturalar_tree.focus()
        if not selected_item_iid:
            return
        
        fatura_id = self.ilgili_faturalar_tree.item(selected_item_iid)['values'][0]
        if fatura_id:
            FaturaDetayPenceresi(self.app, self.db, fatura_id)

class YeniKasaBankaEklePenceresi(tk.Toplevel):
    def __init__(self, parent, db_manager, yenile_callback, hesap_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.hesap_duzenle_id = hesap_duzenle[0] if hesap_duzenle else None
        self.app = app_ref

        self.title("Yeni Kasa/Banka Hesabı Ekle" if not hesap_duzenle else "Hesap Düzenle")
        self.geometry("480x450")
        self.transient(parent)
        self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        ttk.Label(main_frame, text=self.title(), font=("Segoe UI", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0,15))

        labels_entries_kb = {
            "Hesap Adı (*):": "entry_hesap_adi",
            "Hesap Tipi (*):": "combo_tip",
            "Banka Adı (Banka ise):": "entry_banka_adi",
            "Şube Adı (Banka ise):": "entry_sube_adi",
            "Hesap No/IBAN (Banka ise):": "entry_hesap_no",
            "Açılış Bakiyesi:": "entry_bakiye",
            "Para Birimi:": "entry_para_birimi",
            "Açılış Tarihi (YYYY-AA-GG):": "entry_acilis_tarihi",
            "Varsayılan Ödeme Türü:": "combo_varsayilan_odeme_turu"
        }
        self.entries_kb = {}
        row_idx = 1
        for label_text, entry_name in labels_entries_kb.items():
            ttk.Label(main_frame, text=label_text).grid(row=row_idx, column=0, padx=5, pady=7, sticky=tk.W)
            if entry_name == "combo_tip":
                self.entries_kb[entry_name] = ttk.Combobox(main_frame, values=["KASA", "BANKA"], state="readonly", width=25)
                self.entries_kb[entry_name].bind("<<ComboboxSelected>>", self.tip_degisince_banka_alanlarini_ayarla)
            elif entry_name == "combo_varsayilan_odeme_turu":
                self.entries_kb[entry_name] = ttk.Combobox(main_frame, values=["YOK", "NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"], state="readonly", width=25)
                self.entries_kb[entry_name].set("YOK")
            else:
                self.entries_kb[entry_name] = ttk.Entry(main_frame, width=30)
            self.entries_kb[entry_name].grid(row=row_idx, column=1, padx=5, pady=7, sticky=tk.EW)

            if entry_name == "entry_acilis_tarihi":
                setup_date_entry(self.app, self.entries_kb["entry_acilis_tarihi"])
                ttk.Button(main_frame, text="🗓️", command=lambda: self._open_date_picker(self.entries_kb["entry_acilis_tarihi"]), width=3).grid(row=row_idx, column=2, padx=2, pady=7, sticky=tk.W)

            row_idx += 1
        main_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(2, weight=0) # Takvim butonu sütunu

        self.entries_kb["entry_bakiye"].insert(0, "0,00")
        self.entries_kb["entry_para_birimi"].insert(0, "TL")
        self.entries_kb["combo_tip"].current(0)
        self.tip_degisince_banka_alanlarini_ayarla()

        if hesap_duzenle:
            self.entries_kb["entry_hesap_adi"].insert(0, hesap_duzenle[1])
            self.entries_kb["combo_tip"].set(hesap_duzenle[5])
            self.entries_kb["entry_banka_adi"].insert(0, hesap_duzenle[7] or "")
            self.entries_kb["entry_sube_adi"].insert(0, hesap_duzenle[8] or "")
            self.entries_kb["entry_hesap_no"].insert(0, hesap_duzenle[2] or "")
            self.entries_kb["entry_bakiye"].delete(0, tk.END)
            self.entries_kb["entry_bakiye"].insert(0, f"{hesap_duzenle[3]:.2f}".replace('.',','))
            self.entries_kb["entry_para_birimi"].delete(0, tk.END)
            self.entries_kb["entry_para_birimi"].insert(0, hesap_duzenle[4])
            self.entries_kb["entry_acilis_tarihi"].insert(0, hesap_duzenle[6] or "")
            self.tip_degisince_banka_alanlarini_ayarla()
            varsayilan_odeme_turu_db = hesap_duzenle[9] if len(hesap_duzenle) > 9 and hesap_duzenle[9] else "YOK"
            self.entries_kb["combo_varsayilan_odeme_turu"].set(varsayilan_odeme_turu_db)
            # DEĞİŞİKLİK BİTİŞİ

        button_frame_kb_alt = ttk.Frame(main_frame)
        button_frame_kb_alt.grid(row=row_idx, column=0, columnspan=3, pady=(15,0), sticky=tk.E)
        # DEĞİŞİKLİK BİTİŞİ

        ttk.Button(button_frame_kb_alt, text="Kaydet", command=self.kaydet_hesap, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame_kb_alt, text="İptal", command=self.destroy).pack(side=tk.LEFT)

    def _open_date_picker(self, target_entry):
        """Bir Entry widget'ı için tarih seçici penceresi açar."""
        from yardimcilar import DatePickerDialog
        DatePickerDialog(self.app, target_entry)
        self.app.set_status("Tarih seçici açıldı.")

    def tip_degisince_banka_alanlarini_ayarla(self, event=None):
        secili_tip = self.entries_kb["combo_tip"].get()
        banka_alanlari = ["entry_banka_adi", "entry_sube_adi", "entry_hesap_no"]
        for alan_adi in banka_alanlari:
            self.entries_kb[alan_adi].config(state=tk.NORMAL if secili_tip == "BANKA" else tk.DISABLED)

        if secili_tip != "BANKA":
            for alan_adi in banka_alanlari:
                self.entries_kb[alan_adi].delete(0, tk.END)

    def kaydet_hesap(self):
        h_adi = self.entries_kb["entry_hesap_adi"].get().strip()
        h_tip = self.entries_kb["combo_tip"].get()
        b_adi = self.entries_kb["entry_banka_adi"].get().strip() if h_tip == "BANKA" else None
        s_adi = self.entries_kb["entry_sube_adi"].get().strip() if h_tip == "BANKA" else None
        h_no = self.entries_kb["entry_hesap_no"].get().strip() if h_tip == "BANKA" else None
        bakiye_str = self.entries_kb["entry_bakiye"].get().strip()
        p_birimi = self.entries_kb["entry_para_birimi"].get().strip()
        a_tarihi = self.entries_kb["entry_acilis_tarihi"].get().strip() or None
        varsayilan_odeme_turu_secilen = self.entries_kb["combo_varsayilan_odeme_turu"].get()
        varsayilan_odeme_turu_to_db = None if varsayilan_odeme_turu_secilen == "YOK" else varsayilan_odeme_turu_secilen

        if not (h_adi and h_tip):
            messagebox.showerror("Eksik Bilgi", "Hesap Adı ve Hesap Tipi zorunludur.", parent=self)
            return

        if self.hesap_duzenle_id:
            success, message = self.db.kasa_banka_guncelle(self.hesap_duzenle_id, h_adi, h_no, bakiye_str, p_birimi, h_tip, a_tarihi, b_adi, s_adi, varsayilan_odeme_turu_to_db)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                if self.app: self.app.set_status(message)
                self.yenile_callback()
                self.destroy()
            else:
                messagebox.showerror("Hata", message, parent=self)
        else:
            success, message_or_id = self.db.kasa_banka_ekle(h_adi, h_no, bakiye_str, p_birimi, h_tip, a_tarihi, b_adi, s_adi, varsayilan_odeme_turu_to_db)
            if success:
                messagebox.showinfo("Başarılı", message_or_id, parent=self)
                if self.app: self.app.set_status(message_or_id)
                self.yenile_callback()
                self.destroy()
            else:
                messagebox.showerror("Hata", message_or_id, parent=self)
class YeniTedarikciEklePenceresi(tk.Toplevel):
    def __init__(self, parent, db_manager, yenile_callback, tedarikci_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.tedarikci_duzenle_id = tedarikci_duzenle[0] if tedarikci_duzenle else None
        self.app = app_ref

        self.title("Yeni Tedarikçi Ekle" if not tedarikci_duzenle else "Tedarikçi Düzenle")
        self.geometry("500x420") # <-- DÜZELTME: Pencere boyutu ayarlandı
        self.transient(parent) # Ana pencerenin üzerinde kalır
        self.grab_set() # Diğer pencerelere tıklamayı engeller

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        ttk.Label(main_frame, text=self.title(), font=("Segoe UI", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0,15))

        labels_entries = {
            "Tedarikçi Kodu:": "entry_kod",
            "Ad Soyad:": "entry_ad",
            "Telefon:": "entry_tel",
            "Adres:": "entry_adres",
            "Vergi Dairesi:": "entry_vd",
            "Vergi No:": "entry_vn"
        }
        self.entries = {} # Entry widget'larını saklamak için sözlük

        for i, (label_text, entry_name) in enumerate(labels_entries.items(), 1):
            ttk.Label(main_frame, text=label_text).grid(row=i, column=0, padx=5, pady=8, sticky=tk.W)
            if entry_name == "entry_adres":
                self.entries[entry_name] = tk.Text(main_frame, height=3, width=30) # <-- DÜZELTME: Genişlik 30
            else:
                self.entries[entry_name] = ttk.Entry(main_frame, width=30) # <-- DÜZELTME: Genişlik 30
            self.entries[entry_name].grid(row=i, column=1, padx=5, pady=8, sticky=tk.EW)
        
        main_frame.columnconfigure(1, weight=1) # Entry'lerin genişlemesi için

        # Tedarikçi kodu otomatik oluşturulacak ve düzenlenemez olacak
        if not tedarikci_duzenle: # Sadece yeni tedarikçi eklerken kodu otomatik oluştur
            generated_code = self.db.get_next_tedarikci_kodu() # veritabani.py'den metodu çağır
            self.entries["entry_kod"].insert(0, generated_code)
            self.entries["entry_kod"].config(state=tk.DISABLED) # Otomatik kodu düzenlenemez yap
        else: # Düzenleme modu
            # tedarikci_duzenle: (id, tedarikci_kodu, ad, telefon, adres, vergi_dairesi, vergi_no)
            self.entries["entry_kod"].insert(0, tedarikci_duzenle[1])
            self.entries["entry_ad"].insert(0, tedarikci_duzenle[2])
            self.entries["entry_tel"].insert(0, tedarikci_duzenle[3] if tedarikci_duzenle[3] else "")
            if isinstance(self.entries["entry_adres"], tk.Text):
                self.entries["entry_adres"].insert("1.0", tedarikci_duzenle[4] if tedarikci_duzenle[4] else "")
            self.entries["entry_vd"].insert(0, tedarikci_duzenle[5] if tedarikci_duzenle[5] else "")
            self.entries["entry_vn"].insert(0, tedarikci_duzenle[6] if tedarikci_duzenle[6] else "")

            # Düzenleme modunda da tedarikçi kodunu düzenlenemez yapıyoruz
            self.entries["entry_kod"].config(state=tk.DISABLED) 

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=len(labels_entries)+1, column=0, columnspan=2, pady=(20,0), sticky=tk.E)
        ttk.Button(button_frame, text="Kaydet", command=self.kaydet, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.LEFT)
        
    def kaydet(self):
        kod = self.entries["entry_kod"].get().strip() 
        ad = self.entries["entry_ad"].get().strip()
        tel = self.entries["entry_tel"].get().strip()
        adres = self.entries["entry_adres"].get("1.0", tk.END).strip() if isinstance(self.entries["entry_adres"], tk.Text) else ""
        vd = self.entries["entry_vd"].get().strip()
        vn = self.entries["entry_vn"].get().strip()

        if not (kod and ad):
            messagebox.showerror("Eksik Bilgi", "Tedarikçi Kodu ve Ad Soyad boş bırakılamaz.", parent=self)
            return

        if self.tedarikci_duzenle_id: # Güncelleme işlemi
            success, message = self.db.tedarikci_guncelle(self.tedarikci_duzenle_id, kod, ad, tel, adres, vd, vn)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                if self.app: self.app.set_status(message)
                self.yenile_callback()
                self.destroy()
            else:
                messagebox.showerror("Hata", message, parent=self)
        else: 
            success, message_or_id = self.db.tedarikci_ekle(kod, ad, tel, adres, vd, vn)
            if success:
            
                messagebox.showinfo("Başarılı", f"'{ad}' tedarikçisi başarıyla eklendi (ID: {message_or_id}).", parent=self)
                if self.app: self.app.set_status(f"Yeni tedarikçi '{ad}' eklendi (ID: {message_or_id}).")
                self.yenile_callback()
                self.destroy()
            else:
                messagebox.showerror("Hata", message_or_id, parent=self)

class YeniMusteriEklePenceresi(tk.Toplevel):
    def __init__(self, parent, db_manager, yenile_callback, musteri_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.musteri_duzenle_id = musteri_duzenle[0] if musteri_duzenle else None
        self.app = app_ref

        # Eğer müşteri düzenleniyorsa ve ID'si perakende müşteri ID'si ile aynıysa True olur.
        self.is_perakende_duzenleme = (musteri_duzenle and str(self.musteri_duzenle_id) == str(self.db.perakende_musteri_id))

        self.title("Yeni Müşteri Ekle" if not musteri_duzenle else ("Perakende Müşteri Düzenle" if self.is_perakende_duzenleme else "Müşteri Düzenle"))
        self.geometry("500x420") # <-- DÜZELTME: Pencere boyutu ayarlandı
        self.transient(parent); self.grab_set()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)
        ttk.Label(main_frame, text=self.title(), font=("Segoe UI", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0,15))

        labels_entries = {
            "Müşteri Kodu:": "entry_kod", "Ad Soyad:": "entry_ad", "Telefon:": "entry_tel",
            "Adres:": "entry_adres", "Vergi Dairesi:": "entry_vd", "Vergi No:": "entry_vn"
        }
        self.entries = {}

        for i, (label_text, entry_name) in enumerate(labels_entries.items(), 1):
            ttk.Label(main_frame, text=label_text).grid(row=i, column=0, padx=5, pady=8, sticky=tk.W)
            if entry_name == "entry_adres":
                self.entries[entry_name] = tk.Text(main_frame, height=3, width=30) # <-- DÜZELTME: Genişlik 30
            else:
                self.entries[entry_name] = ttk.Entry(main_frame, width=30) # <-- DÜZELTME: Genişlik 30
            self.entries[entry_name].grid(row=i, column=1, padx=5, pady=8, sticky=tk.EW)
        
        main_frame.columnconfigure(1, weight=1)

        # Müşteri kodu otomatik oluşturulacak ve düzenlenemez olacak
        if not musteri_duzenle: # Sadece yeni müşteri eklerken kodu otomatik oluştur
            generated_code = self.db.get_next_musteri_kodu() # Yeni metodu çağır
            self.entries["entry_kod"].insert(0, generated_code)
            self.entries["entry_kod"].config(state=tk.DISABLED) # Otomatik kodu düzenlenemez yap
        else: # Düzenleme modu
            # musteri_duzenle: (id, musteri_kodu, ad, telefon, adres, vergi_dairesi, vergi_no)
            self.entries["entry_kod"].insert(0, musteri_duzenle[1])
            self.entries["entry_ad"].insert(0, musteri_duzenle[2])
            self.entries["entry_tel"].insert(0, musteri_duzenle[3] if musteri_duzenle[3] else "")
            if isinstance(self.entries["entry_adres"], tk.Text):
                self.entries["entry_adres"].insert("1.0", musteri_duzenle[4] if musteri_duzenle[4] else "")
            self.entries["entry_vd"].insert(0, musteri_duzenle[5] if musteri_duzenle[5] else "")
            self.entries["entry_vn"].insert(0, musteri_duzenle[6] if musteri_duzenle[6] else "")

            # Düzenleme modunda da müşteri kodunu düzenlenemez yapıyoruz
            self.entries["entry_kod"].config(state=tk.DISABLED) 
            
            # DEĞİŞİKLİK BAŞLANGICI: Perakende müşterinin alanlarını kısıtlama
            if self.is_perakende_duzenleme:
                # Sadece misafir adı ve kodu düzenlenebilir olmalı, diğerleri kilitli.
                self.entries["entry_tel"].config(state=tk.DISABLED)
                if isinstance(self.entries["entry_adres"], tk.Text): # Text widget'ı için ayrı kontrol
                    self.entries["entry_adres"].config(state=tk.DISABLED)
                else: # Entry widget'ı için
                    self.entries["entry_adres"].config(state=tk.DISABLED)
                self.entries["entry_vd"].config(state=tk.DISABLED)
                self.entries["entry_vn"].config(state=tk.DISABLED)
            # DEĞİŞİKLİK BİTİŞİ

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=len(labels_entries)+1, column=0, columnspan=2, pady=(20,0), sticky=tk.E)
        ttk.Button(button_frame, text="Kaydet", command=self.kaydet, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.LEFT)

    def kaydet(self):
        kod = self.entries["entry_kod"].get().strip() 
        ad = self.entries["entry_ad"].get().strip()
        tel = self.entries["entry_tel"].get().strip()
        adres = self.entries["entry_adres"].get("1.0", tk.END).strip() if isinstance(self.entries["entry_adres"], tk.Text) else ""
        vd = self.entries["entry_vd"].get().strip()
        vn = self.entries["entry_vn"].get().strip()

        if not (kod and ad):
            messagebox.showerror("Eksik Bilgi", "Müşteri Kodu ve Ad Soyad boş bırakılamaz.", parent=self)
            return

        if self.is_perakende_duzenleme:
            kod = self.db.PERAKENDE_MUSTERI_KODU

        if self.musteri_duzenle_id: # Güncelleme işlemi
            success, message = self.db.musteri_guncelle(self.musteri_duzenle_id, kod, ad, tel, adres, vd, vn)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                if self.app: self.app.set_status(message)
                self.yenile_callback()
                self.destroy()
            else:
                messagebox.showerror("Hata", message, parent=self)
        else: 
            success, message_or_id = self.db.musteri_ekle(kod, ad, tel, adres, vd, vn)
            if success:
            
                messagebox.showinfo("Başarılı", f"'{ad}' müşterisi başarıyla eklendi (ID: {message_or_id}).", parent=self)
                if self.app: self.app.set_status(f"Yeni müşteri '{ad}' eklendi (ID: {message_or_id}).")
                self.yenile_callback()
                self.destroy()
            else:
                messagebox.showerror("Hata", message_or_id, parent=self)

class KalemDuzenlePenceresi(tk.Toplevel):
    def __init__(self, parent_page, kalem_index, kalem_verisi, islem_tipi, fatura_id_duzenle=None):
        super().__init__(parent_page)
        self.parent_page = parent_page
        self.db = parent_page.db
        self.kalem_index = kalem_index
        self.islem_tipi = islem_tipi # SATIŞ, ALIŞ, SATIŞ_SIPARIS, ALIŞ_SIPARIS
        self.fatura_id_duzenle = fatura_id_duzenle # Fatura veya Sipariş ID'si

        self.urun_id = kalem_verisi[0]
        self.urun_adi = kalem_verisi[1]
        self.mevcut_miktar = kalem_verisi[2]
        self.orijinal_birim_fiyat_kdv_haric = kalem_verisi[3]
        self.kdv_orani = kalem_verisi[4]
        self.mevcut_alis_fiyati_fatura_aninda = kalem_verisi[8] # Bu, fatura anı alış fiyatı (KDV Dahil)
        self.kdv_orani_fatura_aninda_db = kalem_verisi[9] # Bu, fatura anı KDV oranı (kaydedilen)

        self.initial_iskonto_yuzde_1 = kalem_verisi[10]
        self.initial_iskonto_yuzde_2 = kalem_verisi[11]

        # KDV Hariç fiyatı ve KDV oranını kullanarak KDV Dahil orijinal fiyatı hesapla
        self.orijinal_birim_fiyat_kdv_dahil = self.orijinal_birim_fiyat_kdv_haric * (1 + self.kdv_orani / 100)
        
        self.title(f"Kalem Düzenle: {self.urun_adi}")
        self.geometry("450x550")
        self.transient(parent_page)
        self.grab_set()
        self.resizable(False, False)

        self.sv_miktar = tk.StringVar(self)
        self.sv_fiyat = tk.StringVar(self) # Birim Fiyat (KDV Dahil) için
        self.sv_alis_fiyati_aninda = tk.StringVar(self) # Fatura Anı Alış Fiyatı için (sadece satış faturası için geçerli)
        self.sv_iskonto_yuzde_1 = tk.StringVar(self)
        self.sv_iskonto_yuzde_2 = tk.StringVar(self)
 
        main_f = ttk.Frame(self, padding="15")
        main_f.pack(expand=True, fill=tk.BOTH)

        ttk.Label(main_f, text=f"Ürün: {self.urun_adi}", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=3, pady=5, sticky=tk.W)
        main_f.columnconfigure(1, weight=1)

        current_row = 1
        ttk.Label(main_f, text="Miktar:").grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.miktar_e = ttk.Entry(main_f, width=15, textvariable=self.sv_miktar)
        self.miktar_e.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        self.sv_miktar.set(f"{self.mevcut_miktar:.2f}".rstrip('0').rstrip('.'))

        setup_numeric_entry(self.parent_page.app, self.miktar_e, decimal_places=2) # Yenilenmiş çağrı
        self.miktar_e.bind("<KeyRelease>", self._anlik_hesaplama_ve_guncelleme) # KeyRelease'i manuel bırakın

        current_row += 1

        ttk.Label(main_f, text="Birim Fiyat (KDV Dahil):").grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.fiyat_e = ttk.Entry(main_f, width=15, textvariable=self.sv_fiyat)
        self.fiyat_e.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        self.sv_fiyat.set(f"{self.orijinal_birim_fiyat_kdv_dahil:.2f}".replace('.',','))

        setup_numeric_entry(self.parent_page.app, self.fiyat_e, decimal_places=2) # Yenilenmiş çağrı
        self.fiyat_e.bind("<KeyRelease>", self._anlik_hesaplama_ve_guncelleme) # KeyRelease'i manuel bırakın

        current_row += 1

        if self.islem_tipi == 'SATIŞ' or self.islem_tipi == 'SATIŞ_SIPARIS':
            ttk.Label(main_f, text="Fatura Anı Alış Fiyatı (KDV Dahil):").grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
            self.alis_fiyati_aninda_e = ttk.Entry(main_f, width=15, textvariable=self.sv_alis_fiyati_aninda)
            self.alis_fiyati_aninda_e.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
            self.sv_alis_fiyati_aninda.set(f"{self.mevcut_alis_fiyati_fatura_aninda:.2f}".replace('.',','))

            setup_numeric_entry(self.parent_page.app, self.alis_fiyati_aninda_e, decimal_places=2) # Yenilenmiş çağrı
            self.alis_fiyati_aninda_e.bind("<KeyRelease>", self._anlik_hesaplama_ve_guncelleme) # KeyRelease'i manuel bırakın

            current_row += 1
        else:
            self.alis_fiyati_aninda_e = None
            self.sv_alis_fiyati_aninda.set("0,00")

        ttk.Separator(main_f, orient='horizontal').grid(row=current_row, column=0, columnspan=3, sticky='ew', pady=(10,5))
        current_row += 1

        ttk.Label(main_f, text="İskonto 1 (%):").grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.iskonto_yuzde_1_e = ttk.Entry(main_f, width=10, textvariable=self.sv_iskonto_yuzde_1)
        self.iskonto_yuzde_1_e.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        self.sv_iskonto_yuzde_1.set(f"{self.initial_iskonto_yuzde_1:.2f}".replace('.',','))

        setup_numeric_entry(self.parent_page.app, self.iskonto_yuzde_1_e, decimal_places=2) # Yenilenmiş çağrı
        self.iskonto_yuzde_1_e.bind("<KeyRelease>", self._anlik_hesaplama_ve_guncelleme) # KeyRelease'i manuel bırakın

        ttk.Label(main_f, text="%").grid(row=current_row, column=2, padx=(0,5), pady=8, sticky=tk.W)
        current_row += 1

        ttk.Label(main_f, text="İskonto 2 (%):").grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.iskonto_yuzde_2_e = ttk.Entry(main_f, width=10, textvariable=self.sv_iskonto_yuzde_2)
        self.iskonto_yuzde_2_e.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        self.sv_iskonto_yuzde_2.set(f"{self.initial_iskonto_yuzde_2:.2f}".replace('.',','))

        vcmd_iskonto2 = (self.register(lambda P, S, e=self.iskonto_yuzde_2_e, a_n=False: self._validate_numeric_input_urun_karti(P, S, e, a_n)), '%P', '%S')
        self.iskonto_yuzde_2_e.config(validate="key", validatecommand=vcmd_iskonto2)
        self.iskonto_yuzde_2_e.bind("<FocusOut>", lambda e: self._format_on_focus_out_urun_karti(e, self.sv_iskonto_yuzde_2, 2))
        self.iskonto_yuzde_2_e.bind("<KeyRelease>", self._anlik_hesaplama_ve_guncelleme)

        ttk.Label(main_f, text="%", anchor=tk.W).grid(row=current_row, column=2, padx=(0,5), pady=8, sticky=tk.W)
        current_row += 1

        ttk.Separator(main_f, orient='horizontal').grid(row=current_row, column=0, columnspan=3, sticky='ew', pady=(10,5))
        current_row += 1

        ttk.Label(main_f, text="Toplam İskonto Yüzdesi:", font=("Segoe UI", 9, "bold")).grid(row=current_row, column=0, padx=5, pady=2, sticky=tk.W)
        self.lbl_toplam_iskonto_yuzdesi = ttk.Label(main_f, text="0,00 %", font=("Segoe UI", 9))
        self.lbl_toplam_iskonto_yuzdesi.grid(row=current_row, column=1, columnspan=2, padx=5, pady=2, sticky=tk.EW)
        current_row += 1

        ttk.Label(main_f, text="Uygulanan İskonto Tutarı (KDV Dahil):", font=("Segoe UI", 9, "bold")).grid(row=current_row, column=0, padx=5, pady=2, sticky=tk.W)
        self.lbl_uygulanan_iskonto_dahil = ttk.Label(main_f, text="0,00 TL", font=("Segoe UI", 9))
        self.lbl_uygulanan_iskonto_dahil.grid(row=current_row, column=1, columnspan=2, padx=5, pady=2, sticky=tk.EW)
        current_row += 1

        ttk.Label(main_f, text="İskontolu Birim Fiyat (KDV Dahil):", font=("Segoe UI", 9, "bold")).grid(row=current_row, column=0, padx=5, pady=2, sticky=tk.W)
        self.lbl_iskontolu_bf_dahil = ttk.Label(main_f, text="0,00 TL", font=("Segoe UI", 9))
        self.lbl_iskontolu_bf_dahil.grid(row=current_row, column=1, columnspan=2, padx=5, pady=2, sticky=tk.EW)
        current_row += 1

        ttk.Label(main_f, text="Kalem Toplam (KDV Dahil):", font=("Segoe UI", 10, "bold")).grid(row=current_row, column=0, padx=5, pady=2, sticky=tk.W)
        self.lbl_kalem_toplam_dahil = ttk.Label(main_f, text="0,00 TL", font=("Segoe UI", 10, "bold"))
        self.lbl_kalem_toplam_dahil.grid(row=current_row, column=1, columnspan=2, padx=5, pady=2, sticky=tk.EW)
        current_row += 1

        btn_f = ttk.Frame(main_f)
        btn_f.grid(row=current_row, column=0, columnspan=3, pady=(15,0), sticky=tk.E)

        ttk.Button(btn_f, text="Güncelle", command=self._kalemi_kaydet, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_f, text="İptal", command=self.destroy).pack(side=tk.LEFT)

        self._anlik_hesaplama_ve_guncelleme() # İlk hesaplamayı yap

        self.miktar_e.focus()
        self.miktar_e.selection_range(0, tk.END)


    def _anlik_hesaplama_ve_guncelleme(self, event=None):
        try:
            # Buradaki değişkenlerin doğru StringVar'dan çekildiğinden emin olun
            miktar = self.db.safe_float(self.sv_miktar.get())
            birim_fiyat_kdv_dahil_orijinal = self.db.safe_float(self.sv_fiyat.get())

            # NameError'ı önlemek için burada da yuzde_iskonto_1 ve yuzde_iskonto_2'yi almalıyız.
            yuzde_iskonto_1 = self.db.safe_float(self.sv_iskonto_yuzde_1.get())
            yuzde_iskonto_2 = self.db.safe_float(self.sv_iskonto_yuzde_2.get())

            # Yüzde iskonto doğrulaması (mesaj kutusu göstermeden sadece değeri sıfıra çek)
            if not (0 <= yuzde_iskonto_1 <= 100):
                self.iskonto_yuzde_1_e.delete(0, tk.END)
                self.iskonto_yuzde_1_e.insert(0, "0,00")
                yuzde_iskonto_1 = 0.0

            if not (0 <= yuzde_iskonto_2 <= 100):
                self.iskonto_yuzde_2_e.delete(0, tk.END)
                self.iskonto_yuzde_2_e.insert(0, "0,00")
                yuzde_iskonto_2 = 0.0

            # Ardışık İskonto Hesaplaması:
            fiyat_iskonto_1_sonrasi_dahil = birim_fiyat_kdv_dahil_orijinal * (1 - yuzde_iskonto_1 / 100)
            iskontolu_birim_fiyat_dahil = fiyat_iskonto_1_sonrasi_dahil * (1 - yuzde_iskonto_2 / 100)
            
            if iskontolu_birim_fiyat_dahil < 0:
                iskontolu_birim_fiyat_dahil = 0.0

            toplam_uygulanan_iskonto_dahil = birim_fiyat_kdv_dahil_orijinal - iskontolu_birim_fiyat_dahil
            
            kalem_toplam_dahil = miktar * iskontolu_birim_fiyat_dahil

            if birim_fiyat_kdv_dahil_orijinal > 0:
                toplam_iskonto_yuzdesi = (toplam_uygulanan_iskonto_dahil / birim_fiyat_kdv_dahil_orijinal) * 100
            else:
                toplam_iskonto_yuzdesi = 0.0 

            self.lbl_toplam_iskonto_yuzdesi.config(text=f"{toplam_iskonto_yuzdesi:,.2f} %")
            self.lbl_uygulanan_iskonto_dahil.config(text=self.db._format_currency(toplam_uygulanan_iskonto_dahil))
            self.lbl_iskontolu_bf_dahil.config(text=self.db._format_currency(iskontolu_birim_fiyat_dahil))
            self.lbl_kalem_toplam_dahil.config(text=self.db._format_currency(kalem_toplam_dahil))

        except ValueError:
            self.lbl_toplam_iskonto_yuzdesi.config(text="0,00 %")
            self.lbl_uygulanan_iskonto_dahil.config(text="0,00 TL")
            self.lbl_iskontolu_bf_dahil.config(text="0,00 TL")
            self.lbl_kalem_toplam_dahil.config(text="0,00 TL")
        except Exception as e:
            print(f"Anlık hesaplama hatası: {e}\n{traceback.format_exc()}")
            messagebox.showerror("Hata", f"Hesaplama sırasında beklenmeyen bir hata oluştu: {e}", parent=self)

    def _kalemi_kaydet(self):
        """
        Kalem düzenleme penceresindeki 'Güncelle' butonuna basıldığında tetiklenir.
        Girişleri doğrular, stok kontrolü yapar ve ana sayfadaki kalemi günceller.
        """
        # Tüm değişkenleri fonksiyonun başında başlatarak NameError riskini sıfırla
        yeni_miktar = 0.0
        yeni_fiyat_kdv_dahil_orijinal = 0.0
        yuzde_iskonto_1 = 0.0
        yuzde_iskonto_2 = 0.0
        yeni_alis_fiyati_aninda = self.mevcut_alis_fiyati_fatura_aninda # Varsayılan olarak mevcut değeri al

        try:
            # Kullanıcı girişlerini al ve güvenli bir şekilde float'a dönüştür.
            yeni_miktar = self.db.safe_float(self.sv_miktar.get())
            yeni_fiyat_kdv_dahil_orijinal = self.db.safe_float(self.sv_fiyat.get())
            yuzde_iskonto_1 = self.db.safe_float(self.sv_iskonto_yuzde_1.get())
            yuzde_iskonto_2 = self.db.safe_float(self.sv_iskonto_yuzde_2.get())
            
            # Fatura Anı Alış Fiyatı sadece belirli tiplerde aktifse alınır.
            if (self.islem_tipi == 'SATIŞ' or self.islem_tipi == 'SATIŞ_SIPARIS') and self.alis_fiyati_aninda_e:
                yeni_alis_fiyati_aninda = self.db.safe_float(self.sv_alis_fiyati_aninda.get())

            # --- Giriş Doğrulamaları ---
            if yeni_miktar <= 0:
                messagebox.showerror("Geçersiz Miktar", "Miktar pozitif bir sayı olmalıdır.", parent=self)
                return
            if yeni_fiyat_kdv_dahil_orijinal < 0:
                messagebox.showerror("Geçersiz Fiyat", "Birim fiyat negatif olamaz.", parent=self)
                return
            if not (0 <= yuzde_iskonto_1 <= 100):
                messagebox.showerror("Geçersiz İskonto 1 Yüzdesi", "İskonto 1 yüzdesi 0 ile 100 arasında olmalıdır.", parent=self)
                return
            if not (0 <= yuzde_iskonto_2 <= 100):
                messagebox.showerror("Geçersiz İskonto 2 Yüzdesi", "İskonto 2 yüzdesi 0 ile 100 arasında olmalıdır.", parent=self)
                return
            if (self.islem_tipi == 'SATIŞ' or self.islem_tipi == 'SATIŞ_SIPARIS') and self.alis_fiyati_aninda_e and yeni_alis_fiyati_aninda < 0:
                messagebox.showerror("Geçersiz Fiyat", "Fatura anı alış fiyatı negatif olamaz.", parent=self)
                return

            # --- Stok Kontrolü (SATIŞ Faturaları/Siparişleri için) ---
            if self.islem_tipi == 'SATIŞ' or self.islem_tipi == 'SATIŞ_SIPARIS':
                mevcut_stok_kontrol_icin = self.db.get_stok_miktari_for_kontrol(
                    self.urun_id, 
                    fatura_id_hariç=self.fatura_id_duzenle # Fatura ID'si veya Sipariş ID'si
                )

                # Sepetteki aynı üründen olan diğer miktarlar
                sepetteki_diger_miktar = sum(k[2] for i,k in enumerate(self.parent_page.fatura_kalemleri_ui) 
                                            if i != self.kalem_index and k[0] == self.urun_id)
                
                net_stok_bu_kalem_haric_ve_mevcut_fatura_etkisi_iptal = mevcut_stok_kontrol_icin - sepetteki_diger_miktar

                if yeni_miktar > net_stok_bu_kalem_haric_ve_mevcut_fatura_etkisi_iptal:
                    onay = messagebox.askyesno(
                        "Stok Uyarısı", 
                        f"'{self.urun_adi}' için stok yetersiz!\n"
                        f"Mevcut (anlık) Stok: {net_stok_bu_kalem_haric_ve_mevcut_fatura_etkisi_iptal:.2f} adet\n"
                        f"Talep Edilen Miktar: {yeni_miktar:.2f} adet\n"
                        f"Bu işlem negatif stok yaratacaktır. Devam etmek istiyor musunuz?",
                        icon='warning',
                        parent=self # Messagebox'ın parent'ı KalemDuzenlePenceresi olsun
                    )
                    if not onay:
                        return # Kullanıcı devam etmek istemezse işlemi iptal et.

            # kalem_guncelle metodunu çağırarak ana listeyi güncelle.
            self.parent_page.kalem_guncelle(
                self.kalem_index, 
                yeni_miktar, 
                yeni_fiyat_kdv_dahil_orijinal, 
                yeni_alis_fiyati_aninda,
                yuzde_iskonto_1, 
                yuzde_iskonto_2  
            )
            self.destroy() # Kalem düzenleme penceresini kapat.

        except ValueError as ve:
            # Hata ayıklama için daha spesifik mesajlar
            messagebox.showerror("Giriş Hatası", f"Sayısal alanlarda geçersiz değerler var: {ve}", parent=self)
            print(f"Kalem Guncelle ValueError: {ve}\n{traceback.format_exc()}")
        except IndexError as ie:
            messagebox.showerror("Hata", f"Güncellenecek kalem bulunamadı (indeks hatası): {ie}", parent=self)
            print(f"Kalem Guncelle IndexError: {ie}\n{traceback.format_exc()}")
        except Exception as e:
            # Beklenmeyen diğer hatalar için genel hata yakalama
            messagebox.showerror("Hata", f"Kalem güncellenirken beklenmeyen bir hata oluştu: {e}\n{traceback.format_exc()}", parent=self)
            print(f"Kalem Guncelle Genel Hata: {e}\n{traceback.format_exc()}")

class FiyatGecmisiPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, cari_id, urun_id, fatura_tipi, update_callback, current_kalem_index):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.cari_id = cari_id
        self.urun_id = urun_id
        self.fatura_tipi = fatura_tipi
        self.update_callback = update_callback # FaturaOlusturmaSayfasi'ndaki kalemi güncelleme callback'i
        self.current_kalem_index = current_kalem_index # Sepetteki güncel kalemin indeksi

        self.title("Fiyat Geçmişi Seç")
        self.geometry("600x400") # Boyut ayarı
        self.transient(parent_app) # Ana pencerenin üzerinde kalır
        self.grab_set() # Diğer pencerelere tıklamayı engeller
        self.resizable(False, False) # Boyutlandırılamaz

        ttk.Label(self, text="Geçmiş Fiyat Listesi", font=("Segoe UI", 14, "bold")).pack(pady=10)

        # Fiyat Geçmişi Listesi (Treeview)
        tree_frame = ttk.Frame(self, padding="10")
        tree_frame.pack(expand=True, fill=tk.BOTH)

        # Sütunlar: Fatura No, Tarih, Fiyat (KDV Dahil), İskonto 1 (%), İskonto 2 (%)
        cols = ("Fatura No", "Tarih", "Fiyat (KDV Dahil)", "İskonto 1 (%)", "İskonto 2 (%)")
        self.price_history_tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode="browse")

        col_defs = [
            ("Fatura No", 120, tk.W, tk.NO),
            ("Tarih", 90, tk.CENTER, tk.NO),
            ("Fiyat (KDV Dahil)", 120, tk.E, tk.NO),
            ("İskonto 1 (%)", 90, tk.E, tk.NO),
            ("İskonto 2 (%)", 90, tk.E, tk.NO)
        ]

        for cn, w, a, s in col_defs:
            self.price_history_tree.column(cn, width=w, anchor=a, stretch=s)
            self.price_history_tree.heading(cn, text=cn)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.price_history_tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.price_history_tree.configure(yscrollcommand=vsb.set)
        self.price_history_tree.pack(expand=True, fill=tk.BOTH)

        # Çift tıklama veya seçip butona basma ile fiyatı seçme
        self.price_history_tree.bind("<Double-1>", self._on_price_selected_double_click)

        self._load_price_history() # Geçmiş fiyatları yükle

        # Alt Butonlar
        button_frame = ttk.Frame(self, padding="10")
        button_frame.pack(fill=tk.X)
        ttk.Button(button_frame, text="Seç ve Uygula", command=self._on_price_selected_button, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Kapat", command=self.destroy).pack(side=tk.RIGHT)

    def _load_price_history(self):
        """Veritabanından geçmiş fiyat bilgilerini çeker ve Treeview'e doldurur."""
        # db.get_gecmis_fatura_kalemi_bilgileri metodunu çağır
        history_data = self.db.get_gecmis_fatura_kalemi_bilgileri(self.cari_id, self.urun_id, self.fatura_tipi)

        if not history_data:
            self.price_history_tree.insert("", tk.END, values=("", "", "Geçmiş Fiyat Yok", "", ""))
            return

        for item in history_data:
            # item: (fatura_id, fatura_no, formatted_date, nihai_iskontolu_kdv_dahil_bf, iskonto_yuzde_1, iskonto_yuzde_2)
            fatura_no = item[1]
            tarih = item[2]
            fiyat = self.db._format_currency(item[3])
            iskonto_1 = f"{item[4]:.2f}".replace('.', ',').rstrip('0').rstrip(',')
            iskonto_2 = f"{item[5]:.2f}".replace('.', ',').rstrip('0').rstrip(',')

            self.price_history_tree.insert("", tk.END, values=(
                fatura_no, tarih, fiyat, iskonto_1, iskonto_2
            ), iid=f"history_item_{item[0]}") # iid olarak fatura ID'si kullanılabilir

    def _on_price_selected_double_click(self, event):
        self._on_price_selected_button()

    def _on_price_selected_button(self):
        """Seçilen fiyatı alır ve FaturaOlusturmaSayfasi'na geri gönderir."""
        selected_item_iid = self.price_history_tree.focus()
        if not selected_item_iid:
            messagebox.showwarning("Uyarı", "Lütfen uygulamak için bir geçmiş fiyat seçin.", parent=self)
            return

        item_values = self.price_history_tree.item(selected_item_iid, 'values')
        
        # item_values formatı: ("Fatura No", "Tarih", "Fiyat (KDV Dahil)", "İskonto 1 (%)", "İskonto 2 (%)")
        # Fiyatı, İskonto 1 ve İskonto 2'yi al
        selected_price_str = item_values[2] # Örn: "1.620,00 TL"
        selected_iskonto1_str = item_values[3] # Örn: "10,00" veya "0"
        selected_iskonto2_str = item_values[4] # Örn: "0"

        try:
            cleaned_price_str = selected_price_str.replace(' TL', '').replace('₺', '').strip()
            cleaned_iskonto1_str = selected_iskonto1_str.replace('%', '').strip()
            cleaned_iskonto2_str = selected_iskonto2_str.replace('%', '').strip()

            selected_price = self.db.safe_float(cleaned_price_str)
            selected_iskonto1 = self.db.safe_float(cleaned_iskonto1_str)
            selected_iskonto2 = self.db.safe_float(cleaned_iskonto2_str)

            print(f"DEBUG: Secilen Fiyat (temizlenmis): '{cleaned_price_str}' -> {selected_price}")
            print(f"DEBUG: Secilen Iskonto 1 (temizlenmis): '{cleaned_iskonto1_str}' -> {selected_iskonto1}")
            print(f"DEBUG: Secilen Iskonto 2 (temizlenmis): '{cleaned_iskonto2_str}' -> {selected_iskonto2}")

        except ValueError:
            # safe_float'ın içinde zaten ValueError yakalanıyor ama burada da bir kontrol iyi olur.
            messagebox.showerror("Hata", "Seçilen fiyat verisi geçersiz. (Dönüştürme hatası)", parent=self)
            return
        except Exception as e:
            messagebox.showerror("Hata", f"Fiyat geçmişi verisi işlenirken beklenmeyen bir hata oluştu: {e}", parent=self)
            return

        # update_callback metodu, (kalem_index, yeni_birim_fiyat_kdv_dahil, yeni_iskonto_1, yeni_iskonto_2) alacak.
        self.update_callback(self.current_kalem_index, selected_price, selected_iskonto1, selected_iskonto2)
        self.destroy() # Pencereyi kapat


class KullaniciYonetimiPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app # Ana App referansı
        self.title("Kullanıcı Yönetimi")
        self.geometry("600x650")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Kullanıcı Listesi ve Yönetimi", font=("Segoe UI", 16, "bold")).pack(pady=10)

        # Kullanıcı Listesi
        list_frame = ttk.Frame(self, padding="10")
        list_frame.pack(expand=True, fill=tk.BOTH, pady=5)
        
        cols_kul = ("ID", "Kullanıcı Adı", "Yetki")
        self.tree_kul = ttk.Treeview(list_frame, columns=cols_kul, show='headings', selectmode="browse")
        
        for col_name in cols_kul:
            self.tree_kul.heading(col_name, text=col_name, command=lambda _col=col_name: sort_treeview_column(self.tree_kul, _col, False))
        
        self.tree_kul.column("ID", width=50, stretch=tk.NO, anchor=tk.E)
        self.tree_kul.column("Kullanıcı Adı", width=200)
        self.tree_kul.column("Yetki", width=100, anchor=tk.CENTER)
        self.tree_kul.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
        vsb_kul = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree_kul.yview)
        vsb_kul.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_kul.configure(yscrollcommand=vsb_kul.set)
        self.kullanici_listesini_yenile() # İlk yüklemede listeyi doldur

        # Yeni Kullanıcı Ekleme Formu
        form_frame = ttk.LabelFrame(self, text="Yeni Kullanıcı Ekle / Güncelle", padding="10")
        form_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(form_frame, text="Kullanıcı Adı:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.k_adi_yeni_e = ttk.Entry(form_frame, width=25)
        self.k_adi_yeni_e.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Label(form_frame, text="Yeni Şifre:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.sifre_yeni_e = ttk.Entry(form_frame, show="*", width=25)
        self.sifre_yeni_e.grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Label(form_frame, text="Yetki:").grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
        self.yetki_yeni_cb = ttk.Combobox(form_frame, values=["kullanici", "admin"], state="readonly", width=10)
        self.yetki_yeni_cb.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        self.yetki_yeni_cb.set("kullanici") # Varsayılan
        form_frame.columnconfigure(1, weight=1) # Entry'lerin genişlemesi için

        # Butonlar
        button_frame_kul = ttk.Frame(self, padding="5")
        button_frame_kul.pack(fill=tk.X, padx=10, pady=(0,10))
        
        # "Ekle / Güncelle" butonu: command'i burda atayın
        self.ekle_guncelle_btn = ttk.Button(button_frame_kul, text="Ekle / Güncelle", command=self.yeni_kullanici_ekle, style="Accent.TButton")
        self.ekle_guncelle_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame_kul, text="Seçili Kullanıcıyı Sil", command=self.secili_kullanici_sil).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame_kul, text="Kapat", command=self.destroy).pack(side=tk.RIGHT)

        self.tree_kul.bind("<<TreeviewSelect>>", self.secili_kullaniciyi_forma_yukle) # Seçim değiştiğinde formu doldur

    def kullanici_listesini_yenile(self):
        """Kullanıcı listesini Treeview'de günceller."""
        for i in self.tree_kul.get_children(): self.tree_kul.delete(i)
        kullanicilar = self.db.kullanici_listele()
        
        # <<< DÜZELTME BURADA: Gelen her bir kullanıcı verisini doğru sütunlara ayırıyoruz >>>
        for kul in kullanicilar:
            # kul objesi bir sqlite3.Row objesidir, değerlere anahtar veya indeks ile erişebiliriz.
            self.tree_kul.insert("", "end", values=(kul['id'], kul['kullanici_adi'], kul['yetki']), iid=kul['id'])
            
        self.app.set_status(f"{len(kullanicilar)} kullanıcı listelendi.")
    
    def secili_kullaniciyi_forma_yukle(self, event=None):
        """Treeview'de seçili kullanıcının bilgilerini form alanlarına yükler."""
        selected_item_iid = self.tree_kul.focus()
        if selected_item_iid:
            item_values = self.tree_kul.item(selected_item_iid, "values")
            self.k_adi_yeni_e.delete(0, tk.END)
            self.k_adi_yeni_e.insert(0, item_values[1]) # Kullanıcı adı
            self.yetki_yeni_cb.set(item_values[2]) # Yetki
            self.sifre_yeni_e.delete(0, tk.END) # Şifre alanı temizlensin
            self.ekle_guncelle_btn.config(text="Güncelle") # Buton metnini değiştir
        else: # Seçim yoksa formu temizle
            self.k_adi_yeni_e.delete(0, tk.END)
            self.sifre_yeni_e.delete(0, tk.END)
            self.yetki_yeni_cb.set("kullanici")
            self.ekle_guncelle_btn.config(text="Ekle / Güncelle") # Buton metnini varsayılana döndür

    def yeni_kullanici_ekle(self):
        """Yeni kullanıcı ekler veya seçili kullanıcıyı günceller."""
        k_adi = self.k_adi_yeni_e.get().strip()
        sifre = self.sifre_yeni_e.get().strip() # Yeni şifre (boş olabilir)
        yetki = self.yetki_yeni_cb.get()

        if not (k_adi and yetki):
            messagebox.showerror("Eksik Bilgi", "Kullanıcı adı ve yetki boş bırakılamaz.", parent=self)
            return

        selected_item_iid = self.tree_kul.focus()
        
        # --- MEVCUT KULLANICIYI GÜNCELLEME KISMI ---
        if selected_item_iid: # Treeview'de bir kullanıcı seçiliyse, güncelleme yapıyoruz
            user_id = selected_item_iid
            mevcut_k_adi = self.tree_kul.item(selected_item_iid, "values")[1] # Mevcut kullanıcı adını al

            # Kullanıcı adı değişmişse, kullanıcı adını güncellemeye çalış
            if k_adi != mevcut_k_adi:
                # db.kullanici_adi_guncelle artık (success, message) dönecek
                success_name_update, message_name_update = self.db.kullanici_adi_guncelle(user_id, k_adi)
                if not success_name_update: # Kullanıcı adı güncelleme başarısız olursa
                    messagebox.showerror("Hata", message_name_update, parent=self) # db'den gelen hata mesajını göster
                    return # İşlemi durdur

            # Şifre veya yetki değişmişse veya kullanıcı adı güncellendiyse (yani bir değişiklik olduysa)
            # Şifre alanı boşsa, mevcut şifrenin hash'ini tekrar almalıyız ki şifre değişmesin.
            sifre_to_hash = None
            if sifre: # Eğer yeni bir şifre girilmişse, onu hash'le
                sifre_to_hash = self.db._hash_sifre(sifre)
            else: # Eğer şifre alanı boş bırakılmışsa, mevcut hash'lenmiş şifreyi veritabanından çek.
                try:
                    self.db.c.execute("SELECT sifre FROM kullanicilar WHERE id=?", (user_id,))
                    sifre_to_hash = self.db.c.fetchone()[0] # Mevcut hash'lenmiş şifreyi al
                except Exception as e:
                    messagebox.showerror("Hata", f"Mevcut şifre alınırken bir hata oluştu: {e}", parent=self)
                    return

            # Şifre ve yetki güncelleme işlemini çağır
            # db.kullanici_guncelle_sifre_yetki artık (success, message) dönecek
            success_pw_yetki_update, message_pw_yetki_update = self.db.kullanici_guncelle_sifre_yetki(user_id, sifre_to_hash, yetki)
            
            if success_pw_yetki_update:
                messagebox.showinfo("Başarılı", message_pw_yetki_update, parent=self) # db'den gelen başarılı mesajı göster
                self.app.set_status(message_pw_yetki_update) # Durum çubuğunu güncelle
            else:
                messagebox.showerror("Hata", message_pw_yetki_update, parent=self) # db'den gelen hata mesajını göster
            
            # Güncelleme sonrası ortak temizlik ve yenileme
            self.kullanici_listesini_yenile()
            self.k_adi_yeni_e.delete(0, tk.END)
            self.sifre_yeni_e.delete(0, tk.END)
            self.tree_kul.selection_remove(self.tree_kul.selection()) # Seçimi kaldır
            self.secili_kullaniciyi_forma_yukle() # Formu temizle (butonu da "Ekle / Güncelle" yapar)


        # --- YENİ KULLANICI EKLEME KISMI ---
        else: # Treeview'de bir kullanıcı seçili değilse, yeni kullanıcı ekliyoruz
            if not sifre: # Yeni kullanıcı eklerken şifre boş bırakılamaz
                messagebox.showerror("Eksik Bilgi", "Yeni kullanıcı eklerken şifre boş bırakılamaz.", parent=self)
                return

            # db.kullanici_ekle artık (success, message) dönecek
            success_add, message_add = self.db.kullanici_ekle(k_adi, sifre, yetki)
            
            if success_add:
                messagebox.showinfo("Başarılı", message_add, parent=self) # db'den gelen başarılı mesajı göster
                self.app.set_status(message_add) # Durum çubuğunu güncelle
            else:
                messagebox.showerror("Hata", message_add, parent=self) # db'den gelen hata mesajını göster

            # Ekleme sonrası ortak temizlik ve yenileme
            self.kullanici_listesini_yenile()
            self.k_adi_yeni_e.delete(0, tk.END)
            self.sifre_yeni_e.delete(0, tk.END)
            self.tree_kul.selection_remove(self.tree_kul.selection()) # Seçimi kaldır
            self.secili_kullaniciyi_forma_yukle() # Formu temizle (butonu da "Ekle / Güncelle" yapar)

    def secili_kullanici_sil(self):
        """Seçili kullanıcıyı siler."""
        selected_item_iid = self.tree_kul.focus()
        if not selected_item_iid:
            messagebox.showwarning("Seçim Yok", "Lütfen silmek istediğiniz kullanıcıyı seçin.", parent=self)
            return
        
        k_adi_secili = self.tree_kul.item(selected_item_iid, "values")[1]
        # Kendi kendini silme engeli
        if k_adi_secili == self.app.current_user[1]: 
             messagebox.showwarning("Engellendi", "Aktif olarak giriş yapmış olduğunuz kendi kullanıcı hesabınızı silemezsiniz.", parent=self)
             return

        if messagebox.askyesno("Onay", f"'{k_adi_secili}' kullanıcısını silmek istediğinizden emin misiniz?", parent=self):
            # db.kullanici_sil artık (success, message) dönecek
            success, message = self.db.kullanici_sil(selected_item_iid)
            if success:
                messagebox.showinfo("Başarılı", message, parent=self) # db'den gelen başarılı mesajı göster
                self.kullanici_listesini_yenile()
                self.app.set_status(message) # Durum çubuğunu güncelle
            else:
                messagebox.showerror("Hata", message, parent=self)

class YeniGelirGiderEklePenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, yenile_callback, initial_tip=None):
        super().__init__(parent_app)
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.parent_app = parent_app

        self.kasa_banka_map = {}
        # DÜZELTME BAŞLANGICI: Yeni sınıflandırma haritaları
        self.gelir_siniflandirma_map = {}
        self.gider_siniflandirma_map = {}
        # DÜZELTME BİTİŞİ

        self.title("Yeni Manuel Gelir/Gider Kaydı")
        self.resizable(False, False)
        self.transient(parent_app)
        self.grab_set()

        entry_frame = ttk.Frame(self, padding="15")
        entry_frame.pack(expand=True, fill=tk.BOTH, side=tk.TOP)

        current_row = 0 # UI elemanları için satır indeksi

        ttk.Label(entry_frame, text="Tarih (YYYY-AA-GG):").grid(row=current_row,column=0,padx=5,pady=8,sticky=tk.W)
        self.tarih_entry = ttk.Entry(entry_frame, width=25)
        self.tarih_entry.grid(row=current_row,column=1,padx=5,pady=8,sticky=tk.EW)
        self.tarih_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        setup_date_entry(self.parent_app, self.tarih_entry) 
        ttk.Button(entry_frame, text="🗓️", command=lambda: DatePickerDialog(self.parent_app, self.tarih_entry), width=3).grid(row=current_row, column=2, padx=2, pady=8, sticky=tk.W)
        current_row += 1

        ttk.Label(entry_frame, text="İşlem Tipi:").grid(row=current_row,column=0,padx=5,pady=8,sticky=tk.W)
        self.tip_combo = ttk.Combobox(entry_frame, width=25, values=["GELİR", "GİDER"], state="readonly")
        self.tip_combo.grid(row=current_row,column=1,padx=5,pady=8,sticky=tk.EW)
        
        # initial_tip parametresine göre varsayılanı ayarla
        if initial_tip and initial_tip in ["GELİR", "GİDER"]:
            self.tip_combo.set(initial_tip)
        else:
            self.tip_combo.current(0)
        
        # DÜZELTME BAŞLANGICI: Tip değişiminde sınıflandırma combobox'larını ayarla
        self.tip_combo.bind("<<ComboboxSelected>>", self._on_tip_changed)
        # DÜZELTME BİTİŞİ
        current_row += 1

        # DÜZELTME BAŞLANGICI: Sınıflandırma Combobox'ları ve Etiketleri
        ttk.Label(entry_frame, text="Sınıflandırma:").grid(row=current_row, column=0, padx=5, pady=8, sticky=tk.W)
        self.siniflandirma_combo = ttk.Combobox(entry_frame, width=25, state="readonly")
        self.siniflandirma_combo.grid(row=current_row, column=1, padx=5, pady=8, sticky=tk.EW)
        current_row += 1
        # DÜZELTME BİTİŞİ

        ttk.Label(entry_frame, text="Tutar (TL):").grid(row=current_row,column=0,padx=5,pady=8,sticky=tk.W)
        self.tutar_entry = ttk.Entry(entry_frame, width=25)
        self.tutar_entry.grid(row=current_row,column=1,padx=5,pady=8,sticky=tk.EW)
        setup_numeric_entry(self.parent_app, self.tutar_entry, allow_negative=False, decimal_places=2)
        current_row += 1

        ttk.Label(entry_frame, text="İşlem Kasa/Banka (*):").grid(row=current_row, column=0, sticky=tk.W, padx=5, pady=5)
        self.kasa_banka_combobox = ttk.Combobox(entry_frame, width=25, state="readonly")
        self.kasa_banka_combobox.grid(row=current_row, column=1, padx=5, pady=5, sticky=tk.EW)
        current_row += 1
        
        ttk.Label(entry_frame, text="Açıklama:").grid(row=current_row,column=0,padx=5,pady=8,sticky=tk.W)
        self.aciklama_entry = ttk.Entry(entry_frame, width=25)
        self.aciklama_entry.grid(row=current_row,column=1,padx=5,pady=8,sticky=tk.EW)
        current_row += 1
        
        entry_frame.columnconfigure(1, weight=1)

        ttk.Separator(self, orient='horizontal').pack(fill='x', pady=5, side=tk.TOP)
        button_frame = ttk.Frame(self, padding=(0,5,0,15))
        button_frame.pack(fill=tk.X, side=tk.TOP)
        center_buttons_frame = ttk.Frame(button_frame)
        center_buttons_frame.pack()
        ttk.Button(center_buttons_frame,text="Kaydet",command=self._kaydet,style="Accent.TButton").pack(side=tk.LEFT,padx=10)
        ttk.Button(center_buttons_frame,text="İptal",command=self.destroy).pack(side=tk.LEFT,padx=10)

        self.protocol("WM_DELETE_WINDOW", self.destroy)
        
        # DÜZELTME BAŞLANGICI: İlk yüklemede sınıflandırmaları ve kasa/bankaları yükle
        self._yukle_kasa_banka_hesaplarini()
        self._yukle_siniflandirmalar_comboboxlari_ve_ayarla() # Yeni çağrı
        # DÜZELTME BİTİŞİ

        self.tarih_entry.focus()
        self.update_idletasks()
        self.geometry(f"{self.winfo_reqwidth()}x{self.winfo_reqheight()}")

    # DÜZELTME BAŞLANGICI: _yukle_siniflandirmalar_comboboxlari_ve_ayarla metodu
    def _yukle_siniflandirmalar_comboboxlari_ve_ayarla(self):
        """
        Kasa/Banka hesaplarını ve Gelir/Gider sınıflandırmalarını yükler.
        Sınıflandırma combobox'larını seçili işlem tipine göre ayarlar.
        """
        # Kasa/Banka yüklemesi (mevcut metodunuz)
        self._yukle_kasa_banka_hesaplarini() 

        # Gelir Sınıflandırmalarını yükle
        self.gelir_siniflandirma_map = self.db.get_gelir_siniflandirmalari_for_combobox()
        # Gider Sınıflandırmalarını yükle
        self.gider_siniflandirma_map = self.db.get_gider_siniflandirmalari_for_combobox()

        # İlk ayarlamayı yap
        self._on_tip_changed()

    def _on_tip_changed(self, event=None):
        """İşlem tipi değiştiğinde sınıflandırma combobox'ını günceller."""
        selected_tip = self.tip_combo.get()
        display_values = ["Seçim Yok"]
        selected_map = {}

        if selected_tip == "GELİR":
            selected_map = self.gelir_siniflandirma_map
        elif selected_tip == "GİDER":
            selected_map = self.gider_siniflandirma_map

        display_values.extend(sorted(selected_map.keys()))
        self.siniflandirma_combo['values'] = display_values
        self.siniflandirma_combo.set("Seçim Yok") # Varsayılan olarak "Seçim Yok" seçili olsun
        self.siniflandirma_combo.config(state="readonly")
    # DÜZELTME BİTİŞI


    def _yukle_kasa_banka_hesaplarini(self):
        self.kasa_banka_combobox['values'] = []
        self.kasa_banka_map.clear() # Harita temizlenir
        hesaplar = self.db.kasa_banka_listesi_al()
        display_values = [""]

        if hesaplar:
            for h_id, h_ad, h_no, h_bakiye, h_para_birimi, h_tip, h_acilis_tarihi, h_banka, h_sube_adi, h_varsayilan_odeme_turu in hesaplar:
                bakiye_formatted = self.db._format_currency(h_bakiye)
                display_text = f"{h_ad} ({h_tip}) - Bakiye: {bakiye_formatted}"
                if h_tip == "BANKA" and h_banka:
                    display_text += f" ({h_banka})"
                if h_tip == "BANKA" and h_no:
                    display_text += f" ({h_no})"
                self.kasa_banka_map[display_text] = h_id 
                display_values.append(display_text)

            self.kasa_banka_combobox['values'] = display_values
            self.kasa_banka_combobox.config(state="readonly")
            
            default_hesap_text = None
            for text in display_values:
                # "MERKEZİ NAKİT" ile başlayan metni bul
                if text.strip().startswith("MERKEZİ NAKİT"):
                    default_hesap_text = text
                    break

            if default_hesap_text:
                # Eğer bulunduysa, onu varsayılan olarak ayarla
                self.kasa_banka_combobox.set(default_hesap_text)
            elif len(display_values) > 1:
                # Eğer bulunamadıysa ama listede başka hesap varsa, ilk hesabı seç
                self.kasa_banka_combobox.current(1)
            else:
                # Hiç hesap yoksa boş bırak
                self.kasa_banka_combobox.set("")
        else:
            self.kasa_banka_combobox['values'] = ["Hesap Yok"]
            self.kasa_banka_combobox.set("Hesap Yok")
            self.kasa_banka_combobox.config(state=tk.DISABLED)

    def _kaydet(self):
        tarih_str = self.tarih_entry.get().strip()
        tip_str = self.tip_combo.get()
        tutar_str_val = self.tutar_entry.get().strip()
        aciklama_str = self.aciklama_entry.get().strip()

        secili_hesap_display = self.kasa_banka_combobox.get()
        kasa_banka_id_val = None
        if secili_hesap_display and secili_hesap_display != "Hesap Yok":
            kasa_banka_id_val = self.kasa_banka_map.get(secili_hesap_display) 

        secili_siniflandirma_display = self.siniflandirma_combo.get()
        gelir_siniflandirma_id_val = None
        gider_siniflandirma_id_val = None

        if secili_siniflandirma_display and secili_siniflandirma_display != "Seçim Yok":
            if tip_str == "GELİR":
                gelir_siniflandirma_id_val = self.gelir_siniflandirma_map.get(secili_siniflandirma_display)
            elif tip_str == "GİDER":
                gider_siniflandirma_id_val = self.gider_siniflandirma_map.get(secili_siniflandirma_display)
        else:
            messagebox.showwarning("Uyarı", "Lütfen bir sınıflandırma seçin.", parent=self)
            return

        if kasa_banka_id_val is None:
            messagebox.showerror("Eksik Bilgi", "Lütfen bir İşlem Kasa/Banka hesabı seçin.", parent=self)
            return

        # DÜZELTME BAŞLANGICI: tutar_str yerine tutar_str_val kullanıldı
        if not all([tarih_str, tutar_str_val, aciklama_str]):
            messagebox.showerror("Eksik Bilgi", "Lütfen tüm zorunlu (*) alanları doldurun.", parent=self.parent_app)
            return
        # DÜZELTME BİTİŞİ

        try:
            tutar_f = float(tutar_str_val.replace(',', '.'))
            if tutar_f <= 0:
                messagebox.showerror("Geçersiz Tutar", "Tutar pozitif bir sayı olmalıdır.", parent=self.parent_app)
                return
        except ValueError:
            messagebox.showerror("Giriş Hatası", "Tutar sayısal bir değer olmalıdır.", parent=self.parent_app)
            return

        success, message = self.db.gelir_gider_ekle(
            tarih_str, tip_str, tutar_f, aciklama_str, kasa_banka_id_val,
            gelir_siniflandirma_id=gelir_siniflandirma_id_val,
            gider_siniflandirma_id=gider_siniflandirma_id_val
        )
        if success:
            messagebox.showinfo("Başarılı", message, parent=self.parent_app)
            if self.yenile_callback:
                self.yenile_callback()
            self.destroy() # <-- DÜZELTME: Başarılı kaydetme sonrası pencereyi kapat
        else:
            messagebox.showerror("Hata", message, parent=self.parent_app)

class TarihAraligiDialog(simpledialog.Dialog):
    def __init__(self, parent, title=None, baslangic_gun_sayisi=30):
        self.bas_tarih_str = (datetime.now() - timedelta(days=baslangic_gun_sayisi)).strftime('%Y-%m-%d')
        self.bit_tarih_str = datetime.now().strftime('%Y-%m-%d')
        self.sonuc = None # Kullanıcının seçtiği tarih aralığını tutacak
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Başlangıç Tarihi (YYYY-AA-GG):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.bas_tarih_entry_dialog = ttk.Entry(master, width=15)
        self.bas_tarih_entry_dialog.grid(row=0, column=1, padx=5, pady=2)
        self.bas_tarih_entry_dialog.insert(0, self.bas_tarih_str)

        ttk.Label(master, text="Bitiş Tarihi (YYYY-AA-GG):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.bit_tarih_entry_dialog = ttk.Entry(master, width=15)
        self.bit_tarih_entry_dialog.grid(row=1, column=1, padx=5, pady=2)
        self.bit_tarih_entry_dialog.insert(0, self.bit_tarih_str)
        return self.bas_tarih_entry_dialog # İlk odaklanılacak widget

    def apply(self):
        # Kullanıcı OK'a bastığında çağrılır.
        bas_t_str_dialog = self.bas_tarih_entry_dialog.get()
        bit_t_str_dialog = self.bit_tarih_entry_dialog.get()
        try:
            bas_dt_dialog = datetime.strptime(bas_t_str_dialog, '%Y-%m-%d')
            bit_dt_dialog = datetime.strptime(bit_t_str_dialog, '%Y-%m-%d')
            if bas_dt_dialog > bit_dt_dialog:
                messagebox.showerror("Tarih Hatası", "Başlangıç tarihi, bitiş tarihinden sonra olamaz.", parent=self) # parent=self ile dialog üzerinde göster
                self.sonuc=None # Hata durumunda sonucu None yap
                return # Fonksiyondan çık, dialog kapanmaz
            self.sonuc = (bas_t_str_dialog, bit_t_str_dialog) # Sonucu tuple olarak sakla
        except ValueError:
            messagebox.showerror("Format Hatası", "Tarih formatı YYYY-AA-GG olmalıdır (örn: 2023-12-31).", parent=self)
            self.sonuc=None
            return

class OdemeTuruSecimDialog(tk.Toplevel):
    def __init__(self, parent_app, db_manager, fatura_tipi, initial_cari_id, callback_func):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.fatura_tipi = fatura_tipi # 'SATIŞ' veya 'ALIŞ'
        self.initial_cari_id = initial_cari_id
        self.callback_func = callback_func # Seçim sonrası çağrılacak fonksiyon

        self.title("Ödeme Türü Seçimi")
        self.geometry("400x300")
        self.transient(parent_app)
        self.grab_set()
        self.resizable(False, False)

        self.kasa_banka_map = {} # Kasa/Banka hesaplarını display_text -> ID olarak tutar
        
        ttk.Label(self, text="Fatura Ödeme Türünü Seçin", font=("Segoe UI", 12, "bold")).pack(pady=10)

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # Ödeme Türü Seçimi Combobox
        ttk.Label(main_frame, text="Ödeme Türü (*):").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.odeme_turu_cb = ttk.Combobox(main_frame, width=25, state="readonly")
        # Perakende satışsa 'AÇIK HESAP' ve 'ETKİSİZ FATURA' hariç, değilse 'ETKİSİZ FATURA' hariç
        self._set_odeme_turu_values() # Değerleri burada ayarla
        self.odeme_turu_cb.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        self.odeme_turu_cb.bind("<<ComboboxSelected>>", self._odeme_turu_degisince_hesap_combobox_ayarla)
        self.odeme_turu_cb.current(0) # İlk değeri varsayılan yap

        # İşlem Kasa/Banka Seçimi Combobox
        ttk.Label(main_frame, text="İşlem Kasa/Banka (*):").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.islem_hesap_cb = ttk.Combobox(main_frame, width=25, state=tk.DISABLED)
        self.islem_hesap_cb.grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)

        # Vade Tarihi Alanı (isteğe bağlı, "AÇIK HESAP" için)
        self.lbl_vade_tarihi = ttk.Label(main_frame, text="Vade Tarihi:")
        self.entry_vade_tarihi = ttk.Entry(main_frame, width=15, state=tk.DISABLED) 
        self.btn_vade_tarihi = ttk.Button(main_frame, text="🗓️", command=lambda: DatePickerDialog(self.app, self.entry_vade_tarihi), width=3, state=tk.DISABLED)
        self.lbl_vade_tarihi.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.entry_vade_tarihi.grid(row=2, column=1, padx=5, pady=5, sticky=tk.EW)
        self.btn_vade_tarihi.grid(row=2, column=2, padx=2, pady=5, sticky=tk.W)
        setup_date_entry(self.app, self.entry_vade_tarihi)
        self.lbl_vade_tarihi.grid_remove() # Başlangıçta gizle
        self.entry_vade_tarihi.grid_remove()
        self.btn_vade_tarihi.grid_remove()

        main_frame.columnconfigure(1, weight=1) # Entry/Combobox sütunu genişleyebilir

        button_frame = ttk.Frame(self, padding="10")
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)

        ttk.Button(button_frame, text="Onayla", command=self._onayla, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.RIGHT, padx=5)

        self._yukle_kasa_banka_hesaplarini() # Kasa/Banka hesaplarını yükle
        self._odeme_turu_degisince_hesap_combobox_ayarla() # İlk seçime göre combobox'ı ayarla

    def _set_odeme_turu_values(self):
        """Ödeme türü combobox'ının değerlerini fatura tipine göre ayarlar."""
        all_payment_values = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET", "AÇIK HESAP", "ETKİSİZ FATURA"]
        
        # Perakende müşteri mi kontrol et
        is_perakende_musteri = False
        if self.fatura_tipi == 'SATIŞ' and self.initial_cari_id is not None and \
           str(self.initial_cari_id) == str(self.db.perakende_musteri_id):
            is_perakende_musteri = True

        if is_perakende_musteri:
            # Perakende satışsa 'AÇIK HESAP' ve 'ETKİSİZ FATURA' hariç
            self.odeme_turu_cb['values'] = [p for p in all_payment_values if p != "AÇIK HESAP" and p != "ETKİSİZ FATURA"]
        else:
            # Diğer durumlarda 'ETKİSİZ FATURA' hariç (çünkü faturalara dönüştürülürken bu tür kullanılmaz)
            self.odeme_turu_cb['values'] = [p for p in all_payment_values if p != "ETKİSİZ FATURA"]

    def _yukle_kasa_banka_hesaplarini(self):
        self.islem_hesap_cb['values'] = [""] # İlk seçenek boş olsun
        self.kasa_banka_map.clear()
        hesaplar = self.db.kasa_banka_listesi_al()
        display_values = [""] 

        if hesaplar:
            for h_id, h_ad, h_no, h_bakiye, h_para_birimi, h_tip, h_acilis_tarihi, h_banka, h_sube_adi, h_varsayilan_odeme_turu in hesaplar:
                bakiye_formatted = self.db._format_currency(h_bakiye)
                display_text = f"{h_ad} ({h_tip}) - Bakiye: {bakiye_formatted}"
                if h_tip == "BANKA" and h_banka:
                    display_text += f" ({h_banka})"
                self.kasa_banka_map[display_text] = h_id
                display_values.append(display_text)
    
            self.islem_hesap_cb['values'] = display_values
            self.islem_hesap_cb.config(state="readonly")
            self.islem_hesap_cb.set("") # Başlangıçta boş bırak
        else:
            self.islem_hesap_cb['values'] = ["Hesap Yok"]
            self.islem_hesap_cb.current(0)
            self.islem_hesap_cb.config(state=tk.DISABLED)

    def _odeme_turu_degisince_hesap_combobox_ayarla(self, event=None):
        secili_odeme_turu = self.odeme_turu_cb.get()
        pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"]

        # Vade tarihi alanlarının görünürlüğünü ve aktifliğini ayarla
        if secili_odeme_turu == "AÇIK HESAP":
            self.lbl_vade_tarihi.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W) # 2. satıra taşıdık
            self.entry_vade_tarihi.grid(row=2, column=1, padx=5, pady=5, sticky=tk.EW)
            self.btn_vade_tarihi.grid(row=2, column=2, padx=2, pady=5, sticky=tk.W)
            self.entry_vade_tarihi.config(state=tk.NORMAL)
            self.btn_vade_tarihi.config(state=tk.NORMAL)
            self.entry_vade_tarihi.insert(0, datetime.now().strftime('%Y-%m-%d')) # Varsayılan olarak bugünün tarihini atayalım
        else:
            self.lbl_vade_tarihi.grid_remove()
            self.entry_vade_tarihi.grid_remove()
            self.btn_vade_tarihi.grid_remove()
            self.entry_vade_tarihi.config(state=tk.DISABLED)
            self.entry_vade_tarihi.delete(0, tk.END)

        # Kasa/Banka alanının görünürlüğünü ve aktifliğini ayarla
        if secili_odeme_turu in pesin_odeme_turleri:
            self.islem_hesap_cb.config(state="readonly")
            # Varsayılan kasa/bankayı ayarla
            varsayilan_kb_db = self.db.get_kasa_banka_by_odeme_turu(secili_odeme_turu)
            if varsayilan_kb_db:
                varsayilan_kb_id = varsayilan_kb_db[0]
                found_and_set_default = False
                for text, id_val in self.kasa_banka_map.items():
                    if id_val == varsayilan_kb_id:
                        self.islem_hesap_cb.set(text)
                        found_and_set_default = True
                        break
                if not found_and_set_default and len(self.islem_hesap_cb['values']) > 1:
                    self.islem_hesap_cb.current(1)
            elif len(self.islem_hesap_cb['values']) > 1:
                self.islem_hesap_cb.current(1)
            else:
                self.islem_hesap_cb.set("")
        else: # "AÇIK HESAP" veya "ETKİSİZ FATURA" seçilirse
            self.islem_hesap_cb.set("")
            self.islem_hesap_cb.config(state=tk.DISABLED)

    def _onayla(self):
        """Kullanıcının seçtiği ödeme türü ve kasa/banka bilgilerini ana forma geri gönderir."""
        secili_odeme_turu = self.odeme_turu_cb.get()
        secili_hesap_display = self.islem_hesap_cb.get()
        vade_tarihi_val = self.entry_vade_tarihi.get().strip()

        kasa_banka_id_val = None
        if secili_hesap_display and secili_hesap_display != "Hesap Yok":
            kasa_banka_id_val = self.kasa_banka_map.get(secili_hesap_display)

        # Zorunlu alan kontrolü
        if not secili_odeme_turu:
            messagebox.showerror("Eksik Bilgi", "Lütfen bir Ödeme Türü seçin.", parent=self)
            return

        pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"]
        if secili_odeme_turu in pesin_odeme_turleri and kasa_banka_id_val is None:
            messagebox.showerror("Eksik Bilgi", "Peşin ödeme türleri için bir İşlem Kasa/Banka hesabı seçmelisiniz.", parent=self)
            return
        
        if secili_odeme_turu == "AÇIK HESAP":
            if not vade_tarihi_val:
                messagebox.showerror("Eksik Bilgi", "Açık Hesap ödeme türü için Vade Tarihi boş olamaz.", parent=self)
                return
            try:
                datetime.strptime(vade_tarihi_val, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Tarih Formatı Hatası", "Vade Tarihi formatı (YYYY-AA-GG) olmalıdır.", parent=self)
                return


        # Callback fonksiyonunu çağır
        self.callback_func(secili_odeme_turu, kasa_banka_id_val, vade_tarihi_val)
        self.destroy() # Pencereyi kapat

class TopluVeriEklePenceresi(tk.Toplevel): # <<< Bu sınıf doğru hizada (BeklemePenceresi ve AciklamaDetayPenceresi ile aynı)
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.title("Toplu Veri Ekleme (Excel)")
        self.geometry("600x650")
        self.transient(parent_app)
        self.grab_set()
        self.resizable(False, False)

        ttk.Label(self, text="Toplu Veri Ekleme (Excel)", font=("Segoe UI", 16, "bold")).pack(pady=10)

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # Veri Tipi Seçimi
        ttk.Label(main_frame, text="Veri Tipi:").grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
        self.veri_tipi_combo = ttk.Combobox(main_frame, values=["Müşteri", "Tedarikçi", "Stok/Ürün Ekle/Güncelle"], state="readonly", width=30)
        self.veri_tipi_combo.grid(row=0, column=1, padx=5, pady=10, sticky=tk.EW)
        self.veri_tipi_combo.set("Müşteri")
        self.veri_tipi_combo.bind("<<ComboboxSelected>>", self._show_template_info_and_options)

        # Excel Dosyası Seçimi
        ttk.Label(main_frame, text="Excel Dosyası:").grid(row=1, column=0, padx=5, pady=10, sticky=tk.W)
        self.dosya_yolu_entry = ttk.Entry(main_frame, width=40)
        self.dosya_yolu_entry.grid(row=1, column=1, padx=5, pady=10, sticky=tk.EW)
        ttk.Button(main_frame, text="Gözat...", command=self._gozat_excel_dosyasi).grid(row=1, column=2, padx=5, pady=10, sticky=tk.W)

        # Stok/Ürün Güncelleme Seçenekleri Çerçevesi (Başlangıçta gizli)
        self.stok_guncelleme_options_frame = ttk.LabelFrame(main_frame, text="Stok/Ürün Güncelleme Seçenekleri", padding="10")
        self.stok_guncelleme_options_frame.grid(row=2, column=0, columnspan=3, padx=5, pady=10, sticky=tk.EW)
        self.stok_guncelleme_options_frame.grid_remove()

        # Checkbox'ları tanımlama
        self.cb_vars = {}
        self.cb_vars['fiyat_bilgileri'] = tk.BooleanVar(self, value=False)
        ttk.Checkbutton(self.stok_guncelleme_options_frame, text="Fiyat Bilgileri (Alış/Satış/KDV)", variable=self.cb_vars['fiyat_bilgileri']).pack(anchor=tk.W, pady=2)
        self.cb_vars['urun_nitelikleri'] = tk.BooleanVar(self, value=False)
        ttk.Checkbutton(self.stok_guncelleme_options_frame, text="Ürün Nitelikleri (Kategori/Marka/Grup/Birim/Menşe/Detay)", variable=self.cb_vars['urun_nitelikleri']).pack(anchor=tk.W, pady=2)
        self.cb_vars['stok_miktari'] = tk.BooleanVar(self, value=False)
        ttk.Checkbutton(self.stok_guncelleme_options_frame, text="Stok Miktarı (Mevcut/Minimum)", variable=self.cb_vars['stok_miktari']).pack(anchor=tk.W, pady=2)
        
        self.cb_vars['tumu'] = tk.BooleanVar(self, value=False)
        self.cb_tumu = ttk.Checkbutton(self.stok_guncelleme_options_frame, text="Tümü (Yukarıdakilerin hepsi)", variable=self.cb_vars['tumu'], command=self._toggle_all_checkboxes)
        self.cb_tumu.pack(anchor=tk.W, pady=5)
        
        # Şablon Bilgisi (kısa özet için)
        self.sv_template_info = tk.StringVar(self)
        self.template_info_label = ttk.Label(main_frame, textvariable=self.sv_template_info, wraplength=550, justify=tk.LEFT)
        self.template_info_label.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky=tk.W)

        # Yeni buton: Detaylı Şablon Açıklaması
        self.detayli_aciklama_button = ttk.Button(main_frame, text="Detaylı Bilgi / Şablon Açıklaması", command=self._show_detayli_aciklama_penceresi)
        self.detayli_aciklama_button.grid(row=3, column=2, padx=5, pady=(5,0), sticky=tk.SE)
        self.detayli_aciklama_button.grid_remove()

        main_frame.columnconfigure(1, weight=1)

        button_frame = ttk.Frame(main_frame, padding="10")
        button_frame.grid(row=4, column=0, columnspan=3, sticky=tk.EW, padx=0, pady=(10,0))

        ttk.Button(button_frame, text="Verileri Yükle", command=self._verileri_yukle, style="Accent.TButton").pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Örnek Şablon İndir", command=self._excel_sablonu_indir).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.RIGHT, padx=10)
        self.analysis_results = None # Başlangıçta None
        self._show_template_info_and_options()
        self.update_idletasks()

    def _show_template_info_and_options(self, event=None):
        selected_type = self.veri_tipi_combo.get()
        short_info_text = ""

        if selected_type == "Stok/Ürün Ekle/Güncelle":
            self.stok_guncelleme_options_frame.grid()
            self.detayli_aciklama_button.grid()
        else:
            self.stok_guncelleme_options_frame.grid_remove()
            self.detayli_aciklama_button.grid_remove()
            self.cb_vars['tumu'].set(False)
            self._toggle_all_checkboxes(event=None, force_off=True)

        if selected_type == "Müşteri":
            short_info_text = "Müşteri Excel dosyası:\n`Müşteri Kodu`, `Ad Soyad` (ZORUNLU) ve diğer detaylar."
        elif selected_type == "Tedarikçi":
            short_info_text = "Tedarikçi Excel dosyası:\n`Tedarikçi Kodu`, `Ad Soyad` (ZORUNLU) ve diğer detaylar."
        elif selected_type == "Stok/Ürün Ekle/Güncelle":
            short_info_text = "Stok/Ürün Excel dosyası:\n`Ürün Kodu`, `Ürün Adı` (ZORUNLU) ve diğer detaylar.\n" \
                              "Güncellemek istediğiniz alanları yukarıdan seçin. Detaylı şablon bilgisi için butona tıklayın."
        
        self.sv_template_info.set(short_info_text)

    def _excel_sablonu_indir(self):
        """Seçilen veri tipine göre örnek Excel şablonu oluşturur ve kaydettirir."""
        veri_tipi = self.veri_tipi_combo.get()
        if not veri_tipi:
            messagebox.showwarning("Uyarı", "Lütfen şablon indirmek için bir veri tipi seçin.", parent=self)
            return

        file_name_prefix = ""
        headers = []

        if veri_tipi == "Müşteri":
            file_name_prefix = "Musteri_Sablonu"
            headers = [
                "Müşteri Kodu", "Ad Soyad", "Telefon", "Adres",
                "Vergi Dairesi", "Vergi No"
            ]
        elif veri_tipi == "Tedarikçi":
            file_name_prefix = "Tedarikci_Sablonu"
            headers = [
                "Tedarikçi Kodu", "Ad Soyad", "Telefon", "Adres",
                "Vergi Dairesi", "Vergi No"
            ]
        elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
            file_name_prefix = "Stok_Urun_Sablonu"
            headers = [
                "Ürün Kodu", "Ürün Adı", "Miktar", "Alış Fiyatı (KDV Dahil)",
                "Satış Fiyatı (KDV Dahil)", "KDV Oranı (%)", "Minimum Stok Seviyesi",
                "Kategori Adı", "Marka Adı", "Ürün Grubu Adı", "Ürün Birimi Adı",
                "Menşe Ülke Adı", "Ürün Detayı", "Ürün Resmi Yolu"
            ]
        else:
            messagebox.showerror("Hata", "Geçersiz veri tipi seçimi.", parent=self)
            return

        default_file_name = f"{file_name_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            initialfile=default_file_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx")],
            title="Excel Şablonunu Kaydet",
            parent=self
        )

        if file_path:
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Veri Şablonu"
                
                sheet.append(headers)

                # Başlıkları kalın yapma ve sütun genişliklerini ayarlama
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = openpyxl.styles.Font(bold=True)
                    # Basit bir otomatik sütun genişliği ayarlaması (tahmini)
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) + 2, 15)

                workbook.save(file_path)
                messagebox.showinfo("Başarılı", f"'{veri_tipi}' şablonu başarıyla oluşturuldu:\n{file_path}", parent=self)
            except Exception as e:
                messagebox.showerror("Hata", f"Şablon oluşturulurken bir hata oluştu: {e}", parent=self)
        else:
            self.app.set_status("Şablon kaydetme işlemi iptal edildi.")

    def _show_detayli_aciklama_penceresi(self):
        """Detaylı şablon açıklamasını içeren yeni bir pencere açar."""
        selected_type = self.veri_tipi_combo.get()
        title = f"{selected_type} Şablon Açıklaması"
        message = ""

        if selected_type == "Müşteri":
            message = (
                "Müşteri Veri Şablonu Detayları:\n\n"
                "Excel dosyasının ilk satırı başlık (header) olmalıdır. Veriler ikinci satırdan başlamalıdır.\n\n"
                "Sütun Sırası ve Açıklamaları:\n"
                "1.  **Müşteri Kodu (ZORUNLU):** Müşterinin benzersiz kodu. Eğer boş bırakılırsa otomatik atanır.\n"
                "2.  **Ad Soyad (ZORUNLU):** Müşterinin tam adı veya şirket adı.\n"
                "3.  **Telefon (İsteğe Bağlı):** Müşterinin telefon numarası.\n"
                "4.  **Adres (İsteğe Bağlı):** Müşterinin adresi.\n"
                "5.  **Vergi Dairesi (İsteğe Bağlı):** Müşterinin kayıtlı olduğu vergi dairesi.\n"
                "6.  **Vergi No (İsteğe Bağlı):** Müşterinin vergi numarası veya T.C. kimlik numarası."
            )
        elif selected_type == "Tedarikçi":
            message = (
                "Tedarikçi Veri Şablonu Detayları:\n\n"
                "Excel dosyasının ilk satırı başlık (header) olmalıdır. Veriler ikinci satırdan başlamalıdır.\n\n"
                "Sütun Sırası ve Açıklamaları:\n"
                "1.  **Tedarikçi Kodu (ZORUNLU):** Tedarikçinin benzersiz kodu. Eğer boş bırakılırsa otomatik atanır.\n"
                "2.  **Ad Soyad (ZORUNLU):** Tedarikçinin tam adı veya şirket adı.\n"
                "3.  **Telefon (İsteğe Bağlı):** Tedarikçinin telefon numarası.\n"
                "4.  **Adres (İsteğe Bağlı):** Tedarikçinin adresi.\n"
                "5.  **Vergi Dairesi (İsteğe Bağlı):** Tedarikçinin kayıtlı olduğu vergi dairesi.\n"
                "6.  **Vergi No (İsteğe Bağlı):** Tedarikçinin vergi numarası."
            )
        elif selected_type == "Stok/Ürün Ekle/Güncelle":
            message = (
                "Stok/Ürün Veri Şablonu Detayları:\n\n"
                "Excel dosyasının ilk satırı başlık (header) olmalıdır. Veriler ikinci satırdan başlamalıdır.\n"
                "Varolan bir ürünü güncellemek için 'Ürün Kodu' eşleşmelidir. Yeni ürünler için boş bırakılabilir veya benzersiz olmalıdır.\n\n"
                "Sütun Sırası ve Açıklamaları:\n"
                "1.  **Ürün Kodu (ZORUNLU):** Ürünün benzersiz kodu. Yeni ürünlerde boş bırakılırsa otomatik atanır.\n"
                "2.  **Ürün Adı (ZORUNLU):** Ürünün adı.\n"
                "3.  **Miktar (İsteğe Bağlı):** Mevcut stok miktarı. Sayısal.\n"
                "4.  **Alış Fiyatı (KDV Dahil) (İsteğe Bağlı):** Ürünün KDV dahil alış fiyatı. Sayısal.\n"
                "5.  **Satış Fiyatı (KDV Dahil) (İsteğe Bağlı):** Ürünün KDV dahil satış fiyatı. Sayısal.\n"
                "6.  **KDV Oranı (%) (İsteğe Bağlı):** Ürünün KDV oranı (örn. 20). Sayısal.\n"
                "7.  **Minimum Stok Seviyesi (İsteğe Bağlı):** Ürünün minimum stok seviyesi. Sayısal.\n"
                "8.  **Kategori Adı (İsteğe Bağlı):** Ürünün kategorisi. Eğer kategori yoksa oluşturulur.\n"
                "9.  **Marka Adı (İsteğe Bağlı):** Ürünün markası. Eğer marka yoksa oluşturulur.\n"
                "10. **Ürün Grubu Adı (İsteğe Bağlı):** Ürünün grubu. Eğer grup yoksa oluşturulur.\n"
                "11. **Ürün Birimi Adı (İsteğe Bağlı):** Ürünün birimi (örn. Adet, Kg, Metre). Eğer birim yoksa oluşturulur.\n"
                "12. **Menşe Ülke Adı (İsteğe Bağlı):** Ürünün menşe ülkesi. Eğer ülke yoksa oluşturulur.\n"
                "13. **Ürün Detayı (İsteğe Bağlı):** Ürün hakkında ek detaylar. Metin.\n"
                "14. **Ürün Resmi Yolu (İsteğe Bağlı):** Ürün resminin tam dosya yolu. Dosya kopyalanacaktır."
            )
        else:
            message = "Lütfen bir veri tipi seçin."

        AciklamaDetayPenceresi(self, title, message)

    def _gozat_excel_dosyasi(self):
        dosya_yolu = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx;*.xls")],
            parent=self
        )
        if dosya_yolu:
            self.dosya_yolu_entry.delete(0, tk.END)
            self.dosya_yolu_entry.insert(0, dosya_yolu)

    def _toggle_all_checkboxes(self, event=None, force_off=False):
        if force_off:
            is_checked = False
        else:
            is_checked = self.cb_vars['tumu'].get()
        
        for key, var in self.cb_vars.items():
            if key != 'tumu':
                var.set(is_checked)

    def _verileri_yukle(self):
        dosya_yolu = self.dosya_yolu_entry.get().strip()
        veri_tipi = self.veri_tipi_combo.get()

        if not dosya_yolu or not os.path.exists(dosya_yolu):
            messagebox.showerror("Dosya Hatası", "Lütfen geçerli bir Excel dosyası seçin.", parent=self)
            return

        selected_update_fields = []
        if veri_tipi == "Stok/Ürün Ekle/Güncelle":
            if self.cb_vars['tumu'].get():
                selected_update_fields = ['fiyat_bilgileri', 'urun_nitelikleri', 'stok_miktari']
            else:
                for key, var in self.cb_vars.items():
                    if key != 'tumu' and var.get():
                        selected_update_fields.append(key)
        
        # <<< DÜZELTME 1: Bekleme penceresini oluşturup bir değişkene atıyoruz >>>
        bekleme_penceresi_analiz = BeklemePenceresi(self, message="Excel okunuyor ve veriler analiz ediliyor...")
        
        # <<< DÜZELTME 2: Arka plan görevini başlatırken, oluşturduğumuz pencereyi ona argüman olarak veriyoruz >>>
        threading.Thread(target=self._analiz_et_ve_onizle_threaded, args=(
            dosya_yolu, 
            veri_tipi, 
            selected_update_fields, 
            bekleme_penceresi_analiz
        )).start()

    def _analiz_et_ve_onizle_threaded(self, dosya_yolu, veri_tipi, selected_update_fields, bekleme_penceresi):
        """
        Excel dosyasını okur, veritabanı analiz metodunu çağırır ve sonucu UI'da gösterir.
        """
        analysis_success = False
        analysis_message = ""
        analysis_results = {}
        
        try:
            workbook = openpyxl.load_workbook(dosya_yolu, data_only=True)
            sheet = workbook.active
            
            raw_data_from_excel_list = [
                [cell.value for cell in row_obj] 
                for row_obj in sheet.iter_rows(min_row=2) 
                if any(cell.value is not None and str(cell.value).strip() != '' for cell in row_obj)
            ]

            if not raw_data_from_excel_list:
                analysis_message = "Excel dosyasında okunacak geçerli veri bulunamadı."
                analysis_success = False
            else:
                if veri_tipi == "Müşteri":
                    analysis_results = self.db.toplu_musteri_analiz_et(raw_data_from_excel_list)
                elif veri_tipi == "Tedarikçi":
                    analysis_results = self.db.toplu_tedarikci_analiz_et(raw_data_from_excel_list)
                elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                    analysis_results = self.db.toplu_stok_analiz_et(raw_data_from_excel_list, selected_update_fields)
                else:
                    analysis_results = {}
                    analysis_success = False
                    analysis_message = f"Bilinmeyen veri tipi: {veri_tipi}"
                
                if analysis_results:
                    analysis_success = True
                    analysis_message = "Veri analizi tamamlandı."

        except Exception as e:
            analysis_message = f"Analiz sırasında beklenmeyen bir hata oluştu: {e}"
            analysis_success = False
            from arayuz import logging
            logging.error(f"Toplu veri analizi thread'inde hata: {traceback.format_exc()}")
        
        finally:
            # <<< DÜZELTME 3: İşlem ne olursa olsun, bize gönderilen bekleme penceresini kapatıyoruz. >>>
            self.app.after(0, bekleme_penceresi.kapat)
            
            if analysis_success and analysis_results:
                self.app.after(0, self._onizleme_penceresini_ac, veri_tipi, analysis_results)
            else:
                if not analysis_message:
                    analysis_message = "Analiz sırasında bilinmeyen bir hata oluştu."
                self.app.after(0, lambda: messagebox.showerror("Hata", f"Veri analizi başarısız oldu:\n{analysis_message}", parent=self.app))
                self.app.after(0, self.app.set_status, f"Toplu {veri_tipi} analizi başarısız.")

    def _onizleme_penceresini_ac(self, veri_tipi, analysis_results):
        TopluVeriOnizlemePenceresi(
            self.app, 
            self.db, 
            veri_tipi, 
            analysis_results, 
            callback_on_confirm=self._gercek_yazma_islemini_yap_threaded_from_onizleme
        )

    def _gercek_yazma_islemini_yap_threaded_from_onizleme(self, veri_tipi, analysis_results):
        self.bekleme_penceresi_gercek_islem = BeklemePenceresi(
            self.app, 
            message=f"Toplu {veri_tipi} veritabanına yazılıyor, lütfen bekleyiniz..."
        )
        
        threading.Thread(target=lambda: self._yazma_islemi_threaded(
            veri_tipi, 
            analysis_results, 
            self.bekleme_penceresi_gercek_islem
        )).start()

    def _yazma_islemi_threaded(self, veri_tipi, analysis_results):
        final_success = True
        final_message = ""
        
        try:
            # İşlenecek doğru veri listesini alıyoruz
            data_to_process = analysis_results.get('all_processed_data', [])
            
            # <<< DÜZELTME BURADA: Hatalı metot adlarını yenileriyle değiştiriyoruz >>>
            if veri_tipi == "Müşteri":
                # Önceki 'toplu_musteri_ekle' yerine yeni 'toplu_musteri_ekle_guncelle' metodunu çağırıyoruz.
                success, message = self.db.toplu_musteri_ekle_guncelle(data_to_process)
            elif veri_tipi == "Tedarikçi":
                # Tedarikçi için de gelecekte aynı hatayı almamak adına düzeltiyoruz.
                success, message = self.db.toplu_tedarikci_ekle_guncelle(data_to_process)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                success, message = self.db.toplu_stok_ekle_guncelle(
                    analysis_results.get('all_processed_data', []), 
                    analysis_results.get('selected_update_fields_from_ui', [])
                )
            else:
                success = False
                message = f"Bilinmeyen veri tipi: {veri_tipi}"
            
            final_success = success
            final_message = message

        except Exception as e:
            final_success = False
            final_message = f"Veritabanı yazma sırasında kritik hata: {e}\n{traceback.format_exc()}"
            from pencereler import logging
            logging.error(final_message) # Hata loglama eklendi
        
        finally:
            self.app.after(0, self.bekleme_penceresi_gercek_islem.kapat)
            
            if final_success:
                self.app.after(0, lambda: messagebox.showinfo("Başarılı", f"Toplu {veri_tipi} işlemi tamamlandı:\n{final_message}", parent=self.app))
                self.app.after(0, lambda: self.app.set_status(f"Toplu {veri_tipi} işlemi tamamlandı: {final_message}"))
                self.app.after(0, self._refresh_related_lists, veri_tipi) # lambda kaldırıldı
                self.app.after(0, self.destroy)
            else:
                self.app.after(0, lambda: messagebox.showerror("Hata", f"Toplu {veri_tipi} işlemi başarısız oldu:\n{final_message}", parent=self.app))
                self.app.after(0, lambda: self.app.set_status(f"Toplu {veri_tipi} işlemi başarısız oldu: {final_message}"))

    def _refresh_related_lists(self, veri_tipi):
        if veri_tipi == "Müşteri" and hasattr(self.app, 'musteri_yonetimi_sayfasi') and hasattr(self.app.musteri_yonetimi_sayfasi, 'musteri_listesini_yenile'):
            self.app.musteri_yonetimi_sayfasi.musteri_listesini_yenile()
        elif veri_tipi == "Tedarikçi" and hasattr(self.app, 'tedarikci_yonetimi_sayfasi') and hasattr(self.app.tedarikci_yonetimi_sayfasi, 'tedarikci_listesini_yenile'):
            self.app.tedarikci_yonetimi_sayfasi.tedarikci_listesini_yenile()
        elif veri_tipi == "Stok/Ürün Ekle/Güncelle" and hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, 'stok_listesini_yenile'):
            self.app.stok_yonetimi_sayfasi.stok_listesini_yenile()
        if hasattr(self.app, 'ana_sayfa') and hasattr(self.app.ana_sayfa, 'guncelle_ozet_bilgiler'):
            self.app.ana_sayfa.guncelle_ozet_bilgiler()

class TopluVeriOnizlemePenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, veri_tipi, analysis_results, callback_on_confirm):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.veri_tipi = veri_tipi
        self.analysis_results = analysis_results # Analiz sonuçlarını sakla
        self.callback_on_confirm = callback_on_confirm # Onay sonrası çağrılacak callback

        self.title(f"Toplu {veri_tipi} Önizleme")
        self.geometry("1000x700") # Daha büyük önizleme penceresi
        self.transient(parent_app)
        self.grab_set()
        self.resizable(True, True)

        ttk.Label(self, text=f"Toplu {veri_tipi} İşlemi Önizlemesi", font=("Segoe UI", 16, "bold")).pack(pady=10)

        # Özet Bilgiler Çerçevesi
        summary_frame = ttk.LabelFrame(self, text="İşlem Özeti", padding="10")
        summary_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.new_items_label = ttk.Label(summary_frame, text=f"Yeni Eklenecek: {self.analysis_results['new_count']} kayıt", font=("Segoe UI", 10, "bold"))
        self.new_items_label.pack(side=tk.LEFT, padx=10)
        self.updated_items_label = ttk.Label(summary_frame, text=f"Güncellenecek: {self.analysis_results['updated_count']} kayıt", font=("Segoe UI", 10, "bold"))
        self.updated_items_label.pack(side=tk.LEFT, padx=10)
        self.errors_label = ttk.Label(summary_frame, text=f"Hatalı Satır: {self.analysis_results['error_count']} kayıt", font=("Segoe UI", 10, "bold"), foreground="red")
        self.errors_label.pack(side=tk.LEFT, padx=10)

        # Notebook ile sekmeli görünüm (Yeni, Güncellenen, Hatalı)
        self.notebook_onizleme = ttk.Notebook(self)
        self.notebook_onizleme.pack(expand=True, fill=tk.BOTH, padx=10, pady=5)

        # 1. Yeni Eklenecekler Sekmesi
        if self.analysis_results['new_items']:
            new_frame = ttk.Frame(self.notebook_onizleme, padding="10")
            self.notebook_onizleme.add(new_frame, text="🟢 Yeni Eklenecekler")
            self._create_treeview_tab(new_frame, self.analysis_results['new_items'], "new")

        # 2. Güncellenecekler Sekmesi
        if self.analysis_results['updated_items']:
            updated_frame = ttk.Frame(self.notebook_onizleme, padding="10")
            self.notebook_onizleme.add(updated_frame, text="🟡 Güncellenecekler")
            self._create_treeview_tab(updated_frame, self.analysis_results['updated_items'], "updated")

        # 3. Hatalı Satırlar Sekmesi
        if self.analysis_results['errors_details']:
            errors_frame = ttk.Frame(self.notebook_onizleme, padding="10")
            self.notebook_onizleme.add(errors_frame, text="🔴 Hatalı Satırlar")
            self._create_treeview_tab(errors_frame, self.analysis_results['errors_details'], "errors")
            
            # Eğer hatalar varsa, varsayılan olarak hatalar sekmesini aç
            self.notebook_onizleme.select(errors_frame)


        # Butonlar
        button_frame = ttk.Frame(self, padding="10")
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)

        self.onayla_button = ttk.Button(button_frame, text="Onayla ve İşlemi Başlat", command=self._onayla_islemi_baslat, style="Accent.TButton")
        self.onayla_button.pack(side=tk.LEFT, padx=10)
        
        self.iptal_button = ttk.Button(button_frame, text="İptal", command=self.destroy)
        self.iptal_button.pack(side=tk.RIGHT, padx=10)

        # Eğer hiç işlem yoksa (sadece boş dosya veya sadece hatalar), onay butonunu pasif yap
        if self.analysis_results['new_count'] == 0 and self.analysis_results['updated_count'] == 0:
            self.onayla_button.config(state=tk.DISABLED)
            ttk.Label(button_frame, text="Hiçbir kayıt eklenmeyecek veya güncellenmeyecek.", foreground="orange").pack(side=tk.LEFT, padx=5)

    def _create_treeview_tab(self, parent_frame, data_list, tab_type):
        """Her bir sekme için Treeview oluşturur ve verileri doldurur."""
        # --- Stok/Ürün sütun indekslerini burada tanımla ---
        # Bu indeksler hem analiz hem de önizleme için tutarlı olmalı.
        COL_URUN_KODU = 0
        COL_URUN_ADI = 1
        COL_STOK_MIKTARI = 2
        COL_ALIS_FIYATI_KDV_DAHIL = 3
        COL_SATIS_FIYATI_KDV_DAHIL = 4
        COL_KDV_ORANI = 5
        COL_MIN_STOK_SEVIYESI = 6
        COL_KATEGORI_ADI = 7
        COL_MARKA_ADI = 8
        COL_URUN_GRUBU_ADI = 9
        COL_URUN_BIRIMI_ADI = 10
        COL_ULKE_ADI = 11
        COL_URUN_DETAYI = 12
        COL_URUN_RESMI_YOLU = 13

        if self.veri_tipi in ["Müşteri", "Tedarikçi"]:
            cols = ("Kod", "Ad", "Telefon", "Adres", "Vergi Dairesi", "Vergi No", "Durum")
            col_widths = {"Kod": 100, "Ad": 150, "Telefon": 100, "Adres": 200, "Vergi Dairesi": 120, "Vergi No": 100, "Durum": 150}
        elif self.veri_tipi == "Stok/Ürün Ekle/Güncelle":
            cols = ("Ürün Kodu", "Ürün Adı", "Miktar", "Alış Fyt (KDV Dahil)", "Satış Fyt (KDV Dahil)", "KDV %", "Min. Stok", "Kategori", "Marka", "Ürün Grubu", "Ürün Birimi", "Menşe", "Ürün Detayı", "Resim Yolu", "Durum")
            col_widths = {
                "Ürün Kodu": 80, "Ürün Adı": 120, "Miktar": 60, 
                "Alış Fyt (KDV Dahil)": 100, "Satış Fyt (KDV Dahil)": 100, 
                "KDV %": 60, "Min. Stok": 70, "Kategori": 80, "Marka": 80, 
                "Ürün Grubu": 80, "Ürün Birimi": 80, "Menşe": 80, 
                "Ürün Detayı": 100, "Resim Yolu": 100, "Durum": 150
            }
        else:
            cols = ("Veri 1", "Veri 2", "Durum")
            col_widths = {"Veri 1": 100, "Veri 2": 100, "Durum": 300}

        tree = ttk.Treeview(parent_frame, columns=cols, show='headings', selectmode="none")

        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=col_widths.get(col, 80), anchor=tk.W)

        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(parent_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(expand=True, fill=tk.BOTH)

        for item in data_list:
            if tab_type == "new" or tab_type == "updated":
                row_data_excel = list(item[0])
                status_message = item[1]

                if self.veri_tipi == "Stok/Ürün Ekle/Güncelle":
                    # row_data_excel'in yeterli uzunlukta olduğundan emin olun
                    extended_row = row_data_excel + [None] * (COL_URUN_RESMI_YOLU + 1 - len(row_data_excel))

                    row_for_tree = [
                        extended_row[COL_URUN_KODU] if extended_row[COL_URUN_KODU] is not None else "",
                        extended_row[COL_URUN_ADI] if extended_row[COL_URUN_ADI] is not None else "",
                        f"{self.db.safe_float(extended_row[COL_STOK_MIKTARI]):.2f}".rstrip('0').rstrip('.') if extended_row[COL_STOK_MIKTARI] is not None else "",
                        self.db._format_currency(self.db.safe_float(extended_row[COL_ALIS_FIYATI_KDV_DAHIL])) if extended_row[COL_ALIS_FIYATI_KDV_DAHIL] is not None else "",
                        self.db._format_currency(self.db.safe_float(extended_row[COL_SATIS_FIYATI_KDV_DAHIL])) if extended_row[COL_SATIS_FIYATI_KDV_DAHIL] is not None else "",
                        f"{self.db.safe_float(extended_row[COL_KDV_ORANI]):.0f}%" if extended_row[COL_KDV_ORANI] is not None else "",
                        f"{self.db.safe_float(extended_row[COL_MIN_STOK_SEVIYESI]):.2f}".rstrip('0').rstrip('.') if extended_row[COL_MIN_STOK_SEVIYESI] is not None else "",
                        extended_row[COL_KATEGORI_ADI] if extended_row[COL_KATEGORI_ADI] is not None else "",
                        extended_row[COL_MARKA_ADI] if extended_row[COL_MARKA_ADI] is not None else "",
                        extended_row[COL_URUN_GRUBU_ADI] if extended_row[COL_URUN_GRUBU_ADI] is not None else "",
                        extended_row[COL_URUN_BIRIMI_ADI] if extended_row[COL_URUN_BIRIMI_ADI] is not None else "",
                        extended_row[COL_ULKE_ADI] if extended_row[COL_ULKE_ADI] is not None else "",
                        extended_row[COL_URUN_DETAYI] if extended_row[COL_URUN_DETAYI] is not None else "",
                        extended_row[COL_URUN_RESMI_YOLU] if extended_row[COL_URUN_RESMI_YOLU] is not None else "",
                        status_message
                    ]
                elif self.veri_tipi in ["Müşteri", "Tedarikçi"]:
                    extended_row = row_data_excel + [None] * (5 - len(row_data_excel))
                    row_for_tree = [
                        extended_row[0], extended_row[1], extended_row[2], 
                        extended_row[3], extended_row[4], extended_row[5], 
                        status_message
                    ]
                else:
                    row_for_tree = list(row_data_excel) + [status_message]

                tree.insert("", tk.END, values=row_for_tree)

            elif tab_type == "errors":
                row_data_for_error = list(item[0])
                error_message = item[1]

                if self.veri_tipi == "Stok/Ürün Ekle/Güncelle":
                    # extended_row'u COL_URUN_RESMI_YOLU'na göre ayarlayın
                    extended_row = row_data_for_error + [None] * (COL_URUN_RESMI_YOLU + 1 - len(row_data_for_error))
                    display_cols_for_error = [
                        extended_row[COL_URUN_KODU] if extended_row[COL_URUN_KODU] is not None else "",
                        extended_row[COL_URUN_ADI] if extended_row[COL_URUN_ADI] is not None else "",
                        f"{self.db.safe_float(extended_row[COL_STOK_MIKTARI]):.2f}".rstrip('0').rstrip('.') if extended_row[COL_STOK_MIKTARI] is not None else "",
                        self.db._format_currency(self.db.safe_float(extended_row[COL_ALIS_FIYATI_KDV_DAHIL])) if extended_row[COL_ALIS_FIYATI_KDV_DAHIL] is not None else "",
                        self.db._format_currency(self.db.safe_float(extended_row[COL_SATIS_FIYATI_KDV_DAHIL])) if extended_row[COL_SATIS_FIYATI_KDV_DAHIL] is not None else "",
                        f"{self.db.safe_float(extended_row[COL_KDV_ORANI]):.0f}%" if extended_row[COL_KDV_ORANI] is not None else "",
                        f"{self.db.safe_float(extended_row[COL_MIN_STOK_SEVIYESI]):.2f}".rstrip('0').rstrip('.') if extended_row[COL_MIN_STOK_SEVIYESI] is not None else "",
                        extended_row[COL_KATEGORI_ADI] if extended_row[COL_KATEGORI_ADI] is not None else "",
                        extended_row[COL_MARKA_ADI] if extended_row[COL_MARKA_ADI] is not None else "",
                        extended_row[COL_URUN_GRUBU_ADI] if extended_row[COL_URUN_GRUBU_ADI] is not None else "",
                        extended_row[COL_URUN_BIRIMI_ADI] if extended_row[COL_URUN_BIRIMI_ADI] is not None else "",
                        extended_row[COL_ULKE_ADI] if extended_row[COL_ULKE_ADI] is not None else "",
                        extended_row[COL_URUN_DETAYI] if extended_row[COL_URUN_DETAYI] is not None else "",
                        extended_row[COL_URUN_RESMI_YOLU] if extended_row[COL_URUN_RESMI_YOLU] is not None else "",
                        error_message
                    ]
                elif self.veri_tipi in ["Müşteri", "Tedarikçi"]:
                    display_cols_for_error = [
                        row_data_for_error[0] if len(row_data_for_error) > 0 and row_data_for_error[0] is not None else "",
                        row_data_for_error[1] if len(row_data_for_error) > 1 and row_data_for_error[1] is not None else "",
                        row_data_for_error[2] if len(row_data_for_error) > 2 and row_data_for_error[2] is not None else "",
                        "", "", "", # Boşlukları doldur
                        error_message
                    ]
                else:
                    display_cols_for_error = list(row_data_for_error) + [error_message]

                tree.insert("", tk.END, values=display_cols_for_error, tags=('error_row',))
                tree.tag_configure('error_row', background='#FFCCCC', foreground='red')

    def _onayla_islemi_baslat(self):
        self.destroy() # Önizleme penceresini kapat

        self.bekleme_penceresi_gercek_islem = BeklemePenceresi(
            self.app, 
            message=f"Toplu {self.veri_tipi} veritabanına yazılıyor, lütfen bekleyiniz..."
        )

        # Gerçek veritabanı yazma işlemini ayrı bir thread'de başlat
        threading.Thread(target=lambda: self._gercek_yazma_islemini_yap_threaded(
            self.veri_tipi, 
            self.analysis_results # Önizlemede analiz edilmiş sonuçları gönder
        )).start()

    def _gercek_yazma_islemini_yap_threaded(self, veri_tipi, analysis_results):
        final_success = True
        final_message = ""
        temp_db_manager = None

        try:
            # Geçici bir veritabanı bağlantısı aç
            temp_db_manager = self.db.__class__(db_name=self.db.db_name)
            if not hasattr(temp_db_manager, 'app') or temp_db_manager.app is None:
                temp_db_manager.app = self.app # Geçici manager'a app referansını ver

            # Başlangıçta gerekli varsayılan kayıtları kontrol et/oluştur
            temp_db_manager._ensure_genel_tedarikci()
            temp_db_manager._ensure_perakende_musteri()
            temp_db_manager._ensure_default_kasa()
            temp_db_manager._ensure_default_urun_birimi()
            temp_db_manager._ensure_default_ulke()

            # <<< DÜZELTME BURADA >>>
            # Doğru veri listesini ('all_processed_data') ve doğru metot adlarını kullanıyoruz.
            data_to_process = analysis_results.get('all_processed_data', [])

            if veri_tipi == "Müşteri":
                success, message = temp_db_manager.toplu_musteri_ekle_guncelle(data_to_process)
            elif veri_tipi == "Tedarikçi":
                success, message = temp_db_manager.toplu_tedarikci_ekle_guncelle(data_to_process)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                success, message = temp_db_manager.toplu_stok_ekle_guncelle(
                    analysis_results.get('all_processed_data', []), 
                    analysis_results.get('selected_update_fields_from_ui', [])
                )
            else:
                success = False
                message = f"Bilinmeyen veri tipi: {veri_tipi}"
            
            final_success = success
            final_message = message

        except Exception as e:
            final_success = False
            final_message = f"Veritabanı yazma sırasında kritik hata: {e}\n{traceback.format_exc()}"
            from arayuz import logging
            logging.error(final_message)
        
        finally:
            if temp_db_manager and temp_db_manager.conn:
                try:
                    temp_db_manager.conn.close()
                except Exception as close_e:
                    print(f"UYARI: Thread bağlantısı kapatılırken hata: {close_e}")

            # Bekleme penceresini kapat
            self.app.after(0, self.bekleme_penceresi_gercek_islem.kapat)
            
            if final_success:
                self.app.after(0, lambda: messagebox.showinfo("Başarılı", f"Toplu {veri_tipi} işlemi tamamlandı:\n{final_message}", parent=self.app))
                self.app.after(0, lambda: self.app.set_status(f"Toplu {veri_tipi} işlemi tamamlandı: {final_message}"))
                self.app.after(0, self._refresh_related_lists, veri_tipi)
                self.app.after(0, self.destroy)
            else:
                self.app.after(0, lambda: messagebox.showerror("Hata", f"Toplu {veri_tipi} işlemi başarısız oldu:\n{final_message}", parent=self.app))
                self.app.after(0, lambda: self.app.set_status(f"Toplu {veri_tipi} işlemi başarısız oldu: {final_message}"))

    def _refresh_related_lists(self, veri_tipi):
        if veri_tipi == "Müşteri" and hasattr(self.app, 'musteri_yonetimi_sayfasi') and hasattr(self.app.musteri_yonetimi_sayfasi, 'musteri_listesini_yenile'):
            self.app.musteri_yonetimi_sayfasi.musteri_listesini_yenile()
        elif veri_tipi == "Tedarikçi" and hasattr(self.app, 'tedarikci_yonetimi_sayfasi') and hasattr(self.app.tedarikci_yonetimi_sayfasi, 'tedarikci_listesini_yenile'):
            self.app.tedarikci_yonetimi_sayfasi.tedarikci_listesini_yenile()
        elif veri_tipi == "Stok/Ürün Ekle/Güncelle" and hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, 'stok_listesini_yenile'):
            self.app.stok_yonetimi_sayfasi.stok_listesini_yenile()
        if hasattr(self.app, 'ana_sayfa') and hasattr(self.app.ana_sayfa, 'guncelle_ozet_bilgiler'):
            self.app.ana_sayfa.guncelle_ozet_bilgiler()

class AciklamaDetayPenceresi(tk.Toplevel): # <<< Bu sınıf doğru hizada (BeklemePenceresi ile aynı)
    def __init__(self, parent, title="Detaylı Bilgi", message_text=""):
        super().__init__(parent)
        self.title(title)
        self.geometry("600x400")
        self.transient(parent)
        self.grab_set()
        self.resizable(False, False)

        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

        self.text_widget = tk.Text(self, wrap=tk.WORD, font=("Segoe UI", 10), padx=10, pady=10)
        self.text_widget.pack(expand=True, fill=tk.BOTH)
        self.text_widget.insert(tk.END, message_text)
        self.text_widget.config(state=tk.DISABLED)

        vsb = ttk.Scrollbar(self.text_widget, orient="vertical", command=self.text_widget.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_widget.config(yscrollcommand=vsb.set)

        ttk.Button(self, text="Kapat", command=self.destroy).pack(pady=10)

class CariSecimPenceresi(tk.Toplevel):
    def __init__(self, parent_window, db_manager, fatura_tipi, callback_func): # parent_app -> parent_window olarak adlandırdım
        super().__init__(parent_window) 
        self.app = parent_window.app # parent_window'un içindeki app referansını al
        self.db = db_manager
        self.fatura_tipi = fatura_tipi
        self.callback_func = callback_func

        self.title("Cari Seçimi")
        self.geometry("600x450")
        self.transient(parent_window) 
        self.grab_set()
        self.resizable(False, False)

        self.tum_cariler_cache_data = [] # Data tuple'larını saklar: (id, kod, ad, ...)
        self.cari_map_display_to_id = {} # Sadece pop-up içinde kullanılacak, ana formunkinden farklı

        baslik_text = "Müşteri Seçimi" if self.fatura_tipi == 'SATIŞ' else "Tedarikçi Seçimi"
        ttk.Label(self, text=baslik_text, font=("Segoe UI", 14, "bold")).pack(pady=10)

        # Arama Çerçevesi
        search_frame = ttk.Frame(self, padding="10")
        search_frame.pack(fill=tk.X)

        ttk.Label(search_frame, text="Ara (Ad/Kod):").pack(side=tk.LEFT, padx=(0,5))
        self.search_entry = ttk.Entry(search_frame, width=40)
        self.search_entry.pack(side=tk.LEFT, padx=(0,10), fill=tk.X, expand=True)
        self.search_entry.bind("<KeyRelease>", self._filtre_liste)

        # Cari Listesi Treeview
        tree_frame = ttk.Frame(self, padding="10")
        tree_frame.pack(expand=True, fill=tk.BOTH)

        self.cari_tree = ttk.Treeview(tree_frame, columns=("Cari Adı", "Kodu"), show="headings", selectmode="browse")
        self.cari_tree.heading("Cari Adı", text="Cari Adı")
        self.cari_tree.heading("Kodu", text="Kodu")
        self.cari_tree.column("Cari Adı", width=300, stretch=tk.YES)
        self.cari_tree.column("Kodu", width=100, stretch=tk.NO)
        self.cari_tree.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.cari_tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.cari_tree.configure(yscrollcommand=vsb.set)
        
        self.cari_tree.bind("<Double-1>", self._sec) # Çift tıklama ile seçim

        # Butonlar
        button_frame = ttk.Frame(self, padding="10")
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)

        ttk.Button(button_frame, text="Seç", command=self._sec, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.RIGHT, padx=5)

        # Başlangıç yüklemesi
        self._yukle_carileri()
        self.search_entry.focus() # Arama kutusuna odaklan
    
    def _yukle_carileri(self):
        """Tüm carileri (müşteri veya tedarikçi) veritabanından çeker ve listeler."""
        self.tum_cariler_cache_data = [] # Tüm cari data tuple'larını saklar
        self.cari_map_display_to_id = {} # Display text -> ID map
        
        if self.fatura_tipi == 'SATIŞ':
            cariler_db = self.db.musteri_listesi_al(perakende_haric=False) # Perakende müşteriyi de al
        else: # ALIŞ
            cariler_db = self.db.tedarikci_listesi_al()
        
        for c in cariler_db: # c: (id, kod, ad, ...)
            display_text = f"{c[2]} (Kod: {c[1]})" # Ad (Kod)
            self.cari_map_display_to_id[display_text] = str(c[0]) # ID'yi string olarak sakla
            self.tum_cariler_cache_data.append(c) # Tüm cari data tuple'larını cache'le
        
        self._filtre_liste() # Tüm listeyi göster (boş arama terimiyle)

        # Varsayılan seçimi yap
        default_id_str = None
        if self.fatura_tipi == 'SATIŞ' and self.db.perakende_musteri_id is not None:
            default_id_str = str(self.db.perakende_musteri_id)
        elif self.fatura_tipi == 'ALIŞ' and self.db.genel_tedarikci_id is not None:
            default_id_str = str(self.db.genel_tedarikci_id)
        
        if default_id_str:
            # Treeview'de bu varsayılan öğeyi bul ve seçili yap
            for item_id in self.cari_tree.get_children():
                if item_id == default_id_str: # item_id zaten string ID
                    self.cari_tree.selection_set(item_id)
                    self.cari_tree.focus(item_id)
                    self.cari_tree.see(item_id) # Öğeyi görünür alana kaydır
                    break

    def _filtre_liste(self, event=None):
        arama_terimi = self.search_entry.get().lower().strip()

        for i in self.cari_tree.get_children():
            self.cari_tree.delete(i)

        for cari_tuple in self.tum_cariler_cache_data:
            cari_id, cari_kodu, cari_ad, _, _, _, _ = cari_tuple

            if (not arama_terimi or
                (cari_ad and arama_terimi in cari_ad.lower()) or  # Cari adı boş değilse ve içeriyorsa
                (cari_kodu and arama_terimi in cari_kodu.lower())  # Cari kodu boş değilse ve içeriyorsa
               ):
                self.cari_tree.insert("", tk.END, iid=str(cari_id), values=(cari_ad, cari_kodu))


    def _sec(self, event=None):
        """Seçili cariyi onaylar ve callback fonksiyonunu çağırır."""
        selected_item_iid = self.cari_tree.focus()
        if not selected_item_iid:
            messagebox.showwarning("Seçim Yok", "Lütfen bir cari seçin.", parent=self)
            return

        selected_cari_id = int(selected_item_iid) # iid zaten ID'dir
        item_values = self.cari_tree.item(selected_item_iid, 'values')
        selected_cari_display_text = item_values[0] # Cari Adı sütunu
        
        self.callback_func(selected_cari_id, selected_cari_display_text) # Callback'i çağır
        self.destroy() # Pencereyi kapat        

class TedarikciSecimDialog(tk.Toplevel):
    def __init__(self, parent_window, db_manager, callback_func): # parent_app -> parent_window olarak adlandırdım
        super().__init__(parent_window) 
        self.app = parent_window.app # parent_window'un içindeki app referansını al
        self.db = db_manager
        self.callback_func = callback_func

        self.title("Tedarikçi Seçimi")
        self.geometry("600x400")
        self.transient(parent_window) 
        self.grab_set()
        self.resizable(False, False)

        self.tum_tedarikciler_cache = [] # Data tuple'larını saklar: (id, kod, ad, ...)

        ttk.Label(self, text="Tedarikçi Seçimi", font=("Segoe UI", 14, "bold")).pack(pady=10)

        # Arama Çerçevesi
        search_frame = ttk.Frame(self, padding="10")
        search_frame.pack(fill=tk.X)

        ttk.Label(search_frame, text="Ara (Ad/Kod):").pack(side=tk.LEFT, padx=(0,5))
        self.search_entry = ttk.Entry(search_frame, width=40)
        self.search_entry.pack(side=tk.LEFT, padx=(0,10), fill=tk.X, expand=True)
        self.search_entry.bind("<KeyRelease>", self._filtre_liste)

        # Tedarikçi Listesi Treeview
        tree_frame = ttk.Frame(self, padding="10")
        tree_frame.pack(expand=True, fill=tk.BOTH)

        self.tedarikci_tree = ttk.Treeview(tree_frame, columns=("Tedarikçi Adı", "Kodu"), show="headings", selectmode="browse")
        self.tedarikci_tree.heading("Tedarikçi Adı", text="Tedarikçi Adı")
        self.tedarikci_tree.heading("Kodu", text="Kodu")
        self.tedarikci_tree.column("Tedarikçi Adı", width=300, stretch=tk.YES)
        self.tedarikci_tree.column("Kodu", width=100, stretch=tk.NO)
        self.tedarikci_tree.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tedarikci_tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tedarikci_tree.configure(yscrollcommand=vsb.set)
        
        self.tedarikci_tree.bind("<Double-1>", self._sec) # Çift tıklama ile seçim

        # Butonlar
        button_frame = ttk.Frame(self, padding="10")
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)

        ttk.Button(button_frame, text="Seç", command=self._sec, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.RIGHT, padx=5)

        # Başlangıç yüklemesi
        self._yukle_tedarikcileri()
        self.search_entry.focus() # Arama kutusuna odaklan
    
    def _yukle_tedarikcileri(self):
        """Tüm tedarikçileri veritabanından çeker ve listeler."""
        self.tum_tedarikciler_cache = self.db.tedarikci_listesi_al() # Tüm tedarikçileri al
                
        self._filtre_liste() # Tüm listeyi göster (boş arama terimiyle)

    def _filtre_liste(self, event=None):
        """Arama kutusuna yazıldıkça tedarikçi listesini filtreler."""
        arama_terimi = self.search_entry.get().lower().strip()
        
        for i in self.tedarikci_tree.get_children():
            self.tedarikci_tree.delete(i) # Treeview'i temizle
        
        for tedarikci_tuple in self.tum_tedarikciler_cache:
            tedarikci_id, tedarikci_kodu, tedarikci_ad, _, _, _, _ = tedarikci_tuple
            
            # Arama terimi tedarikçi adında VEYA tedarikçi kodunda geçiyorsa
            if (not arama_terimi or
                (tedarikci_ad and arama_terimi in tedarikci_ad.lower()) or
                (tedarikci_kodu and arama_terimi in tedarikci_kodu.lower())
               ):
                self.tedarikci_tree.insert("", tk.END, iid=str(tedarikci_id), values=(tedarikci_ad, tedarikci_kodu))

    def _sec(self, event=None):
        """Seçili tedarikçiyi onaylar ve callback fonksiyonunu çağırır."""
        selected_item_iid = self.tedarikci_tree.focus()
        if not selected_item_iid:
            messagebox.showwarning("Seçim Yok", "Lütfen bir tedarikçi seçin.", parent=self)
            return

        selected_tedarikci_id = int(selected_item_iid) # iid zaten ID'dir
        item_values = self.tedarikci_tree.item(selected_item_iid, 'values')
        selected_tedarikci_ad = item_values[0] # Tedarikçi Adı sütunu
        
        self.callback_func(selected_tedarikci_id, selected_tedarikci_ad) # Callback'i çağır
        self.destroy() # Pencereyi kapat        

class BeklemePenceresi(tk.Toplevel):
    def __init__(self, parent, title="İşlem Devam Ediyor...", message="Lütfen bekleyiniz..."):
        super().__init__(parent)
        self.title(title)
        self.geometry("300x120")
        self.transient(parent)
        self.grab_set()
        self.resizable(False, False)

        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

        ttk.Label(self, text=message, font=("Segoe UI", 10, "bold"), wraplength=280, justify=tk.CENTER).pack(pady=20)
        
        self.progressbar = ttk.Progressbar(self, mode="indeterminate", length=200)
        self.progressbar.pack(pady=10)
        self.progressbar.start()

        self.protocol("WM_DELETE_WINDOW", self._do_nothing)

    def _do_nothing(self):
        pass

    def kapat(self):
        self.progressbar.stop()
        self.destroy()
        
class GelirGiderSiniflandirmaYonetimiPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, yenile_callback):
        super().__init__(parent_app)
        self.db = db_manager
        self.parent_app = parent_app
        self.yenile_callback = yenile_callback # Ana pencereyi yenilemek için

        self.title("Gelir/Gider Sınıflandırma Yönetimi")
        self.geometry("600x450")
        self.transient(parent_app)
        self.grab_set()
        self.resizable(False, False)

        # Notebook (Sekmeler) oluştur
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        # Gelir Sınıflandırmaları Sekmesi
        self.gelir_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.gelir_frame, text="Gelir Sınıflandırmaları")
        self._setup_siniflandirma_sekmesi(self.gelir_frame, "GELİR")

        # Gider Sınıflandırmaları Sekmesi
        self.gider_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.gider_frame, text="Gider Sınıflandırmaları")
        self._setup_siniflandirma_sekmesi(self.gider_frame, "GİDER")

        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.update_idletasks()
        self.geometry(f"{self.winfo_reqwidth()}x{self.winfo_reqheight()}")

        # Sağ tık menüsü
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Güncelle", command=self._siniflandirma_guncelle)
        self.context_menu.add_command(label="Sil", command=self._siniflandirma_sil)

    def _setup_siniflandirma_sekmesi(self, parent_frame, tip):
        print(f"DEBUG: _setup_siniflandirma_sekmesi çağrıldı. Tip: {tip}") # <-- YENİ DEBUG
        # Arama ve Ekleme alanı
        top_frame = ttk.Frame(parent_frame, padding="10")
        top_frame.pack(fill=tk.X)

        ttk.Label(top_frame, text="Yeni Sınıflandırma Adı:").pack(side=tk.LEFT, padx=5)
        entry = ttk.Entry(top_frame, width=30)
        entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        add_button = ttk.Button(top_frame, text="Ekle", command=lambda: self._siniflandirma_ekle(tip, entry.get().strip(), entry))
        add_button.pack(side=tk.LEFT, padx=5)

        # Treeview alanı
        tree_frame = ttk.Frame(parent_frame)
        tree_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=5)

        tree = ttk.Treeview(tree_frame, columns=("ID", "Sınıflandırma Adı"), show="headings")
        tree.heading("ID", text="ID", anchor=tk.W)
        tree.heading("Sınıflandırma Adı", text="Sınıflandırma Adı", anchor=tk.W)
        tree.column("ID", width=50, stretch=tk.NO)
        tree.column("Sınıflandırma Adı", width=250, stretch=tk.YES)
        tree.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=scrollbar.set)

        # Treeview'i kaydet
        if tip == "GELİR":
            self.gelir_tree = tree
        else:
            self.gider_tree = tree
        
        # Sağ tık menüsünü treeview'e bağla
        print(f"DEBUG: Sağ tık menüsü '{tip}' treeview'ine bağlanıyor.") # <-- YENİ DEBUG
        tree.bind("<Button-3>", self._on_treeview_right_click) # <-- Mouse sağ tıklama olayı
        # DİKKAT: <ButtonRelease-3> yerine <Button-3> kullanmak bazı durumlarda daha güvenilir olabilir.
        # Eğer hala çalışmazsa <ButtonRelease-3> deneyin.

        self._load_siniflandirmalar(tip)

    def _load_siniflandirmalar(self, tip):
        tree = self.gelir_tree if tip == "GELİR" else self.gider_tree
        
        for item in tree.get_children():
            tree.delete(item)
        
        siniflandirmalar = []
        if tip == "GELİR":
            siniflandirmalar = self.db.gelir_siniflandirma_listele()
        else:
            siniflandirmalar = self.db.gider_siniflandirma_listele()
        
        for s_id, s_adi in siniflandirmalar:
            tree.insert("", tk.END, values=(s_id, s_adi), iid=s_id) # iid olarak ID'yi kullan

    def _siniflandirma_ekle(self, tip, siniflandirma_adi, entry_widget):
        if not siniflandirma_adi:
            messagebox.showwarning("Uyarı", "Sınıflandırma adı boş olamaz.", parent=self)
            return

        success, message = (False, "")
        if tip == "GELİR":
            success, message = self.db.gelir_siniflandirma_ekle(siniflandirma_adi)
        else:
            success, message = self.db.gider_siniflandirma_ekle(siniflandirma_adi)

        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            entry_widget.delete(0, tk.END) # Giriş alanını temizle
            self._load_siniflandirmalar(tip) # Listeyi yenile
            if self.yenile_callback:
                self.yenile_callback() # Ana pencereyi yenile
        else:
            messagebox.showerror("Hata", message, parent=self)

    # DÜZELTME BAŞLANGICI: Sağ tık menüsü metotları
    def _on_treeview_right_click(self, event):
        """Treeview'e sağ tıklandığında menüyü gösterir."""
        print(f"DEBUG: _on_treeview_right_click çağrıldı. Event: x={event.x}, y={event.y}") # <-- YENİ DEBUG
        current_tab_text = self.notebook.tab(self.notebook.select(), "text")
        
        if "Gelir Sınıflandırmaları" in current_tab_text:
            tree = self.gelir_tree
        else:
            tree = self.gider_tree

        # Seçili öğeyi al
        item_id = tree.identify_row(event.y)
        print(f"DEBUG: identify_row ile bulunan item_id: {item_id}") # <-- YENİ DEBUG

        if item_id:
            tree.selection_set(item_id) # Öğeyi seçili hale getir
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
                print("DEBUG: Sağ tık menüsü başarıyla açıldı.") 
            finally:
                self.context_menu.grab_release()
        else:
            print("DEBUG: Geçerli bir Treeview öğesi üzerinde sağ tıklanmadı. Menü açılmıyor.") # <-- YENİ DEBUG
            # Boş alana tıklandığında menüyü gizle/kapat (eğer açıksa)
            if hasattr(self, 'context_menu') and self.context_menu.winfo_exists():
                self.context_menu.unpost() # Menüyü kapat

    def _siniflandirma_guncelle(self):
        """Seçili sınıflandırmayı güncellemek için düzenleme penceresini açar."""
        current_tab_text = self.notebook.tab(self.notebook.select(), "text")
        
        if "Gelir Sınıflandırmaları" in current_tab_text:
            tree = self.gelir_tree
            tip = "GELİR"
        else:
            tree = self.gider_tree
            tip = "GİDER"

        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen güncellemek istediğiniz sınıflandırmayı seçin.", parent=self)
            return

        # Seçili öğenin ID'sini al (iid olarak saklandı)
        siniflandirma_id = selected_item[0] 
        values = tree.item(siniflandirma_id, 'values')
        siniflandirma_adi = values[1] # Sınıflandırma Adı ikinci sütunda

        siniflandirma_info = {'id': siniflandirma_id, 'siniflandirma_adi': siniflandirma_adi}
        
        SiniflandirmaDuzenlePenceresi(self, self.db, tip, siniflandirma_info, 
                                      lambda: self._load_siniflandirmalar(tip)) # Yenile callback

    def _siniflandirma_sil(self):
        """Seçili sınıflandırmayı siler."""
        current_tab_text = self.notebook.tab(self.notebook.select(), "text")
        
        if "Gelir Sınıflandırmaları" in current_tab_text:
            tree = self.gelir_tree
            tip = "GELİR"
        else:
            tree = self.gider_tree
            tip = "GİDER"

        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen silmek istediğiniz sınıflandırmayı seçin.", parent=self)
            return

        siniflandirma_id = selected_item[0] # iid olarak saklandı

        cevap = messagebox.askyesno("Onay", f"Seçili sınıflandırmayı silmek istediğinizden emin misiniz?", parent=self)
        if cevap:
            success, message = (False, "")
            if tip == "GELİR":
                success, message = self.db.gelir_siniflandirma_sil(siniflandirma_id)
            else:
                success, message = self.db.gider_siniflandirma_sil(siniflandirma_id)

            if success:
                messagebox.showinfo("Başarılı", message, parent=self)
                self._load_siniflandirmalar(tip) # Listeyi yenile
                if self.yenile_callback:
                    self.yenile_callback() # Ana pencereyi yenile
            else:
                messagebox.showerror("Hata", message, parent=self)

class BirimDuzenlePenceresi(tk.Toplevel):
    def __init__(self, parent_window, db_manager, birim_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.birim_id = birim_info['id']
        self.mevcut_birim_adi = birim_info['birim_adi']
        self.yenile_callback = yenile_callback

        self.title(f"Birim Düzenle: {self.mevcut_birim_adi}")
        self.geometry("350x200")
        self.transient(parent_window)
        self.grab_set()
        self.resizable(False, False)

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        ttk.Label(main_frame, text="Birim Adı:").grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
        self.birim_adi_entry = ttk.Entry(main_frame, width=30)
        self.birim_adi_entry.grid(row=0, column=1, padx=5, pady=10, sticky=tk.EW)
        self.birim_adi_entry.insert(0, self.mevcut_birim_adi)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=10, sticky=tk.E)

        ttk.Button(button_frame, text="Kaydet", command=self._kaydet, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.LEFT)

    def _kaydet(self):
        yeni_birim_adi = self.birim_adi_entry.get().strip()
        if not yeni_birim_adi:
            messagebox.showwarning("Uyarı", "Birim adı boş olamaz.", parent=self)
            return

        success, message = self.db.urun_birimi_guncelle(self.birim_id, yeni_birim_adi)

        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.yenile_callback() # Ana listedeki birimleri yenile
            self.destroy() # Pencereyi kapat
        else:
            messagebox.showerror("Hata", message, parent=self)

class GrupDuzenlePenceresi(tk.Toplevel):
    def __init__(self, parent_window, db_manager, grup_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.grup_id = grup_info['id']
        self.mevcut_grup_adi = grup_info['grup_adi']
        self.yenile_callback = yenile_callback

        self.title(f"Grup Düzenle: {self.mevcut_grup_adi}")
        self.geometry("350x200")
        self.transient(parent_window)
        self.grab_set()
        self.resizable(False, False)

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        ttk.Label(main_frame, text="Grup Adı:").grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
        self.grup_adi_entry = ttk.Entry(main_frame, width=30)
        self.grup_adi_entry.grid(row=0, column=1, padx=5, pady=10, sticky=tk.EW)
        self.grup_adi_entry.insert(0, self.mevcut_grup_adi)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=10, sticky=tk.E)

        ttk.Button(button_frame, text="Kaydet", command=self._kaydet, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.LEFT)

    def _kaydet(self):
        yeni_grup_adi = self.grup_adi_entry.get().strip()
        if not yeni_grup_adi:
            messagebox.showwarning("Uyarı", "Grup adı boş olamaz.", parent=self)
            return

        success, message = self.db.urun_grubu_guncelle(self.grup_id, yeni_grup_adi)

        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.yenile_callback()
            self.destroy()
        else:
            messagebox.showerror("Hata", message, parent=self)

# UlkeDuzenlePenceresi sınıfı
class UlkeDuzenlePenceresi(tk.Toplevel):
    def __init__(self, parent_window, db_manager, ulke_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.ulke_id = ulke_info['id']
        self.mevcut_ulke_adi = ulke_info['ulke_adi']
        self.yenile_callback = yenile_callback

        self.title(f"Ülke Düzenle: {self.mevcut_ulke_adi}")
        self.geometry("350x200")
        self.transient(parent_window)
        self.grab_set()
        self.resizable(False, False)

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        ttk.Label(main_frame, text="Ülke Adı:").grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
        self.ulke_adi_entry = ttk.Entry(main_frame, width=30)
        self.ulke_adi_entry.grid(row=0, column=1, padx=5, pady=10, sticky=tk.EW)
        self.ulke_adi_entry.insert(0, self.mevcut_ulke_adi)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=10, sticky=tk.E)

        ttk.Button(button_frame, text="Kaydet", command=self._kaydet, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.LEFT)

    def _kaydet(self):
        yeni_ulke_adi = self.ulke_adi_entry.get().strip()
        if not yeni_ulke_adi:
            messagebox.showwarning("Uyarı", "Ülke adı boş olamaz.", parent=self)
            return

        success, message = self.db.ulke_guncelle(self.ulke_id, yeni_ulke_adi)

        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.yenile_callback()
            self.destroy()
        else:
            messagebox.showerror("Hata", message, parent=self)

class SiniflandirmaDuzenlePenceresi(tk.Toplevel):
    def __init__(self, parent_window, db_manager, tip, siniflandirma_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.tip = tip # "GELİR" veya "GİDER"
        self.siniflandirma_id = siniflandirma_info['id']
        self.mevcut_siniflandirma_adi = siniflandirma_info['siniflandirma_adi']
        self.yenile_callback = yenile_callback

        self.title(f"{tip.capitalize()} Sınıflandırma Düzenle: {self.mevcut_siniflandirma_adi}")
        self.geometry("400x220") # Boyutu biraz büyütüldü
        self.transient(parent_window)
        self.grab_set()
        self.resizable(False, False)

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)

        ttk.Label(main_frame, text="Sınıflandırma Adı:").grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
        self.siniflandirma_adi_entry = ttk.Entry(main_frame, width=35) # Genişlik artırıldı
        self.siniflandirma_adi_entry.grid(row=0, column=1, padx=5, pady=10, sticky=tk.EW)
        self.siniflandirma_adi_entry.insert(0, self.mevcut_siniflandirma_adi)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=10, sticky=tk.E)

        ttk.Button(button_frame, text="Kaydet", command=self._kaydet, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=self.destroy).pack(side=tk.LEFT)

    def _kaydet(self):
        yeni_siniflandirma_adi = self.siniflandirma_adi_entry.get().strip()
        if not yeni_siniflandirma_adi:
            messagebox.showwarning("Uyarı", "Sınıflandırma adı boş olamaz.", parent=self)
            return

        success, message = (False, "")
        if self.tip == "GELİR":
            success, message = self.db.gelir_siniflandirma_guncelle(self.siniflandirma_id, yeni_siniflandirma_adi)
        else: # GİDER
            success, message = self.db.gider_siniflandirma_guncelle(self.siniflandirma_id, yeni_siniflandirma_adi)

        if success:
            messagebox.showinfo("Başarılı", message, parent=self)
            self.yenile_callback() # Ana listedeki sınıflandırmaları yenile
            self.destroy() # Pencereyi kapat
        else:
            messagebox.showerror("Hata", message, parent=self)